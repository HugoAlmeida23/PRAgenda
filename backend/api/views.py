import logging
from rest_framework import viewsets, generics, status
from rest_framework.request import Request
from datetime import datetime, timedelta 
from django.utils import timezone
import json
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.exceptions import PermissionDenied, ValidationError 
from django.contrib.auth.models import User
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from django.db.models import Q, Prefetch, F, Count, Sum, Avg, Exists, OuterRef, Subquery
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.core.cache import cache
from .models import (Organization, Client, TaskCategory, Task, TimeEntry, Expense, 
                    ClientProfitability, Profile, AutoTimeTracking, WorkflowStep,
                    WorkflowDefinition, TaskApproval, WorkflowNotification,NotificationTemplate, 
                    WorkflowHistory,NotificationSettings, NotificationDigest, FiscalObligationDefinition, FiscalSystemSettings,GeneratedReport)

from .serializers import (ClientSerializer, TaskCategorySerializer, TaskSerializer,
                         TimeEntrySerializer, ExpenseSerializer, ClientProfitabilitySerializer,
                         ProfileSerializer, AutoTimeTrackingSerializer, OrganizationSerializer,
                         UserSerializer, WorkflowDefinitionSerializer, 
                         WorkflowStepSerializer, TaskApprovalSerializer, WorkflowNotificationSerializer,
                         WorkflowHistorySerializer, NotificationSettingsSerializer,NotificationTemplateSerializer,
                         NotificationDigestSerializer,FiscalGenerationRequestSerializer,FiscalStatsSerializer,
                         FiscalObligationTestSerializer, FiscalObligationDefinitionSerializer, FiscalSystemSettingsSerializer,GeneratedReportSerializer)
from django.db import models
from .services.notification_service import NotificationService 
from .services.notifications_metrics import NotificationMetricsService
from .services.notification_escalation import NotificationEscalationService
from .services.notifications_reports import NotificationReportsService
from .services.notification_template_service import NotificationTemplateService
from .services.notification_digest_service import NotificationDigestService
from .services.fiscal_obligation_service import FiscalObligationGenerator
from .services.fiscal_notification_service import FiscalNotificationService 
from .services.ai_advisor_service import AIAdvisorService
from .utils import CustomJSONEncoder, update_client_profitability, log_organization_action  
from django.db.models import ExpressionWrapper, fields
from .services.workflow_service import WorkflowService
from .tasks import update_profitability_for_single_organization_task # Import the new Celery task
from dateutil.relativedelta import relativedelta
from django.db.models.expressions import RawSQL # Make sure this is imported
from .permissions import IsOrgAdmin, CanManageClients, CanManageTimeEntry
from .tasks import generate_report_task # <-- Import the new task
from rest_framework.parsers import MultiPartParser
from .serializers import SAFTFileSerializer
from .models import SAFTFile
from .tasks import process_saft_file_task
from .models import InvoiceBatch, ScannedInvoice
from .serializers import InvoiceBatchSerializer, ScannedInvoiceSerializer
from .tasks import process_invoice_file_task
from django.db import transaction
from .models import OrganizationActionLog
from .serializers import OrganizationActionLogSerializer
from rest_framework import permissions

logger = logging.getLogger(__name__)

class InvoiceBatchViewSet(viewsets.ModelViewSet):
    serializer_class = InvoiceBatchSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        
        # Handle superuser
        if user.is_superuser:
            return InvoiceBatch.objects.all().select_related(
                'uploaded_by', 'organization'
            ).prefetch_related('invoices__generated_tasks').order_by('-created_at')
        
        try:
            profile = user.profile
            if not profile or not profile.organization:
                return InvoiceBatch.objects.none()
            
            return InvoiceBatch.objects.filter(
                organization=profile.organization
            ).select_related(
                'uploaded_by', 'organization'
            ).prefetch_related('invoices__generated_tasks').order_by('-created_at')
            
        except AttributeError:
            # User has no profile
            return InvoiceBatch.objects.none()

    def create(self, request, *args, **kwargs):
        files = request.FILES.getlist('files')
        if not files:
            return Response({
                'error': 'Nenhum ficheiro enviado.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check user profile and organization
        try:
            profile = request.user.profile
            if not profile or not profile.organization:
                return Response({
                    'error': 'Utilizador não está numa organização.'
                }, status=status.HTTP_400_BAD_REQUEST)
        except AttributeError:
            return Response({
                'error': 'Perfil de utilizador não encontrado.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create the batch
        try:
            batch = InvoiceBatch.objects.create(
                organization=profile.organization,
                uploaded_by=request.user,
                description=request.data.get('description', f'Lote de {len(files)} faturas')
            )

            # Create invoice records and dispatch processing tasks
            created_invoices = []
            for file in files:
                invoice = ScannedInvoice.objects.create(
                    batch=batch, 
                    original_file=file,
                    original_filename=file.name
                )
                created_invoices.append(invoice)
                
                # Dispatch the processing task
                try:
                    from .tasks import process_invoice_file_task
                    process_invoice_file_task.delay(str(invoice.id))
                except Exception as e:
                    # If Celery is not available or task fails to dispatch
                    invoice.status = 'ERROR'
                    invoice.processing_log = f'Erro ao iniciar processamento: {str(e)}'
                    invoice.save()
            
            # Log organization action
            log_organization_action(
                request,
                action_type='CREATE_INVOICE_BATCH',
                action_description=f"Lote de faturas criado: {batch.description} (ID: {batch.id}) com {len(created_invoices)} faturas.",
                related_object=batch
            )
            # Return the created batch with all related data
            serializer = self.get_serializer(batch)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({
                'error': f'Erro ao criar lote: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def create_batch_tasks(self, request, pk=None):
        """
        Cria tarefas para múltiplas faturas de um lote.
        Suporta dois modos:
        1. Criação simples (sem dados de tarefa específicos)
        2. Criação com dados de tarefa personalizados
        """
        batch = self.get_object()
        
        try:
            profile = request.user.profile
            if not profile.organization:
                return Response({
                    'error': 'Utilizador não está numa organização.'
                }, status=status.HTTP_400_BAD_REQUEST)
        except AttributeError:
            return Response({
                'error': 'Perfil de utilizador não encontrado.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verificar permissões para criar tarefas
        if not (profile.is_org_admin or profile.can_create_tasks):
            return Response({
                'error': 'Sem permissão para criar tarefas'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Obter dados da tarefa (se fornecidos) ou usar defaults
        task_data = request.data.get('task_data', {})
        invoice_ids_to_process = request.data.get('invoices_to_process', [])
        
        # Filtrar faturas elegíveis
        eligible_invoices_qs = batch.invoices.filter(
            status='COMPLETED'
        ).exclude(
            generated_tasks__isnull=False  # Excluir faturas que já têm tarefas
        )
        
        if invoice_ids_to_process:
            eligible_invoices_qs = eligible_invoices_qs.filter(id__in=invoice_ids_to_process)
        
        eligible_invoices = list(eligible_invoices_qs.select_related('batch__organization'))
        
        if not eligible_invoices:
            return Response({
                'success': False,
                'message': 'Nenhuma fatura elegível encontrada para criação de tarefas.',
                'tasks_created': 0,
                'tasks_failed': 0
            }, status=status.HTTP_200_OK)
        
        # Obter clientes disponíveis para matching
        available_clients = Client.objects.filter(
            organization=profile.organization,
            is_active=True
        ).select_related('organization')
        
        # Preparar dados base da tarefa
        base_task_data = {
            'status': task_data.get('status', 'pending'),
            'priority': task_data.get('priority', 3),
            'deadline': task_data.get('deadline'),
            'estimated_time_minutes': task_data.get('estimated_time_minutes'),
            'category_id': task_data.get('category'),
            'workflow_id': task_data.get('workflow'),
            'assigned_to_id': task_data.get('assigned_to'),
            'description_template': task_data.get('description', ''),
            'title_template': task_data.get('title', 'Lançar Fatura: {invoice_ref}')
        }
        
        # Se um cliente específico foi selecionado, usar esse
        default_client = None
        if task_data.get('client'):
            try:
                default_client = available_clients.get(id=task_data['client'])
                if not profile.can_access_client(default_client):
                    return Response({
                        'error': f'Sem acesso ao cliente selecionado: {default_client.name}'
                    }, status=status.HTTP_403_FORBIDDEN)
            except Client.DoesNotExist:
                return Response({
                    'error': 'Cliente selecionado não encontrado'
                }, status=status.HTTP_404_NOT_FOUND)
        
        tasks_created = 0
        tasks_failed = 0
        errors = []
        created_task_ids = []
        
        with transaction.atomic():
            try:
                for invoice in eligible_invoices:
                    try:
                        # Determinar cliente para esta fatura
                        target_client = default_client
                        
                        # Se não há cliente padrão, tentar encontrar por NIF
                        if not target_client and invoice.nif_acquirer:
                            target_client = available_clients.filter(
                                nif=invoice.nif_acquirer
                            ).first()
                        
                        # Se ainda não há cliente, usar o primeiro disponível
                        if not target_client:
                            target_client = available_clients.first()
                        
                        if not target_client:
                            errors.append(f"Fatura {invoice.original_filename}: Nenhum cliente disponível")
                            tasks_failed += 1
                            continue
                        
                        # Verificar se utilizador tem acesso ao cliente
                        if not profile.can_access_client(target_client):
                            errors.append(f"Fatura {invoice.original_filename}: Sem acesso ao cliente {target_client.name}")
                            tasks_failed += 1
                            continue
                        
                        # Gerar título e descrição personalizados
                        invoice_ref = invoice.atcud or invoice.original_filename
                        task_title = base_task_data['title_template'].format(
                            invoice_ref=invoice_ref,
                            client_name=target_client.name,
                            batch_description=batch.description or f"Lote {batch.id}"
                        )
                        
                        # Criar descrição detalhada
                        if base_task_data['description_template']:
                            task_description = base_task_data['description_template']
                        else:
                            task_description = f"Lançamento contabilístico da fatura de {invoice.nif_emitter or 'N/A'}.\n"
                            task_description += f"Data: {invoice.doc_date or 'N/A'}\n"
                            task_description += f"Valor Total: {invoice.gross_total or '0.00'}€\n"
                            task_description += f"IVA: {invoice.vat_amount or '0.00'}€\n"
                            if batch.description:
                                task_description += f"Processada em lote: {batch.description}"
                        
                        # Criar tarefa
                        task_kwargs = {
                            'title': task_title,
                            'description': task_description,
                            'client': target_client,
                            'created_by': request.user,
                            'source_scanned_invoice': invoice,
                            'status': base_task_data['status'],
                            'priority': base_task_data['priority']
                        }
                        
                        # Adicionar campos opcionais se fornecidos
                        if base_task_data['deadline']:
                            task_kwargs['deadline'] = base_task_data['deadline']
                        if base_task_data['estimated_time_minutes']:
                            task_kwargs['estimated_time_minutes'] = base_task_data['estimated_time_minutes']
                        if base_task_data['category_id']:
                            try:
                                category = TaskCategory.objects.get(id=base_task_data['category_id'])
                                task_kwargs['category'] = category
                            except TaskCategory.DoesNotExist:
                                pass
                        if base_task_data['assigned_to_id']:
                            try:
                                assigned_user = User.objects.get(id=base_task_data['assigned_to_id'])
                                # Verificar se o usuário pertence à mesma organização
                                assigned_profile = assigned_user.profile
                                if assigned_profile.organization == profile.organization:
                                    task_kwargs['assigned_to'] = assigned_user
                            except (User.DoesNotExist, Profile.DoesNotExist):
                                pass
                        
                        task = Task.objects.create(**task_kwargs)
                        
                        # Configurar workflow se especificado
                        if base_task_data['workflow_id']:
                            try:
                                workflow = WorkflowDefinition.objects.get(
                                    id=base_task_data['workflow_id'],
                                    is_active=True
                                )
                                first_step = workflow.steps.order_by('order').first()
                                if first_step:
                                    task.workflow = workflow
                                    task.current_workflow_step = first_step
                                    task.save(update_fields=['workflow', 'current_workflow_step'])
                                    
                                    # Criar histórico de workflow
                                    WorkflowHistory.objects.create(
                                        task=task,
                                        from_step=None,
                                        to_step=first_step,
                                        changed_by=request.user,
                                        action='workflow_assigned',
                                        comment=f"Workflow '{workflow.name}' atribuído na criação em lote."
                                    )
                            except WorkflowDefinition.DoesNotExist:
                                pass
                        
                        tasks_created += 1
                        created_task_ids.append(str(task.id))
                        
                        # Log da criação
                        logger.info(f"Tarefa {task.id} criada para fatura {invoice.id} no batch {batch.id}")
                        
                    except Exception as e:
                        logger.error(f"Erro ao criar tarefa para fatura {invoice.id}: {str(e)}")
                        tasks_failed += 1
                        errors.append(f"Fatura {invoice.original_filename}: {str(e)}")
                
                # Enviar notificações para tarefas criadas
                if tasks_created > 0 and base_task_data.get('assigned_to_id'):
                    try:
                        assigned_user = User.objects.get(id=base_task_data['assigned_to_id'])
                        if assigned_user.id != request.user.id:  # Não notificar o criador
                            NotificationService.create_notification(
                                user=assigned_user,
                                task=None,  # Múltiplas tarefas
                                notification_type='task_assigned_to_you',
                                title=f"{tasks_created} Novas Tarefas Atribuídas (Lote)",
                                message=f"Foram-lhe atribuídas {tasks_created} tarefas do lote '{batch.description or 'Sem descrição'}'. "
                                    f"Criadas por: {request.user.username}.",
                                created_by=request.user,
                                metadata={
                                    'batch_id': str(batch.id),
                                    'task_ids': created_task_ids,
                                    'invoice_count': len(eligible_invoices)
                                }
                            )
                    except (User.DoesNotExist, Exception) as e:
                        logger.warning(f"Erro ao enviar notificação de lote: {e}")
                
                response_data = {
                    'success': True,
                    'message': f'{tasks_created} tarefas criadas com sucesso.',
                    'tasks_created': tasks_created,
                    'tasks_failed': tasks_failed,
                    'batch_id': str(batch.id),
                    'created_task_ids': created_task_ids
                }
                
                if errors:
                    response_data['errors'] = errors
                    response_data['message'] += f' {tasks_failed} falharam.'
                # Log organization action for batch task creation
                log_organization_action(
                    request,
                    action_type='BATCH_TASK_CREATION',
                    action_description=f"{tasks_created} tarefas criadas em lote para batch {batch.id}.",
                    related_object=batch
                )
                return Response(response_data, status=status.HTTP_201_CREATED)
                
            except Exception as e:
                logger.error(f"Erro na criação em lote de tarefas para batch {batch.id}: {str(e)}")
                return Response({
                    'error': f'Erro interno ao criar tarefas: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def batch_status(self, request, pk=None):
        """
        Retorna informações detalhadas sobre o estado do lote e suas tarefas.
        """
        batch = self.get_object()
        
        try:
            profile = request.user.profile
            if not profile.organization:
                return Response({'error': 'Utilizador não está numa organização.'}, 
                            status=status.HTTP_400_BAD_REQUEST)
        except AttributeError:
            return Response({'error': 'Perfil de utilizador não encontrado.'}, 
                        status=status.HTTP_400_BAD_REQUEST)
        
        # Estatísticas das faturas
        invoices = batch.invoices.all()
        completed_invoices = invoices.filter(status='COMPLETED')
        
        # Estatísticas das tarefas
        invoices_with_tasks = completed_invoices.filter(generated_tasks__isnull=False).distinct()
        invoices_without_tasks = completed_invoices.exclude(generated_tasks__isnull=False)
        
        # Obter tarefas relacionadas
        related_tasks = Task.objects.filter(
            source_scanned_invoice__batch=batch
        ).select_related('assigned_to', 'client', 'category')
        
        task_stats = {
            'pending': related_tasks.filter(status='pending').count(),
            'in_progress': related_tasks.filter(status='in_progress').count(),
            'completed': related_tasks.filter(status='completed').count(),
            'cancelled': related_tasks.filter(status='cancelled').count(),
        }
        
        return Response({
            'batch_id': str(batch.id),
            'batch_description': batch.description,
            'created_at': batch.created_at,
            'invoice_stats': {
                'total': invoices.count(),
                'completed': completed_invoices.count(),
                'with_tasks': invoices_with_tasks.count(),
                'without_tasks': invoices_without_tasks.count(),
                'processing': invoices.filter(status='PROCESSING').count(),
                'pending': invoices.filter(status='PENDING').count(),
                'error': invoices.filter(status='ERROR').count(),
            },
            'task_stats': task_stats,
            'can_create_more_tasks': invoices_without_tasks.count() > 0,
            'ready_for_batch_creation': invoices_without_tasks.count() > 1,
        })

    @action(detail=True, methods=['get'])
    def generate_excel(self, request, pk=None):
        """
        Generates an Excel file with details of all invoices in a specific batch.
        """
        batch = self.get_object()
        invoices = batch.invoices.all().order_by('created_at')

        # Create an in-memory workbook
        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Faturas Processadas'

        # Define headers
        headers = [
            'ID Fatura', 'Ficheiro Original', 'Status', 'NIF Emissor', 'NIF Adquirente',
            'Data Documento', 'ATCUD', 'Total Bruto', 'Total IVA', 'Base Tributável'
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Populate rows with invoice data
        for row_num, invoice in enumerate(invoices, 2):
            worksheet.cell(row=row_num, column=1, value=str(invoice.id))
            worksheet.cell(row=row_num, column=2, value=invoice.original_file.name.split('/')[-1])
            worksheet.cell(row=row_num, column=3, value=invoice.get_status_display())
            worksheet.cell(row=row_num, column=4, value=invoice.nif_emitter)
            worksheet.cell(row=row_num, column=5, value=invoice.nif_acquirer)
            worksheet.cell(row=row_num, column=6, value=invoice.doc_date)
            worksheet.cell(row=row_num, column=7, value=invoice.atcud)
            worksheet.cell(row=row_num, column=8, value=float(invoice.gross_total or 0)).number_format = '#,##0.00€'
            worksheet.cell(row=row_num, column=9, value=float(invoice.vat_amount or 0)).number_format = '#,##0.00€'
            worksheet.cell(row=row_num, column=10, value=float(invoice.taxable_amount or 0)).number_format = '#,##0.00€'

        # Auto-fit column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Save the workbook to the buffer
        workbook.save(buffer)
        buffer.seek(0)

        # Create the HTTP response
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="faturas_lote_{batch.id}.xlsx"'

        return response

class ScannedInvoiceViewSet(viewsets.ModelViewSet):
    serializer_class = ScannedInvoiceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        
        if user.is_superuser:
            return ScannedInvoice.objects.all().select_related('batch__organization')
        
        try:
            profile = user.profile
            if not profile or not profile.organization:
                return ScannedInvoice.objects.none()
            
            return ScannedInvoice.objects.filter(
                batch__organization=profile.organization
            ).select_related('batch__organization')
            
        except AttributeError:
            return ScannedInvoice.objects.none()

    def partial_update(self, request, *args, **kwargs):
        """Handle user edits to invoice data."""
        invoice = self.get_object()
        
        # Update edited_data with the new values
        edited_data = request.data.copy()
        
        # Remove non-editable fields
        for field in ['id', 'batch', 'created_at', 'original_file', 'status']:
            edited_data.pop(field, None)
        
        invoice.edited_data = edited_data
        invoice.is_reviewed = True
        invoice.status = 'COMPLETED'  # Mark as completed after review
        invoice.save()
        # Log organization action
        log_organization_action(
            request,
            action_type='REVIEW_INVOICE',
            action_description=f"Fatura revista: {invoice.original_filename} (ID: {invoice.id})",
            related_object=invoice
        )
        serializer = self.get_serializer(invoice)
        return Response(serializer.data)
    
# In views.py - Update SAFTFileViewSet
class SAFTFileViewSet(viewsets.ModelViewSet):
    serializer_class = SAFTFileSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def get_queryset(self):
        try:
            profile = self.request.user.profile
            if not profile or not profile.organization:
                logger.warning(f"User {self.request.user.username} has no profile or organization")
                return SAFTFile.objects.none()
            
            queryset = SAFTFile.objects.filter(
                organization=profile.organization
            ).select_related('uploaded_by', 'organization').order_by('-uploaded_at')
            
            logger.info(f"SAFT queryset for user {self.request.user.username}: {queryset.count()} files")
            return queryset
            
        except Profile.DoesNotExist:
            logger.warning(f"User {self.request.user.username} without a profile tried to access SAFT files")
            return SAFTFile.objects.none()
        except Exception as e:
            logger.error(f"Error in SAFTFileViewSet.get_queryset: {e}")
            return SAFTFile.objects.none()

    def create(self, request, *args, **kwargs):
        """
        Custom create method to provide detailed error responses.
        """
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            # perform_create logic is now here:
            profile = request.user.profile
            file_obj = request.data.get('file')
            
            saft_instance = serializer.save(
                organization=profile.organization,
                uploaded_by=request.user,
                original_filename=file_obj.name
            )
            
            # Dispatch the background task
            process_saft_file_task.delay(saft_instance.id)

            # Log organization action
            log_organization_action(
                request,
                action_type='UPLOAD_SAFT_FILE',
                action_description=f"SAFT carregado: {saft_instance.original_filename} (ID: {saft_instance.id})",
                related_object=saft_instance
            )
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

        except Profile.DoesNotExist:
            logger.error(f"User {request.user.username} has no profile, cannot upload SAFT.")
            return Response({"detail": "O seu perfil de utilizador não foi encontrado. Contacte o suporte."}, status=status.HTTP_403_FORBIDDEN)
        
        except ValidationError as e:
            logger.warning(f"SAFT upload validation error for user {request.user.username}: {e.detail}")
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            logger.error(f"Unhandled exception during SAFT upload for user {request.user.username}: {e}", exc_info=True)
            return Response({"detail": "Ocorreu um erro inesperado no servidor ao processar o seu pedido."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def details(self, request, pk=None):
        """
        Retrieve detailed information about a processed SAFT file
        """
        try:
            saft_file = self.get_object()
            
            # Calculate file size dynamically
            file_size_kb = None
            if saft_file.file:
                try:
                    file_size_kb = saft_file.file.size // 1024  # Convert bytes to KB
                except (OSError, AttributeError):
                    file_size_kb = None
            
            # Basic file information
            details = {
                'id': str(saft_file.id),
                'filename': saft_file.file.name.split('/')[-1] if saft_file.file else 'N/A',
                'original_filename': saft_file.original_filename,
                'status': saft_file.status,
                'processing_log': saft_file.processing_log,
                'uploaded_at': saft_file.uploaded_at,
                'processed_at': saft_file.processed_at,
                'file_size_kb': file_size_kb,
                
                # Processed data
                'fiscal_year': saft_file.fiscal_year,
                'start_date': saft_file.start_date,
                'end_date': saft_file.end_date,
                'company_name': saft_file.company_name,
                'company_tax_id': saft_file.company_tax_id,
                'summary_data': saft_file.summary_data or {},
            }
            
            return Response(details)
            
        except Exception as e:
            logger.error(f"Error in SAFTFile details view: {e}", exc_info=True)
            return Response(
                {'error': f'Erro ao obter detalhes: {str(e)}'}, 
                status=500
            )
    
class GeneratedReportViewSet(viewsets.ReadOnlyModelViewSet): # ReadOnly por agora, criação será via "gerar relatório"
    serializer_class = GeneratedReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        logger.debug(f"[GeneratedReportViewSet] get_queryset for user: {user.username}")
        try:
            profile = user.profile
            if not profile.organization:
                logger.warn(f"[GeneratedReportViewSet] User {user.username} has no organization.")
                return GeneratedReport.objects.none()

            # Permissão para ver relatórios: ser admin da org OU ter permissão 'can_view_analytics' OU 'can_export_reports'
            can_view_reports = (
                profile.is_org_admin or 
                getattr(profile, 'can_view_analytics', False) or # getattr com default False
                getattr(profile, 'can_export_reports', False)
            )

            if can_view_reports:
                logger.debug(f"[GeneratedReportViewSet] User {user.username} can view reports for org {profile.organization.name}.")
                base_qs = GeneratedReport.objects.filter(organization=profile.organization)
            else:
                # Se não tem permissão geral, talvez possa ver relatórios que ele mesmo gerou?
                # (Pode ser uma regra futura, por agora, se não pode ver análises/exportar, não vê a lista)
                logger.warn(f"[GeneratedReportViewSet] User {user.username} does not have general report viewing permissions.")
                return GeneratedReport.objects.none()
        
        except Profile.DoesNotExist:
            if user.is_superuser:
                logger.debug(f"[GeneratedReportViewSet] Superuser {user.username} accessing all reports.")
                base_qs = GeneratedReport.objects.all()
            else:
                logger.warn(f"[GeneratedReportViewSet] Profile does not exist for non-superuser {user.username}.")
                return GeneratedReport.objects.none()

        # Aplicar filtros de query params
        filters_applied = {}
        report_type_param = self.request.query_params.get('report_type')
        if report_type_param:
            base_qs = base_qs.filter(report_type=report_type_param)
            filters_applied['report_type'] = report_type_param

        date_from_param = self.request.query_params.get('created_at__gte')
        if date_from_param:
            base_qs = base_qs.filter(created_at__gte=date_from_param)
            filters_applied['created_at__gte'] = date_from_param

        date_to_param = self.request.query_params.get('created_at__lte')
        if date_to_param:
            # Ajustar para incluir o dia todo
            try:
                date_to_obj = datetime.strptime(date_to_param, '%Y-%m-%d').date()
                datetime_to = datetime.combine(date_to_obj, datetime.max.time())
                base_qs = base_qs.filter(created_at__lte=datetime_to)
                filters_applied['created_at__lte'] = date_to_param
            except ValueError:
                logger.warn(f"Invalid date_to_param: {date_to_param}")


        # Adicionar pesquisa se o frontend enviar `search_term` e você quiser filtrar no backend
        search_term_param = self.request.query_params.get('search') # Supondo que o frontend use 'search'
        if search_term_param:
             base_qs = base_qs.filter(
                 Q(name__icontains=search_term_param) |
                 Q(description__icontains=search_term_param) |
                 Q(parameters__icontains=search_term_param) # Pesquisa dentro do JSON de parâmetros
             )
             filters_applied['search'] = search_term_param
        
        logger.debug(f"[GeneratedReportViewSet] Filters applied: {filters_applied}, Count after filters: {base_qs.count()}")
        return base_qs.select_related('organization', 'generated_by').order_by('-created_at')

class FiscalObligationDefinitionViewSet(viewsets.ModelViewSet):
    serializer_class = FiscalObligationDefinitionSerializer
    permission_classes = [IsAuthenticated] # Permissions will be checked in methods

    def get_queryset(self):
        user = self.request.user
        base_queryset = FiscalObligationDefinition.objects.select_related(
            'organization', 'default_task_category', 'default_workflow'
        )
        if user.is_superuser:
            return base_queryset.all()
        try:
            profile = user.profile
            if profile.organization:
                return base_queryset.filter(
                    Q(organization__isnull=True) | Q(organization=profile.organization)
                )
            else:
                return base_queryset.filter(organization__isnull=True)
        except Profile.DoesNotExist:
            return base_queryset.filter(organization__isnull=True)

    def _check_permission(self, request, instance=None):
        """Helper to check permissions for write operations."""
        user = request.user
        if user.is_superuser:
            return True
        try:
            profile = user.profile
            if not profile.organization or not profile.is_org_admin:
                raise PermissionDenied("Apenas administradores de organização podem gerenciar definições.")
            
            if instance: # This is an update/delete on an existing object
                if instance.organization and instance.organization != profile.organization:
                    raise PermissionDenied("Não pode modificar definições de outra organização.")
                if instance.organization is None: # Global definition
                    raise PermissionDenied("Apenas superusuários podem modificar definições globais.")
            
            # For create, we also need to check the org in the payload
            if not instance and 'organization' in request.data:
                org_id = request.data['organization']
                if org_id and str(org_id) != str(profile.organization.id):
                    raise PermissionDenied("Não pode criar definições para outra organização.")

            return True
        except Profile.DoesNotExist:
            return False

    def perform_create(self, serializer):
        self._check_permission(self.request)
        # If permission check passes, save with the user's organization
        organization = self.request.user.profile.organization
        serializer.save(organization=organization)
        # Log organization action
        log_organization_action(
            self.request,
            action_type='CREATE_FISCAL_OBLIGATION_DEFINITION',
            action_description=f"Definição de obrigação fiscal criada: {serializer.instance.name} (ID: {serializer.instance.id})",
            related_object=serializer.instance
        )

    def perform_update(self, serializer):
        self._check_permission(self.request, instance=serializer.instance)
        instance = serializer.save()
        # Log organization action
        log_organization_action(
            self.request,
            action_type='UPDATE_FISCAL_OBLIGATION_DEFINITION',
            action_description=f"Definição de obrigação fiscal atualizada: {instance.name} (ID: {instance.id})",
            related_object=instance
        )

    def perform_destroy(self, instance):
        self._check_permission(self.request, instance=instance)
        if Task.objects.filter(source_fiscal_obligation=instance).exists():
            raise ValidationError("Esta definição está em uso e não pode ser excluída. Considere desativá-la.")
        instance_id = instance.id
        instance_name = instance.name
        instance.delete()
        # Log organization action
        log_organization_action(
            self.request,
            action_type='DELETE_FISCAL_OBLIGATION_DEFINITION',
            action_description=f"Definição de obrigação fiscal excluída: {instance_name} (ID: {instance_id})",
            related_object=None
        )
                
class ProfileViewSet(viewsets.ModelViewSet):
    serializer_class = ProfileSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        # OPTIMIZATION:
        # - `select_related` for user and organization (ForeignKey).
        # - `prefetch_related` for visible_clients (ManyToManyField).
        # - `annotate` to calculate unread notifications count in the DB, preventing N+1 in the serializer.
        return Profile.objects.filter(user=user).select_related(
            'user', 
            'organization'
        ).prefetch_related(
            # Prefetch clients with their related data to make `visible_clients_info` efficient
            Prefetch('visible_clients', queryset=Client.objects.only('id', 'name'))
        ).annotate(
            unread_notifications_count=Count(
                'user__workflow_notifications', 
                filter=Q(
                    user__workflow_notifications__is_read=False, 
                    user__workflow_notifications__is_archived=False
                )
            )
        )
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        # Log organization action
        log_organization_action(
            self.request,
            action_type='CREATE_PROFILE',
            action_description=f"Perfil criado para usuário: {self.request.user.username}",
            related_object=self.request.user.profile if hasattr(self.request.user, 'profile') else None
        )

class AutoTimeTrackingViewSet(viewsets.ModelViewSet):
    serializer_class = AutoTimeTrackingSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return AutoTimeTracking.objects.select_related(
            'user'
        ).filter(user=self.request.user).order_by('-start_time')
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        # Log organization action
        log_organization_action(
            self.request,
            action_type='CREATE_AUTOTIMETRACKING',
            action_description=f"AutoTimeTracking criado para usuário: {self.request.user.username}",
            related_object=self.request.user
        )
        
class ClientViewSet(viewsets.ModelViewSet):
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated, CanManageClients]

    def get_queryset(self):
        """
        Uses the custom manager to get a secure base queryset and then
        applies filters from the request.
        """
        user = self.request.user
        base_queryset = Client.objects.for_user(user)
        
        # Apply optimizations
        optimized_queryset = base_queryset.select_related('organization', 'account_manager')

        # Apply filters from query parameters
        if self.request.query_params.get('is_active', '').lower() == 'true':
            optimized_queryset = optimized_queryset.filter(is_active=True)
            
        search_term = self.request.query_params.get('search')
        if search_term:
            optimized_queryset = optimized_queryset.filter(
                Q(name__icontains=search_term) |
                Q(nif__icontains=search_term) |
                Q(email__icontains=search_term)
            )

        return optimized_queryset.order_by('name')

    def perform_create(self, serializer):
        """
        Assigns the client to the user's organization upon creation.
        """
        try:
            profile = self.request.user.profile
            if not profile.organization:
                raise ValidationError("Usuário não pertence a nenhuma organização.")
            instance = serializer.save(organization=profile.organization)
            # Log the action
            log_organization_action(
                self.request,
                action_type='CREATE_CLIENT',
                action_description=f"Cliente criado: {instance.name} (NIF: {instance.nif})",
                related_object=instance
            )
            # Notify all org admins
            org_admins = Profile.objects.filter(organization=profile.organization, is_org_admin=True).select_related('user')
            for admin_profile in org_admins:
                NotificationService.create_notification(
                    user=admin_profile.user,
                    notification_type='client_created',
                    title=f"Novo cliente criado",
                    message=f"O cliente {instance.name} (NIF: {instance.nif}) foi criado.",
                    created_by=self.request.user
                )
        except Profile.DoesNotExist:
            raise PermissionDenied("Perfil de usuário não encontrado.")

    def update(self, request, *args, **kwargs):
        try:
            profile = Profile.objects.select_related('organization').get(user=request.user)
            client_instance = self.get_object() # Use client_instance to avoid confusion with serializer.client
            
            if not (profile.is_org_admin or profile.can_edit_clients):
                raise PermissionDenied("Você não tem permissão para editar clientes")
                
            if client_instance.organization != profile.organization:
                # This check might be too strict if an admin from Org A should never edit client from Org B, even if superuser.
                # Superuser bypass might be needed here if intended.
                raise PermissionDenied("Este cliente não pertence à sua organização")
                
            if not (profile.is_org_admin or profile.can_view_all_clients):
                if not profile.visible_clients.filter(id=client_instance.id).exists():
                    raise PermissionDenied("Você não tem acesso a este cliente para edição")
            response = super().update(request, *args, **kwargs)
            # Log organization action
            log_organization_action(
                request,
                action_type='UPDATE_CLIENT',
                action_description=f"Cliente atualizado: {client_instance.name} (NIF: {client_instance.nif})",
                related_object=client_instance
            )
            # Notify all org admins
            org_admins = Profile.objects.filter(organization=client_instance.organization, is_org_admin=True).select_related('user')
            for admin_profile in org_admins:
                NotificationService.create_notification(
                    user=admin_profile.user,
                    notification_type='client_updated',
                    title=f"Cliente atualizado",
                    message=f"O cliente {client_instance.name} (NIF: {client_instance.nif}) foi atualizado.",
                    created_by=request.user
                )
            return response
        except Profile.DoesNotExist:
            if self.request.user.is_superuser: # Allow superuser to update if no profile
                response = super().update(request, *args, **kwargs)
                # Log organization action
                log_organization_action(
                    request,
                    action_type='UPDATE_CLIENT',
                    action_description=f"Cliente atualizado (superuser): {self.get_object().name}",
                    related_object=self.get_object()
                )
                return response
            raise PermissionDenied("Perfil de usuário não encontrado")

    def destroy(self, request, *args, **kwargs):
        try:
            profile = Profile.objects.get(user=request.user)
            client_instance = self.get_object()

            if not (profile.is_org_admin or profile.can_delete_clients):
                raise PermissionDenied("Você não tem permissão para excluir clientes")
            
            # Ensure client belongs to user's org before deleting (unless superuser)
            if client_instance.organization != profile.organization and not request.user.is_superuser:
                 raise PermissionDenied("Não pode excluir clientes de outra organização.")

            response = super().destroy(request, *args, **kwargs)
            # Log organization action
            log_organization_action(
                request,
                action_type='DELETE_CLIENT',
                action_description=f"Cliente excluído: {client_instance.name} (NIF: {client_instance.nif})",
                related_object=client_instance
            )
            # Notify all org admins
            org_admins = Profile.objects.filter(organization=client_instance.organization, is_org_admin=True).select_related('user')
            for admin_profile in org_admins:
                NotificationService.create_notification(
                    user=admin_profile.user,
                    notification_type='client_deleted',
                    title=f"Cliente excluído",
                    message=f"O cliente {client_instance.name} (NIF: {client_instance.nif}) foi excluído.",
                    created_by=request.user
                )
            return response
        except Profile.DoesNotExist:
            if self.request.user.is_superuser:
                response = super().destroy(request, *args, **kwargs)
                # Log organization action
                log_organization_action(
                    request,
                    action_type='DELETE_CLIENT',
                    action_description=f"Cliente excluído (superuser): {self.get_object().name}",
                    related_object=self.get_object()
                )
                return response
            raise PermissionDenied("Perfil de usuário não encontrado")
    
    @action(detail=True, methods=['patch'])
    def toggle_status(self, request, pk=None):
        client = self.get_object()
        
        try:
            profile = Profile.objects.get(user=request.user)
            
            if not (profile.is_org_admin or profile.can_change_client_status):
                raise PermissionDenied("Você não tem permissão para alterar o status do cliente")

            if client.organization != profile.organization and not request.user.is_superuser:
                 raise PermissionDenied("Não pode alterar status de clientes de outra organização.")
                
            client.is_active = not client.is_active
            client.save()
            # Log organization action
            log_organization_action(
                request,
                action_type='TOGGLE_CLIENT_STATUS',
                action_description=f"Status do cliente alterado: {client.name} (NIF: {client.nif}) para {'Ativo' if client.is_active else 'Inativo'}",
                related_object=client
            )
            # Send notification to all org admins
            org_admins = Profile.objects.filter(organization=client.organization, is_org_admin=True).select_related('user')
            for admin_profile in org_admins:
                NotificationService.create_notification(
                    user=admin_profile.user,
                    notification_type='client_status_changed',
                    title=f"Status do cliente alterado",
                    message=f"O status do cliente {client.name} (NIF: {client.nif}) foi alterado para {'Ativo' if client.is_active else 'Inativo'}.",
                    created_by=request.user
                )
            serializer = self.get_serializer(client)
            return Response(serializer.data)
            
        except Profile.DoesNotExist:
            if self.request.user.is_superuser: # Superuser can toggle if client exists
                client.is_active = not client.is_active
                client.save()
                # Log organization action
                log_organization_action(
                    request,
                    action_type='TOGGLE_CLIENT_STATUS',
                    action_description=f"Status do cliente alterado (superuser): {client.name}",
                    related_object=client
                )
                serializer = self.get_serializer(client)
                return Response(serializer.data)
            raise PermissionDenied("Perfil de usuário não encontrado")
        
class TaskCategoryViewSet(viewsets.ModelViewSet):
    queryset = TaskCategory.objects.all() # Categories are typically global or org-specific.
    serializer_class = TaskCategorySerializer
    permission_classes = [IsAuthenticated]
    # Add permission checks for create/update/delete (e.g., only org admins)

class TaskViewSet(viewsets.ModelViewSet):
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        logger.debug(f"[TaskViewSet] get_queryset for user: {user.username}")

        # 1. Start with a secure base queryset using our new manager method.
        # This queryset ONLY contains tasks the user is allowed to see.
        base_queryset = Task.objects.for_user(user)

        # 2. Apply optional filters from the request's query parameters.
        # These filters are now applied ON TOP of the secure base set.
        query_params = self.request.query_params

        # --- Status filter ---
        status_param = query_params.get('status')
        if status_param:
            status_values = [s.strip() for s in status_param.split(',') if s.strip()]
            if status_values:
                base_queryset = base_queryset.filter(status__in=status_values)
        
        # --- Client filter ---
        client_id_param = query_params.get('client')
        if client_id_param:
            base_queryset = base_queryset.filter(client_id=client_id_param)

        # --- Priority filter ---
        priority_param = query_params.get('priority')
        if priority_param and priority_param.isdigit():
            base_queryset = base_queryset.filter(priority=int(priority_param))

        # --- AssignedTo filter ---
        assigned_to_id_param = query_params.get('assignedTo')
        if assigned_to_id_param and assigned_to_id_param.isdigit():
            # This now filters the user's visible tasks to only show those
            # also assigned to the specified user.
            base_queryset = base_queryset.filter(
                 Q(assigned_to_id=assigned_to_id_param) | 
                 Q(collaborators__id=assigned_to_id_param)
                 # Note: Filtering by workflow step assignee here is complex and might be omitted for simplicity
                 # unless it's a critical feature for the filter dropdown.
            ).distinct()

        # --- Category filter ---
        category_id = query_params.get('category')
        if category_id:
            base_queryset = base_queryset.filter(category_id=category_id)

        # --- Overdue filter ---
        if query_params.get('overdue', '').lower() == 'true':
            base_queryset = base_queryset.filter(
                deadline__lt=timezone.now().date(),
                status__in=['pending', 'in_progress']
            )
        
        # --- Search filter ---
        search_term = query_params.get('search')
        if search_term:
            base_queryset = base_queryset.filter(
                Q(title__icontains=search_term) |
                Q(description__icontains=search_term) |
                Q(client__name__icontains=search_term) |
                Q(category__name__icontains=search_term)
            ).distinct()
        
        # 3. Apply ordering.
        ordering = query_params.get('ordering', 'priority') # Default to priority
        valid_ordering_fields = [
            'priority', '-priority', 'deadline', '-deadline', 
            'title', '-title', 'client__name', '-client__name', 'status', '-status'
        ]
        if ordering in valid_ordering_fields:
            base_queryset = base_queryset.order_by(ordering)
        else:
            base_queryset = base_queryset.order_by('priority', 'deadline') # Fallback ordering

        # 4. Return the final, optimized, and secure queryset.
        return base_queryset.select_related(
            'client', 'category', 'assigned_to', 'workflow', 'current_workflow_step'
        ).prefetch_related('collaborators')
         
    def perform_create(self, serializer):
        user_making_request = self.request.user
        logger.debug(f"[TaskViewSet perform_create] User making request: {user_making_request.username}")

        try:
            profile_making_request = Profile.objects.select_related('organization').get(user=user_making_request)
            
            if not (profile_making_request.is_org_admin or profile_making_request.can_create_tasks):
                logger.warning(f"[TaskViewSet perform_create] User {user_making_request.username} does not have permission to create tasks.")
                raise PermissionDenied("Você não tem permissão para criar tarefas")
                
            client_id = self.request.data.get('client')
            client_instance = None
            if client_id:
                try:
                    client_instance = Client.objects.select_related('organization').get(id=client_id)
                    if client_instance.organization != profile_making_request.organization:
                         logger.warning(f"[TaskViewSet perform_create] User {user_making_request.username} trying to create task for client in another org.")
                         raise PermissionDenied("Não pode criar tarefa para cliente de outra organização.")
                    # A verificação profile.can_access_client pode ser redundante se can_create_tasks já implicar isso
                    # ou se a lógica de atribuição de cliente já for feita no frontend com base nos clientes visíveis.
                except Client.DoesNotExist:
                    logger.error(f"[TaskViewSet perform_create] Client with ID {client_id} not found.")
                    raise ValidationError({"client": "Cliente não encontrado."})
            else:
                logger.error("[TaskViewSet perform_create] Client ID is mandatory.")
                raise ValidationError({"client": "Cliente é obrigatório."})

            collaborators_ids_from_request = self.request.data.get('collaborators', []) # Vem do frontend
            assigned_to_id_from_request = self.request.data.get('assigned_to') # Vem do frontend

            # Validar colaboradores pertencem à mesma organização (se houver)
            if collaborators_ids_from_request:
                collaborators_qs = User.objects.filter(id__in=collaborators_ids_from_request)
                for collaborator_user in collaborators_qs:
                    try:
                        collab_profile = collaborator_user.profile
                        if collab_profile.organization != profile_making_request.organization:
                            logger.warning(f"[TaskViewSet perform_create] Collaborator {collaborator_user.username} not in same org.")
                            raise PermissionDenied(f"Colaborador {collaborator_user.username} não pertence à sua organização.")
                    except Profile.DoesNotExist:
                        logger.error(f"[TaskViewSet perform_create] Profile for collaborator {collaborator_user.username} not found.")
                        raise ValidationError({"collaborators": f"Colaborador {collaborator_user.username} não possui perfil válido."})
            
            # O serializer.save() irá criar a tarefa com assigned_to e
            # depois vamos definir os collaborators se existirem.
            # O serializer deve ser configurado para aceitar 'collaborators' como writeable ManyToManyField.
            # No frontend, ao submeter, deve enviar collaborators como uma lista de IDs.
            
            step_assignments = self.request.data.get('workflow_step_assignments', {})
            workflow_id = self.request.data.get('workflow')

            # Save task
            # O serializer.save() precisa que 'collaborators' não esteja no validated_data inicial
            # se for ser tratado separadamente com .set()
            validated_data_for_creation = serializer.validated_data.copy()
            
            # collaborators_ids_for_set é a lista de IDs para o M2M
            collaborators_ids_for_set = validated_data_for_creation.pop('collaborators', []) 
            
            task = serializer.save(
                created_by=user_making_request, 
                client=client_instance,
                # Se 'assigned_to' e 'collaborators' já estão no validated_data do serializer, ótimo.
                # Caso contrário, o 'assigned_to' já estaria sendo salvo e 'collaborators' precisa ser setado depois.
                workflow_step_assignments=step_assignments if workflow_id else {}
            )
            logger.debug(f"[TaskViewSet perform_create] Task {task.id} created successfully by serializer.")

            # Set collaborators if any were provided (and if mode was 'multiple' potentially)
            # This step assumes serializer's `create` method does not handle 'collaborators' M2M directly,
            # or if you prefer to set it explicitly after instance creation.
            if collaborators_ids_for_set:
                task.collaborators.set(collaborators_ids_for_set)
                logger.debug(f"[TaskViewSet perform_create] Collaborators {collaborators_ids_for_set} set for task {task.id}.")


            # Handle workflow assignment
            if workflow_id:
                try:
                    workflow = WorkflowDefinition.objects.get(id=workflow_id, is_active=True)
                    first_step = workflow.steps.order_by('order').first()
                    if first_step:
                        task.workflow = workflow
                        task.current_workflow_step = first_step
                        task.save(update_fields=['workflow', 'current_workflow_step'])
                        WorkflowHistory.objects.create(
                            task=task, from_step=None, to_step=first_step,
                            changed_by=user_making_request, action='workflow_assigned',
                            comment=f"Workflow '{workflow.name}' atribuído na criação da tarefa."
                        )
                        # NotificationService.notify_workflow_assigned(task, user_making_request)
                        # Esta notificação acima é para o primeiro responsável do passo do workflow.
                        # A nova notificação abaixo é para a atribuição geral da tarefa.
                    else:
                        logger.warning(f"[TaskViewSet perform_create] Workflow {workflow_id} for task {task.id} has no steps.")
                except WorkflowDefinition.DoesNotExist:
                    logger.warning(f"[TaskViewSet perform_create] Workflow {workflow_id} not found or inactive for task {task.id}.")

            # --- NOTIFICATION LOGIC ---
            users_to_notify_ids = set()
            if task.assigned_to:
                users_to_notify_ids.add(task.assigned_to.id)
            
            # Get collaborators from the task instance after .set()
            for collaborator in task.collaborators.all():
                users_to_notify_ids.add(collaborator.id)

            logger.debug(f"[TaskViewSet perform_create] Users to notify for new task assignment: {users_to_notify_ids}")

            for user_id_to_notify in users_to_notify_ids:
                # Não notificar o criador da tarefa se ele for um dos atribuídos
                if user_id_to_notify == user_making_request.id:
                    continue
                
                try:
                    target_user = User.objects.get(id=user_id_to_notify)
                    NotificationService.create_notification(
                        user=target_user,
                        task=task,
                        notification_type='task_assigned_to_you', # NOVO TIPO
                        title=f"Nova tarefa atribuída: {task.title}",
                        message=f"Você foi atribuído(a) à tarefa '{task.title}' para o cliente '{task.client.name}'. Criada por: {user_making_request.username}.",
                        created_by=user_making_request
                    )
                    logger.info(f"[TaskViewSet perform_create] Sent 'task_assigned_to_you' notification to {target_user.username} for task {task.id}")
                except User.DoesNotExist:
                    logger.error(f"[TaskViewSet perform_create] User with ID {user_id_to_notify} not found for notification.")
                except Exception as e_notify:
                    logger.error(f"[TaskViewSet perform_create] Error sending task assignment notification to user ID {user_id_to_notify}: {e_notify}")
            
            # Se um workflow foi atribuído E tem um primeiro passo com responsável,
            # a notificação de "passo pronto" para esse responsável já é tratada por notify_workflow_assigned.
            # Mas, se o primeiro passo do workflow não tiver um responsável direto,
            # e a tarefa tiver um `assigned_to` geral, ele já foi notificado acima.
            # Se o primeiro passo do workflow tiver um responsável diferente do `assigned_to` geral e dos `collaborators`,
            # a notificação `notify_workflow_assigned` (que internamente chama `notify_step_ready`) cuidará disso.
            if task.workflow and task.current_workflow_step and task.current_workflow_step.assign_to:
                 if task.current_workflow_step.assign_to.id not in users_to_notify_ids and task.current_workflow_step.assign_to.id != user_making_request.id:
                    NotificationService.notify_step_ready(task, task.current_workflow_step, user_making_request)
                    logger.info(f"[TaskViewSet perform_create] Sent 'step_ready' for initial workflow step to {task.current_workflow_step.assign_to.username}")


        except Profile.DoesNotExist:
            # Esta parte é para o superuser que não tem perfil (não recomendado)
            logger.warning(f"[TaskViewSet perform_create] Profile.DoesNotExist for request user {user_making_request.username}. Is superuser: {user_making_request.is_superuser}")
            if user_making_request.is_superuser:
                client_id = self.request.data.get('client')
                if not client_id: 
                    raise ValidationError({"client": "Cliente é obrigatório."})
                try:
                    client_instance = Client.objects.get(id=client_id)
                    # Superuser path, less restrictive, but needs care
                    # Ensure collaborators are handled if passed for superuser
                    validated_data_for_creation = serializer.validated_data.copy()
                    collaborators_ids_for_set = validated_data_for_creation.pop('collaborators', [])

                    task = serializer.save(
                        created_by=user_making_request, 
                        client=client_instance,
                        # assigned_to handled by serializer
                        # collaborators to be set after save
                        workflow_step_assignments=self.request.data.get('workflow_step_assignments', {}) if self.request.data.get('workflow') else {}
                    )
                    if collaborators_ids_for_set:
                        task.collaborators.set(collaborators_ids_for_set)
                    
                    # ... (notification logic for superuser-created tasks could be added here if needed) ...

                except Client.DoesNotExist:
                    raise ValidationError({"client": "Cliente não encontrado."})
            else:
                raise PermissionDenied("Perfil de usuário não encontrado")
        except Exception as e:
            logger.error(f"[TaskViewSet perform_create] Unexpected error: {e}", exc_info=True)
            raise # Re-raise a generic server error or a specific one

    @action(detail=True, methods=['get'])
    def workflow_status(self, request, pk=None):
        """
        Recupera todos os dados necessários para a visualização do workflow de uma tarefa.
        """
        task = self.get_object()  # Isto irá tratar do erro 404 se a tarefa não existir

        # Permissão para ver o workflow desta tarefa
        profile = request.user.profile
        if not (profile.is_org_admin or profile.can_view_all_tasks or task.can_user_access_task(request.user)):
             return Response({'error': 'Você não tem permissão para ver o workflow desta tarefa.'}, status=status.HTTP_403_FORBIDDEN)

        if not task.workflow:
            return Response({'error': 'Esta tarefa não tem um workflow associado.'}, status=status.HTTP_404_NOT_FOUND)
        
        # Preparar a resposta com todos os dados necessários
        history = task.workflow_history.order_by('-created_at')
        approvals = task.approvals.all()
        all_workflow_steps = task.workflow.steps.select_related('assign_to').order_by('order')
        
        time_by_step = TimeEntry.objects.filter(
            task=task,
            workflow_step__isnull=False
        ).values('workflow_step_id').annotate(
            total_minutes=Sum('minutes_spent')
        ).order_by()
        
        time_by_step_dict = {
            str(item['workflow_step_id']): item['total_minutes'] or 0
            for item in time_by_step
        }

        current_step_order = float('inf')
        if task.current_workflow_step:
            current_step_order = task.current_workflow_step.order

        steps_data = []
        for step in all_workflow_steps:
            is_completed = (task.status == 'completed' or (task.current_workflow_step and step.order < current_step_order))
            if not is_completed:
                 if history.filter(from_step=step, action__in=['step_completed', 'step_advanced', 'workflow_completed']).exists():
                     is_completed = True

            steps_data.append({
                'id': str(step.id),
                'name': step.name,
                'order': step.order,
                'description': step.description,
                'requires_approval': step.requires_approval,
                'approver_role': step.approver_role,
                'assign_to_name': step.assign_to.username if step.assign_to else None,
                'is_current': task.current_workflow_step_id == step.id,
                'is_completed': is_completed,
                'next_steps': [str(ns.id) for ns in step.next_steps.all()]
            })

        response_data = {
            'workflow': {
                'name': task.workflow.name,
                'steps': steps_data,
                'is_completed': task.status == 'completed' and not task.current_workflow_step,
                'progress': task.get_workflow_progress_data(),
                'time_by_step': time_by_step_dict,
            },
            'current_step': WorkflowStepSerializer(task.current_workflow_step).data if task.current_workflow_step else None,
            'approvals': TaskApprovalSerializer(approvals, many=True).data,
            'history': WorkflowHistorySerializer(history, many=True).data,
            'task': self.get_serializer(task).data,
        }

        return Response(response_data)
    @action(detail=True, methods=['post'])
    def assign_users(self, request, pk=None):
        """
        New action to manage user assignments to a task.
        Supports adding/removing primary assignee and collaborators.
        """
        task = self.get_object()
        
        try:
            profile = Profile.objects.get(user=request.user)
            
            # Check permissions
            can_assign = (
                profile.is_org_admin or 
                profile.can_edit_all_tasks or
                (profile.can_edit_assigned_tasks and task.can_user_access_task(request.user))
            )
            
            if not can_assign:
                raise PermissionDenied("Você não tem permissão para gerenciar atribuições desta tarefa")
            
            # Ensure task belongs to the user's organization
            if task.client.organization != profile.organization and not profile.is_org_admin:
                raise PermissionDenied("Não pode gerenciar atribuições de tarefas de outra organização.")

            action_type = request.data.get('action')  # 'set_primary', 'add_collaborators', 'remove_collaborators', 'set_all'
            user_ids = request.data.get('user_ids', [])
            
            if action_type == 'set_primary':
                # Set primary assignee
                if len(user_ids) != 1:
                    return Response({"error": "Deve especificar exatamente um usuário para responsável principal"}, 
                                  status=status.HTTP_400_BAD_REQUEST)
                
                try:
                    new_assignee = User.objects.get(id=user_ids[0])
                    # Validate user belongs to organization
                    new_assignee_profile = new_assignee.profile
                    if new_assignee_profile.organization != profile.organization:
                        raise PermissionDenied("Não pode atribuir tarefa a usuário de outra organização.")
                    
                    # Remove from collaborators if present
                    task.collaborators.remove(new_assignee)
                    task.assigned_to = new_assignee
                    task.save()
                    
                except User.DoesNotExist:
                    return Response({"error": "Usuário não encontrado"}, status=status.HTTP_404_NOT_FOUND)
                except Profile.DoesNotExist:
                    return Response({"error": "Usuário não possui perfil válido"}, status=status.HTTP_400_BAD_REQUEST)
            
            elif action_type == 'add_collaborators':
                # Add collaborators
                users_to_add = User.objects.filter(id__in=user_ids)
                valid_users = []
                
                for user in users_to_add:
                    try:
                        user_profile = user.profile
                        if user_profile.organization == profile.organization and user != task.assigned_to:
                            valid_users.append(user)
                    except Profile.DoesNotExist:
                        continue
                
                task.collaborators.add(*valid_users)
                
            elif action_type == 'remove_collaborators':
                # Remove collaborators
                users_to_remove = User.objects.filter(id__in=user_ids)
                task.collaborators.remove(*users_to_remove)
                
            elif action_type == 'set_all':
                # Set all assignments at once
                primary_user_id = request.data.get('primary_user_id')
                collaborator_ids = request.data.get('collaborator_ids', [])
                
                # Set primary assignee
                if primary_user_id:
                    try:
                        new_assignee = User.objects.get(id=primary_user_id)
                        new_assignee_profile = new_assignee.profile
                        if new_assignee_profile.organization != profile.organization:
                            raise PermissionDenied("Responsável principal deve pertencer à sua organização.")
                        task.assigned_to = new_assignee
                    except (User.DoesNotExist, Profile.DoesNotExist):
                        return Response({"error": "Responsável principal inválido"}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    task.assigned_to = None
                
                # Set collaborators
                if collaborator_ids:
                    collaborators = User.objects.filter(id__in=collaborator_ids)
                    valid_collaborators = []
                    
                    for user in collaborators:
                        try:
                            user_profile = user.profile
                            if (user_profile.organization == profile.organization and 
                                user != task.assigned_to):
                                valid_collaborators.append(user)
                        except Profile.DoesNotExist:
                            continue
                    
                    task.collaborators.set(valid_collaborators)
                else:
                    task.collaborators.clear()
                
                task.save()
            
            else:
                return Response({"error": "Ação inválida"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Return updated task data
            serializer = self.get_serializer(task)
            # Log organization action
            log_organization_action(
                request,
                action_type='ASSIGN_USERS_TO_TASK',
                action_description=f"Atribuição de usuários na tarefa: {task.title} (ID: {task.id}) - ação: {action_type}",
                related_object=task
            )
            return Response(serializer.data)
            
        except Profile.DoesNotExist:
            if request.user.is_superuser:
                # Handle superuser case
                pass
            else:
                raise PermissionDenied("Perfil de usuário não encontrado")

    @action(detail=True, methods=['get'])
    def assignment_suggestions(self, request, pk=None):
        """
        Suggest users for assignment based on:
        - Organization members
        - Previous task history
        - Workflow step assignments
        - Client experience
        """
        task = self.get_object()
        
        try:
            profile = Profile.objects.get(user=request.user)
            
            if not profile.organization:
                return Response({"error": "Usuário não pertence a uma organização"}, 
                              status=status.HTTP_400_BAD_REQUEST)
            
            # Get organization members
            org_members = Profile.objects.filter(
                organization=profile.organization,
                user__is_active=True
            ).select_related('user').exclude(user=request.user)
            
            suggestions = []
            
            for member_profile in org_members:
                user = member_profile.user
                suggestion = {
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'role': member_profile.role,
                    'relevance_score': 0,
                    'reasons': []
                }
                
                # Score based on previous work with this client
                client_tasks_count = Task.objects.filter(
                    client=task.client,
                    assigned_to=user
                ).count()
                
                if client_tasks_count > 0:
                    suggestion['relevance_score'] += client_tasks_count * 2
                    suggestion['reasons'].append(f"Trabalhou em {client_tasks_count} tarefa(s) deste cliente")
                
                # Score based on similar category experience
                if task.category:
                    category_tasks_count = Task.objects.filter(
                        category=task.category,
                        assigned_to=user
                    ).count()
                    
                    if category_tasks_count > 0:
                        suggestion['relevance_score'] += category_tasks_count
                        suggestion['reasons'].append(f"Experiência em {task.category.name}")
                
                # Score based on workflow step assignments
                if task.workflow and task.workflow_step_assignments:
                    if str(user.id) in task.workflow_step_assignments.values():
                        suggestion['relevance_score'] += 5
                        suggestion['reasons'].append("Atribuído a passos do workflow")
                
                # Score based on current workload (lower workload = higher score)
                current_tasks = Task.objects.filter(
                    Q(assigned_to=user) | Q(collaborators=user),
                    status__in=['pending', 'in_progress']
                ).distinct().count()
                
                if current_tasks < 5:
                    suggestion['relevance_score'] += 3
                    suggestion['reasons'].append("Baixa carga de trabalho atual")
                elif current_tasks > 10:
                    suggestion['relevance_score'] -= 2
                    suggestion['reasons'].append("Alta carga de trabalho atual")
                
                suggestions.append(suggestion)
            
            # Sort by relevance score
            suggestions.sort(key=lambda x: x['relevance_score'], reverse=True)
            
            return Response({
                'suggestions': suggestions[:10],  # Top 10 suggestions
                'total_members': len(org_members)
            })
            
        except Profile.DoesNotExist:
            return Response({"error": "Perfil não encontrado"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def my_assignments(self, request):
        """
        Get all tasks assigned to the current user in any capacity.
        """
        user = request.user
        
        try:
            profile = user.profile
            if not profile.organization:
                return Response({"tasks": []})
            
            # Get tasks where user is assigned in any capacity
            my_tasks = Task.objects.filter(
                Q(assigned_to=user) |  # Primary assignee
                Q(collaborators=user) |  # Collaborator
                Q(workflow_step_assignments__has_value=str(user.id)),  # Workflow step assignee
                client__organization=profile.organization
            ).select_related(
                'client', 'category', 'assigned_to', 'workflow', 'current_workflow_step'
            ).prefetch_related('collaborators').distinct()
            
            # Group by assignment type
            assignments = {
                'primary_tasks': [],
                'collaborative_tasks': [],
                'workflow_step_tasks': [],
                'summary': {
                    'total_tasks': 0,
                    'pending_tasks': 0,
                    'in_progress_tasks': 0,
                    'overdue_tasks': 0
                }
            }
            
            today = timezone.now().date()
            
            for task in my_tasks:
                task_data = {
                    'id': task.id,
                    'title': task.title,
                    'client_name': task.client.name,
                    'status': task.status,
                    'priority': task.priority,
                    'deadline': task.deadline,
                    'is_overdue': task.deadline and task.deadline.date() < today and task.status != 'completed',
                    'role_in_task': task.get_user_role_in_task(user)
                }
                
                # Categorize by assignment type
                if task.assigned_to == user:
                    assignments['primary_tasks'].append(task_data)
                elif task.collaborators.filter(id=user.id).exists():
                    assignments['collaborative_tasks'].append(task_data)
                elif (task.workflow_step_assignments and 
                      str(user.id) in task.workflow_step_assignments.values()):
                    assignments['workflow_step_tasks'].append(task_data)
                
                # Update summary
                assignments['summary']['total_tasks'] += 1
                if task.status == 'pending':
                    assignments['summary']['pending_tasks'] += 1
                elif task.status == 'in_progress':
                    assignments['summary']['in_progress_tasks'] += 1
                
                if task_data['is_overdue']:
                    assignments['summary']['overdue_tasks'] += 1
            
            return Response(assignments)
            
        except Profile.DoesNotExist:
            return Response({"error": "Perfil não encontrado"}, status=status.HTTP_404_NOT_FOUND)

class CreateUserView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]

class TimeEntryViewSet(viewsets.ModelViewSet):
    serializer_class = TimeEntrySerializer
    permission_classes = [IsAuthenticated, CanManageTimeEntry]

    def get_queryset(self):
        """
        Uses the custom manager and then applies request filters.
        """
        user = self.request.user
        base_queryset = TimeEntry.objects.for_user(user)

        # Apply optimizations
        optimized_queryset = base_queryset.select_related(
            'user', 'client', 'task', 'category', 'workflow_step'
        )

        # Apply query parameter filters
        query_params = self.request.query_params
        start_date_str = query_params.get('start_date')
        if start_date_str:
            optimized_queryset = optimized_queryset.filter(date__gte=start_date_str)

        end_date_str = query_params.get('end_date')
        if end_date_str:
            optimized_queryset = optimized_queryset.filter(date__lte=end_date_str)
            
        client_id = query_params.get('client')
        if client_id:
            optimized_queryset = optimized_queryset.filter(client_id=client_id)
        
        user_id_param = query_params.get('user_id_param')
        if user_id_param and (user.is_superuser or (hasattr(user, 'profile') and user.profile.is_org_admin)):
            # This allows an admin to filter for a specific user's entries
            # within the already organization-scoped queryset.
            optimized_queryset = optimized_queryset.filter(user_id=user_id_param)

        search_term = query_params.get('search')
        if search_term:
            optimized_queryset = optimized_queryset.filter(description__icontains=search_term)
        
        # Apply ordering
        ordering = query_params.get('ordering', '-date')
        valid_ordering = ['date', '-date', 'minutes_spent', '-minutes_spent']
        if ordering in valid_ordering:
            optimized_queryset = optimized_queryset.order_by(ordering)

        return optimized_queryset

    def perform_create(self, serializer):
        """
        Ensures the user logging time has permission for the specified client/task.
        """
        user = self.request.user
        profile = user.profile

        if not (profile.can_log_time or profile.is_org_admin):
            raise PermissionDenied("Você não tem permissão para registrar tempo.")
            
        client = serializer.validated_data.get('client')
        # Check if user can access this client
        if not profile.can_access_client(client):
            raise PermissionDenied("Você não tem acesso a este cliente para registrar tempo.")

        time_entry = serializer.save(user=user)
        # The processing logic can be moved here from the serializer or called from a service
        # For now, let's assume it's called via a signal or remains in the serializer's save.
        self._process_workflow_and_status_for_time_entry(time_entry)
        # Log organization action
        log_organization_action(
            self.request,
            action_type='CREATE_TIME_ENTRY',
            action_description=f"Registo de tempo criado: {time_entry.description} (ID: {time_entry.id}) para tarefa {time_entry.task.id if time_entry.task else 'N/A'}",
            related_object=time_entry
        )

    def _process_workflow_and_status_for_time_entry(self, time_entry: TimeEntry):
        """
        Processa as mudanças de status da tarefa e do workflow após a criação de um TimeEntry.
        """
        if not time_entry.task:
            return

        task = time_entry.task
        user = time_entry.user
        
        step_being_worked_on = time_entry.workflow_step

        if step_being_worked_on and time_entry.workflow_step_completed:
            if task.current_workflow_step == step_being_worked_on:
                logger.info(f"TimeEntry {time_entry.id}: Tentando completar o passo '{step_being_worked_on.name}' para a tarefa {task.id}.")
                
                # WorkflowService.complete_step_and_advance will handle logging history,
                # notifying for step_completed, and then advancing (which notifies for step_ready).
                # It will also handle task completion if the workflow ends and sets the task status.
                success, message = WorkflowService.complete_step_and_advance(
                    task=task,
                    step_to_complete=step_being_worked_on,
                    user=user,
                    completion_comment=f"Passo concluído via registo de tempo: {time_entry.description}",
                    time_spent_on_step=time_entry.minutes_spent,
                    should_auto_advance=time_entry.advance_workflow
                )
                # If the workflow service completed the task, notify_task_completed is already called from there.
                # We don't need to call it again here unless the task status logic below does something different.
            else:
                logger.warning(
                    f"TimeEntry {time_entry.id} marcou o passo '{step_being_worked_on.name}' como concluído, "
                    f"mas o passo atual da tarefa {task.id} é '{task.current_workflow_step.name if task.current_workflow_step else 'Nenhum'}'. "
                    "O avanço do workflow não será acionado a partir de um passo que não é o atual."
                )
        elif step_being_worked_on: # Work logged on a step, but step not marked as completed via this time entry
            WorkflowService._log_workflow_history(
                task=task, from_step=step_being_worked_on, to_step=None, 
                changed_by=user, action='step_work_logged',
                comment=f"Tempo: {time_entry.minutes_spent}m. Desc: {time_entry.description}",
                time_spent_minutes=time_entry.minutes_spent
            )

        # Handle task status change *after* workflow logic, as workflow might change status
        if time_entry.task_status_after != 'no_change':
            task.refresh_from_db() # Get the latest status (potentially updated by WorkflowService)
            
            if task.status != time_entry.task_status_after: # Only proceed if the requested status is different
                old_status_before_time_entry_logic = task.status
                
                task.status = time_entry.task_status_after
                
                if task.status == 'completed':
                    task.completed_at = timezone.now()
                    # If the workflow completion above ALREADY set the task to completed, this notification might be redundant.
                    # However, this logic handles cases where time entry directly completes a task *without* workflow interaction
                    # or if the workflow logic didn't mark it as completed but this time entry action does.
                    if old_status_before_time_entry_logic != 'completed': # Avoid double notification if workflow already did it
                        task.save(update_fields=['status', 'completed_at'])
                        NotificationService.notify_task_completed(task, user) 
                    else:
                        task.save(update_fields=['status', 'completed_at']) # Still save if status changed to completed
                elif old_status_before_time_entry_logic == 'completed': # Task was completed, now it's being reopened
                    task.completed_at = None
                    task.save(update_fields=['status', 'completed_at'])
                else: # Other status changes (e.g., pending to in_progress)
                    task.save(update_fields=['status'])
                
                logger.info(f"Status da tarefa {task.title} (via time entry status_after) alterado de '{old_status_before_time_entry_logic}' para '{task.status}'.")
            else:
                logger.info(f"Status da tarefa {task.title} não alterado via time entry, pois já é '{task.status}'.")

    def update(self, request, *args, **kwargs):
        try:
            profile = Profile.objects.get(user=request.user)
            time_entry_instance = self.get_object()
            can_edit = (
                profile.is_org_admin or 
                profile.can_edit_all_time or 
                (profile.can_edit_own_time and time_entry_instance.user == request.user)
            )
            if not can_edit:
                raise PermissionDenied("Você não tem permissão para editar este registro de tempo")

            if time_entry_instance.client.organization != profile.organization and not profile.is_org_admin:
                 raise PermissionDenied("Não pode editar registro de tempo de cliente de outra organização.")
                
            new_client_id = request.data.get('client')
            if new_client_id and str(time_entry_instance.client.id) != str(new_client_id):
                try:
                    new_client = Client.objects.get(id=new_client_id)
                    if new_client.organization != profile.organization and not profile.is_org_admin:
                        raise PermissionDenied("Não pode mover registro para cliente de outra organização.")
                    if not profile.can_access_client(new_client):
                        raise PermissionDenied("Você não tem acesso ao novo cliente selecionado")
                except Client.DoesNotExist:
                     raise ValidationError({"client": "Novo cliente não encontrado."})
            return super().update(request, *args, **kwargs)
        except Profile.DoesNotExist:
            if request.user.is_superuser:
                time_entry_instance = self.get_object()
                # Superuser can edit any, but not change client to other orgs unless specific logic allows
                return super().update(request, *args, **kwargs)
            raise PermissionDenied("Perfil de usuário não encontrado")
            
    def destroy(self, request, *args, **kwargs):
        try:
            profile = Profile.objects.get(user=request.user)
            time_entry_instance = self.get_object()
            can_delete = (
                profile.is_org_admin or 
                profile.can_edit_all_time or # Assuming can_edit_all_time implies can_delete_all_time
                (profile.can_edit_own_time and time_entry_instance.user == request.user) # edit_own implies delete_own
            )
            if not can_delete:
                raise PermissionDenied("Você não tem permissão para excluir este registro de tempo")

            if time_entry_instance.client.organization != profile.organization and not profile.is_org_admin:
                 raise PermissionDenied("Não pode excluir registro de tempo de cliente de outra organização.")
            return super().destroy(request, *args, **kwargs)
        except Profile.DoesNotExist:
            if request.user.is_superuser:
                return super().destroy(request, *args, **kwargs)
            raise PermissionDenied("Perfil de usuário não encontrado")
            
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        entries_data = request.data.get('entries', [])
        if not isinstance(entries_data, list):
            return Response({"error": "O campo 'entries' deve ser uma lista."}, status=status.HTTP_400_BAD_REQUEST)
        if not entries_data:
            return Response({"error": "Nenhum registro de tempo fornecido"}, status=status.HTTP_400_BAD_REQUEST)
            
        profile = None
        if not request.user.is_superuser:
            try:
                profile = Profile.objects.get(user=request.user)
                if not profile.can_log_time:
                    raise PermissionDenied("Você não tem permissão para registrar tempo")
            except Profile.DoesNotExist:
                raise PermissionDenied("Perfil de usuário não encontrado e sem permissão para registrar tempo")

        created_entries_data = []
        errors = []

        for entry_data_item in entries_data: # Renamed to avoid conflict
            entry_data_item['user'] = request.user.id 
            
            client_id = entry_data_item.get('client')
            if client_id:
                try:
                    client = Client.objects.get(id=client_id)
                    if profile and client.organization != profile.organization and not profile.is_org_admin:
                        errors.append({(entry_data_item.get('description') or f"Entrada para cliente {client_id}"): f"Cliente {client.name} não pertence à sua organização."})
                        continue
                    if profile and not profile.can_access_client(client):
                        errors.append({(entry_data_item.get('description') or f"Entrada para cliente {client_id}"): f"Sem acesso ao cliente {client.name}."})
                        continue
                except Client.DoesNotExist:
                    errors.append({(entry_data_item.get('description') or f"Entrada para cliente {client_id}"): f"Cliente com ID {client_id} não encontrado."})
                    continue
            
            serializer = self.get_serializer(data=entry_data_item)
            if serializer.is_valid():
                try:
                    time_entry_instance = serializer.save() 
                    created_entries_data.append(serializer.data)
                except PermissionDenied as e: 
                    errors.append({(entry_data_item.get('description') or f"Entrada") : str(e)})
                except Exception as e:
                     errors.append({(entry_data_item.get('description') or f"Entrada") : f"Erro ao salvar: {str(e)}"})
            else:
                errors.append({(entry_data_item.get('description') or f"Entrada") : serializer.errors})

        status_code = status.HTTP_201_CREATED
        response_payload = {"created_entries": created_entries_data}

        if errors:
            response_payload["errors"] = errors
            response_payload["message"] = "Alguns registros não puderam ser criados."
            status_code = status.HTTP_207_MULTI_STATUS if created_entries_data else status.HTTP_400_BAD_REQUEST
        # Log organization action for bulk time entry creation
        log_organization_action(
            request,
            action_type='BULK_CREATE_TIME_ENTRIES',
            action_description=f"{len(created_entries_data)} registros de tempo criados em lote. Erros: {len(errors)}.",
            related_object=None
        )
        # Notify all org admins
        if profile and profile.organization:
            org_admins = Profile.objects.filter(organization=profile.organization, is_org_admin=True).select_related('user')
            for admin_profile in org_admins:
                NotificationService.create_notification(
                    user=admin_profile.user,
                    notification_type='bulk_time_entry_created',
                    title=f"Registros de tempo criados em lote",
                    message=f"{len(created_entries_data)} registros de tempo foram criados em lote por {request.user.username}.",
                    created_by=request.user
                )
        return Response(response_payload, status=status_code)
    
class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = Expense.objects.select_related(
            'client__organization',
            'created_by__profile' # For created_by_name if needed via serializer
        )
        
        if not user.is_superuser:
            try:
                profile = Profile.objects.get(user=user)
                if not profile.organization:
                    return Expense.objects.none()
                queryset = queryset.filter(client__organization=profile.organization)
                
                if not (profile.is_org_admin or profile.can_manage_expenses): # Assuming can_view_expenses is covered by can_manage_expenses
                    # If user cannot manage all, they can only see expenses they created
                    # or expenses for clients they have visibility on (more complex)
                    queryset = queryset.filter(created_by=user) 
            except Profile.DoesNotExist:
                return Expense.objects.none()

        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date and end_date:
            try:
                datetime.strptime(start_date, '%Y-%m-%d')
                datetime.strptime(end_date, '%Y-%m-%d')
                queryset = queryset.filter(date__range=[start_date, end_date])
            except ValueError:
                 raise ValidationError("Formato de data inválido. Use YYYY-MM-DD.")
            
        client_id = self.request.query_params.get('client')
        if client_id:
            queryset = queryset.filter(client_id=client_id)
            
        return queryset
    
    def perform_create(self, serializer):
        user = self.request.user
        if not user.is_superuser:
            try:
                profile = Profile.objects.get(user=user)
                if not (profile.is_org_admin or profile.can_manage_expenses):
                    raise PermissionDenied("Você não tem permissão para criar despesas.")
                
                # Ensure client for expense is in user's org
                client_id = serializer.validated_data.get('client').id if serializer.validated_data.get('client') else None
                if client_id:
                    client_for_expense = Client.objects.get(id=client_id)
                    if client_for_expense.organization != profile.organization:
                        raise PermissionDenied("Não pode criar despesa para cliente de outra organização.")
            except Profile.DoesNotExist:
                raise PermissionDenied("Perfil não encontrado e sem permissão para criar despesas.")
            except Client.DoesNotExist:
                 raise ValidationError({"client": "Cliente para despesa não encontrado."})

        serializer.save(created_by=user)
        # Log organization action
        log_organization_action(
            self.request,
            action_type='CREATE_EXPENSE',
            action_description=f"Despesa criada: {serializer.instance.description} (ID: {serializer.instance.id}) para cliente {serializer.instance.client.name if serializer.instance.client else 'N/A'}",
            related_object=serializer.instance
        )
        # Notify all org admins
        org_admins = Profile.objects.filter(organization=serializer.instance.client.organization, is_org_admin=True).select_related('user')
        for admin_profile in org_admins:
            NotificationService.create_notification(
                user=admin_profile.user,
                notification_type='expense_created',
                title=f"Nova despesa criada",
                message=f"Uma nova despesa foi criada para o cliente {serializer.instance.client.name if serializer.instance.client else 'N/A'}.",
                created_by=user
            )

class ClientProfitabilityViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ClientProfitabilitySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        try:
            profile = user.profile
            if not profile.organization:
                return ClientProfitability.objects.none()
            
            # OPTIMIZATION: `select_related` on nested foreign keys
            base_queryset = ClientProfitability.objects.filter(
                client__organization=profile.organization
            ).select_related(
                'client', 'client__account_manager'
            )
            
            # Permission check
            can_view_any_profitability = (
                profile.is_org_admin or 
                profile.can_view_organization_profitability or 
                profile.can_view_team_profitability or 
                profile.can_view_profitability
            )
            if not can_view_any_profitability:
                return ClientProfitability.objects.none()

            # Apply filters
            year = self.request.query_params.get('year')
            month = self.request.query_params.get('month')
            if year and month:
                try:
                    base_queryset = base_queryset.filter(year=int(year), month=int(month))
                except ValueError:
                    raise ValidationError("Ano e mês devem ser números inteiros.")

            client_id = self.request.query_params.get('client')
            if client_id:
                base_queryset = base_queryset.filter(client_id=client_id)
                
            is_profitable_param = self.request.query_params.get('is_profitable')
            if is_profitable_param is not None:
                is_profitable_bool = is_profitable_param.lower() == 'true'
                base_queryset = base_queryset.filter(is_profitable=is_profitable_bool)
            
            # Apply role-based filtering
            if profile.is_org_admin or profile.can_view_organization_profitability or profile.can_view_team_profitability:
                return base_queryset
            elif profile.can_view_profitability:
                visible_client_ids = profile.visible_clients.values_list('id', flat=True)
                return base_queryset.filter(client_id__in=visible_client_ids)
            
            return ClientProfitability.objects.none()
                
        except Profile.DoesNotExist:
            if user.is_superuser:
                return ClientProfitability.objects.select_related('client', 'client__account_manager')
            return ClientProfitability.objects.none()


class WorkflowDefinitionViewSet(viewsets.ModelViewSet):
    # queryset already defined with prefetch
    serializer_class = WorkflowDefinitionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # OPTIMIZATION: Pre-fetch steps and related users to prevent N+1 queries when serializing steps.
        queryset = WorkflowDefinition.objects.select_related(
            'created_by'
        ).prefetch_related(
            Prefetch('steps', queryset=WorkflowStep.objects.select_related('assign_to').order_by('order'))
        )
        
        # NOTE: Add permission filtering here. For example, if workflows are organization-specific.
        # if hasattr(self.request.user, 'profile') and self.request.user.profile.organization:
        #     queryset = queryset.filter(Q(organization=self.request.user.profile.organization) | Q(organization__isnull=True))

        is_active_param = self.request.query_params.get('is_active')
        if is_active_param is not None:
            is_active = is_active_param.lower() == 'true'
            queryset = queryset.filter(is_active=is_active)
            
        return queryset
    
    def perform_create(self, serializer):
        user = self.request.user
        organization_to_assign = None
        try:
            profile = Profile.objects.get(user=user)
            if not (profile.is_org_admin or profile.can_create_workflows):
                raise PermissionDenied("Você não tem permissão para criar workflows")
            organization_to_assign = profile.organization # Assign to user's org
        except Profile.DoesNotExist:
            if not user.is_superuser: 
                raise PermissionDenied("Perfil de usuário não encontrado e sem permissão.")
        
        # If WorkflowDefinition model gets an 'organization' field:
        # serializer.save(created_by=user, organization=organization_to_assign)
        serializer.save(created_by=user) 
        # Log organization action
        log_organization_action(
            self.request,
            action_type='CREATE_WORKFLOW_DEFINITION',
            action_description=f"Workflow criado: {serializer.instance.name} (ID: {serializer.instance.id})",
            related_object=serializer.instance
        )
    
    def update(self, request, *args, **kwargs):
        user = request.user
        workflow_def = self.get_object() # Get instance before super().update
        try:
            profile = Profile.objects.get(user=user)
            if not (profile.is_org_admin or profile.can_edit_workflows):
                raise PermissionDenied("Você não tem permissão para editar workflows")
            # If WorkflowDefinition is org-specific:
            # if workflow_def.organization and workflow_def.organization != profile.organization:
            #    raise PermissionDenied("Não pode editar workflows de outra organização.")
        except Profile.DoesNotExist:
            if not user.is_superuser:
                raise PermissionDenied("Perfil de usuário não encontrado e sem permissão.")
        response = super().update(request, *args, **kwargs)
        # Log organization action
        log_organization_action(
            request,
            action_type='UPDATE_WORKFLOW_DEFINITION',
            action_description=f"Workflow atualizado: {workflow_def.name} (ID: {workflow_def.id})",
            related_object=workflow_def
        )
        return response
    
    def destroy(self, request, *args, **kwargs):
        user = request.user
        workflow_def = self.get_object()
        try:
            profile = Profile.objects.get(user=user)
            if not (profile.is_org_admin or profile.can_manage_workflows): # Use can_manage for delete
                raise PermissionDenied("Você não tem permissão para excluir workflows")
            # If WorkflowDefinition is org-specific:
            # if workflow_def.organization and workflow_def.organization != profile.organization:
            #    raise PermissionDenied("Não pode excluir workflows de outra organização.")
        except Profile.DoesNotExist:
             if not user.is_superuser:
                raise PermissionDenied("Perfil de usuário não encontrado e sem permissão.")
        
        # Check if workflow is in use before deleting
        if Task.objects.filter(workflow=workflow_def).exists():
            return Response(
                {"error": "Este workflow está em uso por uma ou mais tarefas e não pode ser excluído."},
                status=status.HTTP_400_BAD_REQUEST
            )
        response = super().destroy(request, *args, **kwargs)
        # Log organization action
        log_organization_action(
            request,
            action_type='DELETE_WORKFLOW_DEFINITION',
            action_description=f"Workflow excluído: {workflow_def.name} (ID: {workflow_def.id})",
            related_object=workflow_def
        )
        return response
    
    @action(detail=True, methods=['get'])
    def analyze(self, request, pk=None):
        """Analyzes a workflow for bottlenecks and optimization suggestions."""
        workflow = self.get_object()
        # Permission check needed here
        analysis_data = WorkflowAnalyticsService.analyze_workflow(workflow)
        return Response(analysis_data)
        
    @action(detail=True, methods=['post'])
    def assign_to_task(self, request, pk=None):
        workflow = self.get_object()
        task_id = request.data.get('task_id')
        
        if not task_id:
            return Response({"error": "ID da tarefa é obrigatório"}, status=status.HTTP_400_BAD_REQUEST)
            
        user = request.user
        try:
            profile = Profile.objects.get(user=user)
            if not (profile.is_org_admin or profile.can_assign_workflows):
                raise PermissionDenied("Você não tem permissão para atribuir workflows")
                
            task = Task.objects.select_related('client__organization').get(id=task_id)
            
            # Ensure workflow and task are compatible (e.g., same organization if workflows are org-specific)
            # if hasattr(workflow, 'organization') and workflow.organization and workflow.organization != profile.organization:
            #    raise PermissionDenied("Este workflow não pertence à sua organização.")
            if task.client.organization != profile.organization and not profile.is_org_admin:
                 raise PermissionDenied("Não pode atribuir workflow a uma tarefa de cliente de outra organização.")

            if not profile.can_access_client(task.client): # Check visibility if not admin
                    raise PermissionDenied("Você não tem acesso ao cliente desta tarefa")
                    
            first_step = workflow.steps.order_by('order').first()
            if not first_step:
                return Response({"error": "Este workflow não possui passos definidos e não pode ser atribuído."}, 
                                status=status.HTTP_400_BAD_REQUEST)
                    
            task.workflow = workflow
            task.current_workflow_step = first_step
            # Reset task status if it was completed, to allow workflow to run
            if task.status == 'completed':
                task.status = 'pending' # Or 'in_progress'
                task.completed_at = None
            task.save()
            
            WorkflowHistory.objects.create(
                task=task, from_step=None, to_step=first_step, changed_by=user,
                action='workflow_assigned', comment=f"Workflow '{workflow.name}' atribuído."
            )
            NotificationService.notify_workflow_assigned(task, user)
                
            return Response(TaskSerializer(task).data, status=status.HTTP_200_OK) # Return updated task
                
        except Task.DoesNotExist:
            return Response({"error": "Tarefa não encontrada"}, status=status.HTTP_404_NOT_FOUND)
        except Profile.DoesNotExist:
            if user.is_superuser: # Superuser specific logic if needed
                # ... handle superuser case ...
                pass
            raise PermissionDenied("Perfil de usuário não encontrado")


class WorkflowStepViewSet(viewsets.ModelViewSet):
    queryset = WorkflowStep.objects.select_related('workflow', 'assign_to').order_by('workflow__name', 'order')
    serializer_class = WorkflowStepSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # OPTIMIZATION: Annotate counts and boolean checks.
        base_queryset = WorkflowStep.objects.select_related(
            'workflow', 'assign_to'
        ).annotate(
            time_entries_count=Count('time_entries', distinct=True),
            total_time_spent=Sum('time_entries__minutes_spent'),
            is_approved=Exists(TaskApproval.objects.filter(workflow_step=OuterRef('pk'), approved=True))
        ).order_by('workflow__name', 'order')

        workflow_id = self.request.query_params.get('workflow')
        if workflow_id:
            base_queryset = base_queryset.filter(workflow_id=workflow_id)
        
        # NOTE: Add permission-based filtering here. For now, assuming anyone authenticated can see
        # all workflow step definitions, which might be too broad.
        # A better approach for non-admins might be:
        # `base_queryset.filter(workflow__tasks__client__organization=self.request.user.profile.organization).distinct()`
        # but this depends on your specific security requirements.
        
        return base_queryset

    def perform_create(self, serializer):
        user = self.request.user
        workflow_id = serializer.validated_data.get('workflow').id
        try:
            profile = Profile.objects.get(user=user)
            workflow_def = WorkflowDefinition.objects.get(id=workflow_id)
            # if hasattr(workflow_def, 'organization') and workflow_def.organization and workflow_def.organization != profile.organization:
            #    raise PermissionDenied("Não pode adicionar passos a workflows de outra organização.")
            if not (profile.is_org_admin or profile.can_edit_workflows): # Use edit_workflows for adding steps
                 raise PermissionDenied("Sem permissão para adicionar passos a este workflow.")
        except Profile.DoesNotExist:
            if not user.is_superuser:
                 raise PermissionDenied("Perfil não encontrado e sem permissão.")
        except WorkflowDefinition.DoesNotExist:
            raise ValidationError({"workflow": "WorkflowDefinition não encontrado."})
        serializer.save()
        # Log organization action
        log_organization_action(
            self.request,
            action_type='CREATE_WORKFLOW_STEP',
            action_description=f"Passo de workflow criado: {serializer.instance.name} (ID: {serializer.instance.id})",
            related_object=serializer.instance
        )

    def update(self, request, *args, **kwargs):
        user = request.user
        step_instance = self.get_object()
        try:
            profile = Profile.objects.get(user=user)
            # if hasattr(step_instance.workflow, 'organization') and step_instance.workflow.organization and step_instance.workflow.organization != profile.organization:
            #    raise PermissionDenied("Não pode modificar passos de workflows de outra organização.")
            if not (profile.is_org_admin or profile.can_edit_workflows):
                 raise PermissionDenied("Sem permissão para modificar passos deste workflow.")
        except Profile.DoesNotExist:
            if not user.is_superuser:
                raise PermissionDenied("Perfil não encontrado e sem permissão.")
        response = super().update(request, *args, **kwargs)
        # Log organization action
        log_organization_action(
            request,
            action_type='UPDATE_WORKFLOW_STEP',
            action_description=f"Passo de workflow atualizado: {step_instance.name} (ID: {step_instance.id})",
            related_object=step_instance
        )
        return response

    def destroy(self, request, *args, **kwargs):
        user = request.user
        step_instance = self.get_object()
        try:
            profile = Profile.objects.get(user=user)
            # if hasattr(step_instance.workflow, 'organization') and step_instance.workflow.organization and step_instance.workflow.organization != profile.organization:
            #    raise PermissionDenied("Não pode excluir passos de workflows de outra organização.")
            if not (profile.is_org_admin or profile.can_edit_workflows): # can_manage_workflows might be better here
                 raise PermissionDenied("Sem permissão para excluir passos deste workflow.")
        except Profile.DoesNotExist:
            if not user.is_superuser:
                raise PermissionDenied("Perfil não encontrado e sem permissão.")
        
        # Add check if step is part of an active task's current_workflow_step
        if Task.objects.filter(current_workflow_step=step_instance).exists():
            return Response(
                {"error": "Este passo está ativo em uma ou mais tarefas e não pode ser excluído."},
                status=status.HTTP_400_BAD_REQUEST
            )
        response = super().destroy(request, *args, **kwargs)
        # Log organization action
        log_organization_action(
            request,
            action_type='DELETE_WORKFLOW_STEP',
            action_description=f"Passo de workflow excluído: {step_instance.name} (ID: {step_instance.id})",
            related_object=step_instance
        )
        return response


class TaskApprovalViewSet(viewsets.ModelViewSet):
    serializer_class = TaskApprovalSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        # OPTIMIZATION
        base_queryset = TaskApproval.objects.select_related(
            'task__client', 'workflow_step', 'approved_by'
        )
        
        task_id = self.request.query_params.get('task')
        if task_id:
            base_queryset = base_queryset.filter(task_id=task_id)

        if user.is_superuser:
            return base_queryset

        try:
            profile = user.profile
            if not profile.organization: return TaskApproval.objects.none()
            
            queryset = base_queryset.filter(task__client__organization=profile.organization)
            
            # User should only see approvals if they can see the task, are the approver, or have specific perms.
            if not (profile.is_org_admin or profile.can_view_all_tasks or profile.can_approve_tasks):
                visible_client_ids = profile.visible_clients.values_list('id', flat=True)
                queryset = queryset.filter(
                    Q(task__client_id__in=visible_client_ids) | 
                    Q(task__assigned_to=user) | 
                    Q(approved_by=user)
                )
            return queryset.distinct()
        except Profile.DoesNotExist:
            return TaskApproval.objects.none()
    
    def perform_create(self, serializer):
        user = self.request.user
        validated_data = serializer.validated_data
        task = validated_data.get('task')
        workflow_step = validated_data.get('workflow_step')

        if not task or not workflow_step: # Should be caught by serializer validation
            raise ValidationError("Tarefa e passo do workflow são obrigatórios.")

        try:
            profile = Profile.objects.get(user=user)
            # Permission to approve/reject:
            # 1. Org admin
            # 2. User has general 'can_approve_tasks'
            # 3. User's role matches 'approver_role' defined on the workflow_step
            # 4. (Optional) User is specifically assigned as an approver for this task/step (more complex model needed)
            
            can_approve_this_step = profile.is_org_admin or profile.can_approve_tasks
            if not can_approve_this_step and workflow_step.approver_role:
                # Simple role check (case-insensitive substring match)
                if workflow_step.approver_role.lower() not in profile.role.lower():
                    raise PermissionDenied("Seu papel não permite aprovar/rejeitar este passo.")
            elif not can_approve_this_step: # No general perm and no specific role match
                raise PermissionDenied("Você não tem permissão para aprovar/rejeitar este passo.")

            if task.client.organization != profile.organization and not profile.is_org_admin:
                raise PermissionDenied("Não pode aprovar/rejeitar passo de tarefa de outra organização.")

            if task.current_workflow_step != workflow_step:
                raise ValidationError(f"Só é possível aprovar/rejeitar o passo atual '{task.current_workflow_step.name if task.current_workflow_step else 'N/A'}'. Este passo é '{workflow_step.name}'.")

        except Profile.DoesNotExist:
            if not user.is_superuser:
                raise PermissionDenied("Perfil de usuário não encontrado.")
        
        # Check if an approval/rejection already exists for this user, step, task to prevent duplicates
        if TaskApproval.objects.filter(task=task, workflow_step=workflow_step, approved_by=user).exists():
            raise ValidationError("Você já submeteu uma aprovação/rejeição para este passo desta tarefa.")

        approval = serializer.save(approved_by=user) # approved_at is auto_now_add
        
        action_taken = 'step_approved' if approval.approved else 'step_rejected'
        WorkflowHistory.objects.create(
            task=task, from_step=workflow_step, to_step=workflow_step, 
            changed_by=user, action=action_taken,
            comment=approval.comment or f"Passo {workflow_step.name} {action_taken.split('_')[1]}."
        )
        
        # ==============================================================================
        # NEW: Add notification for approval completion
        # ==============================================================================
        NotificationService.notify_approval_completed(
            task=approval.task,
            workflow_step=approval.workflow_step,
            approval_record=approval,
            approved_by=approval.approved_by
        )
        
        if approval.approved:
            logger.info(f"Passo {workflow_step.name} da tarefa {task.id} APROVADO por {user.username}")
            # Consider auto-advancing workflow if all required approvals are met (if multiple approvers were possible)
            # For now, single approval is sufficient. The _advance_workflow_step will check this approval.
        else: 
            logger.info(f"Passo {workflow_step.name} da tarefa {task.id} REJEITADO por {user.username}")

class WorkflowStepDetailViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = WorkflowStepSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        # OPTIMIZATION: This is the main queryset for the list view.
        # We pre-fetch everything that might be needed by the serializer or related actions.
        queryset = WorkflowStep.objects.select_related(
            'workflow', 
            'assign_to'
        ).prefetch_related(
            'time_entries__user', 
            'task_approvals__approved_by' 
        ).order_by('workflow__name', 'order')
        
        workflow_id_param = self.request.query_params.get('workflow')
        if workflow_id_param:
            queryset = queryset.filter(workflow_id=workflow_id_param)
        
        if user.is_superuser:
            return queryset

        # Apply permission-based filtering
        try:
            profile = user.profile
            if not profile.organization: 
                return WorkflowStep.objects.none()
            
            # This is a simplified permission. It assumes that if a user is in an org,
            # they can see all workflow step definitions for that org.
            # This might need to be more granular in a real application.
            # For now, we assume workflows are not tied to organizations in the model.
            # If they were, you would add: .filter(workflow__organization=profile.organization)
            
            # For non-admins, a stricter rule might be to only show steps from workflows
            # that are actually used in tasks they can see. This query is more complex:
            # accessible_workflow_ids = Task.objects.filter(...user access query...).values_list('workflow_id', flat=True).distinct()
            # queryset = queryset.filter(workflow_id__in=accessible_workflow_ids)
            
            return queryset

        except Profile.DoesNotExist:
            return WorkflowStep.objects.none()

    @action(detail=True, methods=['get'])
    def time_entries(self, request, pk=None):
        workflow_step = self.get_object() 
        user = request.user
        
        # Permission Check: Can user see time entries for this specific step?
        # This could depend on if they can see tasks associated with this step, or if they are an admin.
        # Simplified: check if user's org matches (if step's workflow is org-bound)
        # try:
        #     profile = Profile.objects.get(user=user)
        #     if hasattr(workflow_step.workflow, 'organization') and workflow_step.workflow.organization and workflow_step.workflow.organization != profile.organization:
        #         raise PermissionDenied("Acesso negado às entradas de tempo deste passo.")
        #     # Further checks if user is not admin (e.g., only if they are assigned to tasks on this step)
        # except Profile.DoesNotExist:
        #     if not user.is_superuser: raise PermissionDenied("Acesso negado.")


        time_entries_qs = TimeEntry.objects.filter(
            workflow_step=workflow_step
        ).select_related(
            'user__profile', 'task__client', 'client' # Added client direct from TimeEntry
        ).order_by('-date', '-created_at')
        
        # Further filter time_entries_qs based on user's visibility of tasks/clients if not admin
        # ...
        
        serializer = TimeEntrySerializer(time_entries_qs, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def current_tasks(self, request, pk=None):
        workflow_step = self.get_object()
        user = request.user
        
        current_tasks_qs = Task.objects.filter(
            current_workflow_step=workflow_step
        ).select_related(
            'client__organization', 'assigned_to__profile', 'created_by__profile'
        )
        
        if not user.is_superuser:
            try:
                profile = Profile.objects.get(user=user)
                if not profile.organization: return Task.objects.none()
                
                current_tasks_qs = current_tasks_qs.filter(client__organization=profile.organization)
                
                if not (profile.is_org_admin or profile.can_view_all_tasks):
                    visible_client_ids = list(profile.visible_clients.values_list('id', flat=True))
                    current_tasks_qs = current_tasks_qs.filter(
                        Q(client_id__in=visible_client_ids) | Q(assigned_to=user)
                    ).distinct()
            except Profile.DoesNotExist:
                 return Task.objects.none()
        
        serializer = TaskSerializer(current_tasks_qs.distinct(), many=True)
        return Response(serializer.data)
        
class OrganizationViewSet(viewsets.ModelViewSet):
    serializer_class = OrganizationSerializer
    permission_classes = [IsAuthenticated, IsOrgAdmin]
    
    def get_queryset(self):
        """
        Returns the organization(s) the user is part of.
        An admin sees their org, a superuser sees all. A user with no org sees none.
        """
        user = self.request.user
        
        # Annotate counts for efficiency in the serializer
        base_queryset = Organization.objects.annotate(
            member_count=Count('members', distinct=True),
            client_count=Count('clients', distinct=True)
        )

        if user.is_superuser:
            return base_queryset
            
        try:
            profile = user.profile
            if profile.organization:
                return base_queryset.filter(id=profile.organization.id)
            return Organization.objects.none()
        except Profile.DoesNotExist:
            return Organization.objects.none()

    def perform_create(self, serializer):
        """
        Overrides default behavior to allow a user without an organization to create one
        and become its first administrator.
        """
        user = self.request.user
        
        try:
            profile = Profile.objects.get(user=user)
            if profile.organization:
                raise PermissionDenied("Você já pertence a uma organização. Não pode criar uma nova.")
        except Profile.DoesNotExist:
            # If the user somehow has no profile, create one.
            profile = Profile.objects.create(user=user)
        
        # Create the organization with the provided data
        organization = serializer.save()
        
        # Assign the new organization to the user's profile and make them an admin
        profile.organization = organization
        profile.is_org_admin = True
        profile.role = 'Administrador'
        
        # Grant all permissions to the first admin
        for field in profile._meta.fields:
            if field.name.startswith('can_'):
                setattr(profile, field.name, True)
        
        profile.save()
        
        logger.info(f"Organização '{organization.name}' criada por {user.username}. Usuário definido como administrador.")
        # Log organization action
        log_organization_action(
            self.request,
            action_type='CREATE_ORGANIZATION',
            action_description=f"Organização criada: {organization.name} (ID: {organization.id})",
            related_object=organization
        )

    def get_permissions(self):
        """
        Instantiate and return the list of permissions that this view requires.
        Override to allow any authenticated user to `create` (if they have no org),
        while other actions are protected by `IsOrgAdmin`.
        """
        if self.action == 'create':
            # Anyone authenticated can attempt to create, the logic is in `perform_create`.
            return [IsAuthenticated()]
        # For all other actions (list, retrieve, update, destroy, custom actions),
        # use the default permissions defined in `permission_classes`.
        return super().get_permissions()

    # NOTE: The 'update' and 'destroy' methods are now inherited from ModelViewSet.
    # We no longer need to override them just for permission checks, because the
    # `IsOrgAdmin` class handles it automatically via `has_object_permission`.

    @action(detail=True, methods=['get'])
    def members(self, request, pk=None):
        """
        Lists members of a specific organization.
        The `IsOrgAdmin` check on `get_object()` ensures only admins of this org can access.
        """
        organization = self.get_object() # This triggers the `has_object_permission` check
        members_profiles = Profile.objects.filter(organization=organization).select_related('user')
        serializer = ProfileSerializer(members_profiles, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def clients(self, request, pk=None):
        """
        Lists clients of a specific organization.
        """
        organization = self.get_object() # Triggers permission check
        clients_qs = Client.objects.filter(organization=organization).select_related('account_manager')
        serializer = ClientSerializer(clients_qs, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def add_member_by_code(self, request, pk=None):
        """
        Adds a new member to the organization using an invitation code.
        """
        organization = self.get_object() # Triggers permission check
        
        invitation_code = request.data.get('invitation_code')
        if not invitation_code:
            return Response({"error": "Código de convite é obrigatório"}, status=status.HTTP_400_BAD_REQUEST)
        
        profile_to_add = Profile.find_by_invitation_code(invitation_code)
        if not profile_to_add:
            return Response({"error": "Código de convite inválido ou não encontrado"}, status=status.HTTP_404_NOT_FOUND)
        if profile_to_add.organization:
            return Response({"error": "Este utilizador já pertence a uma organização"}, status=status.HTTP_400_BAD_REQUEST)
            
        # Assign organization and base role
        profile_to_add.organization = organization
        profile_to_add.role = request.data.get('role', 'Colaborador')
        
        # Apply permissions from request
        profile_to_add.is_org_admin = request.data.get('is_admin', False)
        # Add all other 'can_' flags from the request
        for field in Profile._meta.fields:
            if field.name.startswith('can_') and field.name in request.data:
                setattr(profile_to_add, field.name, request.data[field.name])

        profile_to_add.save()
        # Log organization action
        log_organization_action(
            request,
            action_type='ADD_MEMBER_BY_CODE',
            action_description=f"Membro adicionado à organização {organization.name} (ID: {organization.id}) via código de convite: {profile_to_add.user.username}",
            related_object=organization
        )
        # Notify the new member
        NotificationService.create_notification(
            user=profile_to_add.user,
            notification_type='added_to_organization',
            title=f"Bem-vindo à organização {organization.name}",
            message=f"Você foi adicionado à organização '{organization.name}' como {profile_to_add.role}.",
            created_by=request.user
        )
        # Notify all org admins (except the new member)
        org_admins = Profile.objects.filter(organization=organization, is_org_admin=True).exclude(user=profile_to_add.user).select_related('user')
        for admin_profile in org_admins:
            NotificationService.create_notification(
                user=admin_profile.user,
                notification_type='organization_member_added',
                title=f"Novo membro adicionado",
                message=f"{profile_to_add.user.username} foi adicionado à organização como {profile_to_add.role}.",
                created_by=request.user
            )
        return Response(ProfileSerializer(profile_to_add, context={'request': request}).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def remove_member(self, request, pk=None):
        """
        Removes a member from the organization.
        """
        organization = self.get_object() # Triggers permission check
        
        user_id_to_remove = request.data.get('user_id')
        if not user_id_to_remove:
            return Response({"error": "ID do usuário é obrigatório"}, status=status.HTTP_400_BAD_REQUEST)
        
        if str(request.user.id) == str(user_id_to_remove):
            return Response({"error": "Não pode remover-se a si próprio."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            profile_to_remove = Profile.objects.get(user_id=user_id_to_remove, organization=organization)
            
            # Prevent removing the last admin
            if profile_to_remove.is_org_admin:
                admin_count = Profile.objects.filter(organization=organization, is_org_admin=True).count()
                if admin_count <= 1:
                    return Response({"error": "Não é possível remover o único administrador."}, status=status.HTTP_400_BAD_REQUEST)
            
            # Disassociate from org and reset permissions
            profile_to_remove.organization = None
            for field in profile_to_remove._meta.fields:
                if field.name.startswith('can_') or field.name == 'is_org_admin':
                    setattr(profile_to_remove, field.name, False)
            profile_to_remove.visible_clients.clear()
            profile_to_remove.save()
            # Log organization action
            log_organization_action(
                request,
                action_type='REMOVE_MEMBER',
                action_description=f"Membro removido da organização {organization.name} (ID: {organization.id}): {profile_to_remove.user.username}",
                related_object=organization
            )
            # Notify the removed member
            NotificationService.create_notification(
                user=profile_to_remove.user,
                notification_type='removed_from_organization',
                title=f"Removido da organização {organization.name}",
                message=f"Você foi removido da organização '{organization.name}'. Se acredita que isso foi um erro, contate um administrador.",
                created_by=request.user
            )
            # Notify all org admins (except the removed member)
            org_admins = Profile.objects.filter(organization=organization, is_org_admin=True).exclude(user=profile_to_remove.user).select_related('user')
            for admin_profile in org_admins:
                NotificationService.create_notification(
                    user=admin_profile.user,
                    notification_type='organization_member_removed',
                    title=f"Membro removido",
                    message=f"{profile_to_remove.user.username} foi removido da organização.",
                    created_by=request.user
                )
            return Response({"success": "Membro removido da organização."}, status=status.HTTP_200_OK)
        except Profile.DoesNotExist:
            return Response({"error": "Perfil não encontrado nesta organização."}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def update_member(self, request, pk=None):
        """
        Updates a member's role and permissions within the organization.
        """
        organization = self.get_object() # Triggers permission check
        
        user_id_to_update = request.data.get('user_id') or request.query_params.get('user_id')
        if not user_id_to_update:
            return Response({"error": "ID do usuário é obrigatório"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            profile_to_update = Profile.objects.get(user_id=user_id_to_update, organization=organization)
            
            update_data = request.data.copy()
            if 'user_id' in update_data: del update_data['user_id']
            
            m2m_fields = ['visible_clients']
            
            for field_name, field_value in update_data.items():
                if hasattr(profile_to_update, field_name) and field_name not in m2m_fields:
                    setattr(profile_to_update, field_name, field_value)
            
            profile_to_update.save()
            
            if 'visible_clients' in request.data:
                visible_client_ids = request.data.get('visible_clients', [])
                if not profile_to_update.can_view_all_clients:
                    valid_clients = Client.objects.filter(id__in=visible_client_ids, organization=organization)
                    profile_to_update.visible_clients.set(valid_clients)
                else:
                    profile_to_update.visible_clients.clear()
            # Log organization action
            log_organization_action(
                request,
                action_type='UPDATE_MEMBER',
                action_description=f"Permissões/atribuições de membro atualizadas na organização {organization.name} (ID: {organization.id}): {profile_to_update.user.username}",
                related_object=organization
            )
            # Notify the updated member
            NotificationService.create_notification(
                user=profile_to_update.user,
                notification_type='organization_member_updated',
                title=f"Permissões atualizadas na organização {organization.name}",
                message=f"Suas permissões ou atribuições na organização '{organization.name}' foram atualizadas.",
                created_by=request.user
            )
            return Response(ProfileSerializer(profile_to_update, context={'request': request}).data, status=status.HTTP_200_OK)
            
        except Profile.DoesNotExist:
            return Response({"error": "Perfil não encontrado nesta organização."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error updating member: {str(e)}")
            return Response({"error": f"Erro ao atualizar membro: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_summary(request):
        user = request.user
        
        # Create a unique cache key for this specific user's dashboard
        cache_key = f'dashboard_summary_{user.id}'
        
        # 1. Try to get data from the cache
        cached_data = cache.get(cache_key)
        if cached_data:
            logger.info(f"Serving dashboard summary for user {user.username} from CACHE.")
            return Response(cached_data)

        logger.info(f"Cache miss for dashboard summary for user {user.username}. Computing now.")
        
        try:
            profile = Profile.objects.select_related('organization').get(user=user)
            org_id = profile.organization_id if profile.organization else None

            if not org_id:
                return Response({'error': 'Utilizador não associado a uma organização.'}, status=400)

            today = timezone.now().date()
            seven_days_ago = today - timedelta(days=7)

            tasks_qs = Task.objects.for_user(user)
            time_entries_qs = TimeEntry.objects.for_user(user)

            task_stats = tasks_qs.aggregate(
                active_tasks=Count('id', filter=Q(status__in=['pending', 'in_progress'])),
                overdue_tasks=Count('id', filter=Q(deadline__lt=today, status__in=['pending', 'in_progress'])),
                today_tasks=Count('id', filter=Q(deadline=today, status__in=['pending', 'in_progress'])),
                completed_tasks_week=Count('id', filter=Q(status='completed', completed_at__date__gte=seven_days_ago))
            )
            
            time_stats = time_entries_qs.aggregate(
                time_tracked_today=Sum('minutes_spent', filter=Q(date=today)),
                time_tracked_week=Sum('minutes_spent', filter=Q(date__gte=seven_days_ago))
            )
            
            clients_qs = Client.objects.for_user(user)
            client_stats = clients_qs.aggregate(
                active_clients=Count('id', filter=Q(is_active=True))
            )
            
            response_data = {
                'permissions': ProfileSerializer(profile).data,
                'active_tasks': task_stats.get('active_tasks', 0) or 0,
                'overdue_tasks': task_stats.get('overdue_tasks', 0) or 0,
                'today_tasks': task_stats.get('today_tasks', 0) or 0,
                'completed_tasks_week': task_stats.get('completed_tasks_week', 0) or 0,
                'time_tracked_today': time_stats.get('time_tracked_today', 0) or 0,
                'time_tracked_week': time_stats.get('time_tracked_week', 0) or 0,
                'active_clients': client_stats.get('active_clients', 0) or 0,
            }

            if profile.is_org_admin or profile.can_view_organization_profitability:
                # ... (your existing profitability logic)
                profit_stats = ClientProfitability.objects.filter(
                    client__organization_id=org_id, 
                    year=today.year, 
                    month=today.month
                ).aggregate(
                    unprofitable_count=Count('id', filter=Q(is_profitable=False)),
                    avg_margin=Avg('profit_margin')
                )
                response_data['unprofitable_clients'] = profit_stats.get('unprofitable_count', 0) or 0
                response_data['average_profit_margin'] = profit_stats.get('avg_margin', 0) or 0

            if profile.is_org_admin or profile.can_view_all_tasks:
                # ... (your existing approval logic)
                response_data['tasks_needing_approval'] = tasks_qs.filter(
                    current_workflow_step__requires_approval=True
                ).exclude(
                    approvals__workflow_step=F('current_workflow_step'),
                    approvals__approved=True
                ).distinct().count()

            # 2. Store the computed data in the cache before returning it.
            #    Timeout is in seconds. 300 seconds = 5 minutes.
            cache.set(cache_key, response_data, timeout=300)
            
            return Response(response_data)
            
        except Profile.DoesNotExist:
            # ... (your existing exception handling)
            if user.is_superuser:
                return Response({'message': 'Superuser dashboard not implemented yet.'})
            return Response({'error': 'User profile not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Erro no dashboard_summary: {e}", exc_info=True)
            return Response({'error': 'Erro interno ao carregar o sumário do dashboard.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 
    
class WorkflowNotificationViewSet(viewsets.ModelViewSet):
    serializer_class = WorkflowNotificationSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        # OPTIMIZATION
        base_queryset = WorkflowNotification.objects.filter(
            user=user
        ).select_related(
            'task__client', 
            'workflow_step', 
            'created_by'
        ).order_by('-created_at')
        
        is_read_param = self.request.query_params.get('is_read')
        if is_read_param is not None:
            base_queryset = base_queryset.filter(is_read=(is_read_param.lower() == 'true'))

        notification_type = self.request.query_params.get('type')
        if notification_type:
            base_queryset = base_queryset.filter(notification_type=notification_type)
            
        priority = self.request.query_params.get('priority')
        if priority:
            base_queryset = base_queryset.filter(priority=priority)
            
        is_archived_str = self.request.query_params.get('is_archived', 'false')
        is_archived = is_archived_str.lower() == 'true'
        base_queryset = base_queryset.filter(is_archived=is_archived)
        
        limit_str = self.request.query_params.get('limit')
        if limit_str:
            try:
                limit = int(limit_str)
                if limit > 0:
                    base_queryset = base_queryset[:limit]
            except ValueError:
                pass
                
        return base_queryset
    
    @action(detail=False, methods=['get'])
    def summary_stats(self, request):
        """Resumo rápido de estatísticas"""
        user = request.user
        
        # Últimos 7 dias
        week_ago = timezone.now() - timedelta(days=7)
        
        notifications = WorkflowNotification.objects.filter(
            user=user,
            created_at__gte=week_ago,
            is_archived=False
        )
        
        stats = {
            'total_this_week': notifications.count(),
            'unread_this_week': notifications.filter(is_read=False).count(),
            'urgent_unread': notifications.filter(
                is_read=False, 
                priority='urgent'
            ).count(),
            'most_frequent_type': notifications.values('notification_type').annotate(
                count=Count('id')
            ).order_by('-count').first(),
            'oldest_unread_days': None,
        }
        
        # Notificação não lida mais antiga
        oldest_unread = notifications.filter(is_read=False).order_by('created_at').first()
        if oldest_unread:
            days_old = (timezone.now() - oldest_unread.created_at).days
            stats['oldest_unread_days'] = days_old
        
        return Response(stats)
    
    @action(detail=True, methods=['post'])
    def mark_as_read(self, request, pk=None):
        notification = self.get_object()
        if notification.user != request.user: # Ensure user can only modify their own notifications
            raise PermissionDenied("Você não pode modificar notificações de outro usuário.")
        notification.mark_as_read()
        return Response({'status': 'marked_as_read'})
    
    @action(detail=True, methods=['post'])
    def mark_as_unread(self, request, pk=None):
        notification = self.get_object()
        if notification.user != request.user:
            raise PermissionDenied("Você não pode modificar notificações de outro usuário.")
        notification.mark_as_unread()
        return Response({'status': 'marked_as_unread'})
    
    @action(detail=False, methods=['post'])
    def mark_all_as_read(self, request):
        count = NotificationService.mark_all_as_read(request.user)
        return Response({'status': 'marked_as_read', 'count': count})
    
    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        count = NotificationService.get_unread_count(request.user)
        return Response({'unread_count': count})
    
    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        notification = self.get_object()
        if notification.user != request.user:
            raise PermissionDenied("Você não pode modificar notificações de outro usuário.")
        notification.archive()
        return Response({'status': 'archived'})
    
    @action(detail=False, methods=['post'])
    def create_manual_reminder(self, request):
        task_id = request.data.get('task_id')
        user_ids_str = request.data.get('user_ids', []) 
        user_ids = []
        if isinstance(user_ids_str, str): 
            user_ids = [uid.strip() for uid in user_ids_str.split(',') if uid.strip()]
        elif isinstance(user_ids_str, list):
            user_ids = user_ids_str
        
        title = request.data.get('title')
        message = request.data.get('message')
        priority = request.data.get('priority', 'normal')
        scheduled_for_str = request.data.get('scheduled_for')
        
        if not all([task_id, user_ids, title, message]):
            return Response({'error': 'task_id, user_ids, title e message são obrigatórios'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            task = Task.objects.get(id=task_id)
            target_users = User.objects.filter(id__in=user_ids, is_active=True)
            
            if not target_users.exists():
                 return Response({'error': 'Nenhum usuário alvo válido encontrado.'}, status=status.HTTP_400_BAD_REQUEST)

            profile = Profile.objects.get(user=request.user)
            
            can_remind = (
                profile.is_org_admin or 
                profile.can_edit_all_tasks or
                (profile.can_edit_assigned_tasks and task.assigned_to == request.user) or
                (task.current_workflow_step and task.current_workflow_step.assign_to == request.user)
            )
            if not can_remind:
                return Response({'error': 'Sem permissão para criar lembretes para esta tarefa'}, status=status.HTTP_403_FORBIDDEN)
            
            scheduled_for = None
            if scheduled_for_str:
                try:
                    # Ensure timezone-aware datetime objects if your DATETIME_INPUT_FORMATS support 'Z' or '+HH:MM'
                    # For simplicity, assuming ISO format possibly with 'Z'
                    scheduled_for = timezone.datetime.fromisoformat(scheduled_for_str.replace('Z', '+00:00'))
                    if scheduled_for <= timezone.now(): # Ensure it's in the future
                        return Response({'error': 'Data agendada deve ser no futuro.'}, status=status.HTTP_400_BAD_REQUEST)
                except ValueError:
                    return Response({'error': 'Formato de data inválido para scheduled_for. Use ISO 8601 (e.g., YYYY-MM-DDTHH:MM:SSZ).'}, status=status.HTTP_400_BAD_REQUEST)
            
            notifications = NotificationService.create_manual_reminder(
                task=task, target_users=target_users, title=title, message=message,
                created_by=request.user, priority=priority, scheduled_for=scheduled_for
            )
            # Log organization action
            log_organization_action(
                request,
                action_type='CREATE_MANUAL_REMINDER',
                action_description=f"Lembrete manual criado para tarefa {task.title} (ID: {task.id}) para usuários: {[u.username for u in target_users]}",
                related_object=task
            )
            return Response({
                'status': 'created', 
                'count': len(notifications), 
                'notifications': [n.id for n in notifications]
            }, status=status.HTTP_201_CREATED)
            
        except Task.DoesNotExist:
            return Response({'error': 'Tarefa não encontrada'}, status=status.HTTP_404_NOT_FOUND)
        except Profile.DoesNotExist:
            return Response({'error': 'Perfil do solicitante não encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Erro ao criar lembrete manual: {e}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class WorkflowHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = WorkflowHistorySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        # OPTIMIZATION
        base_queryset = WorkflowHistory.objects.select_related(
            'task__client', 'from_step', 'to_step', 'changed_by'
        ).order_by('-created_at')
        
        task_id = self.request.query_params.get('task')
        if task_id:
            base_queryset = base_queryset.filter(task_id=task_id)
        
        if user.is_superuser:
            return base_queryset

        try:
            profile = user.profile
            if not profile.organization:
                return WorkflowHistory.objects.none()

            org_filter = Q(task__client__organization=profile.organization)
            
            if profile.is_org_admin or profile.can_view_all_tasks:
                return base_queryset.filter(org_filter)
            else:
                visible_client_ids = profile.visible_clients.values_list('id', flat=True)
                user_access_filter = Q(task__client_id__in=visible_client_ids) | Q(task__assigned_to=user)
                return base_queryset.filter(org_filter & user_access_filter).distinct()
                
        except Profile.DoesNotExist:
            return WorkflowHistory.objects.none()

class WorkflowStepDetailViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = WorkflowStepSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = WorkflowStep.objects.select_related(
            'workflow', 'assign_to__profile' # for assign_to_name in serializer
        ).prefetch_related(
            'time_entries__user__profile', 
            'task_approvals__approved_by__profile' 
        )
        
        workflow_id_param = self.request.query_params.get('workflow')
        if workflow_id_param:
            queryset = queryset.filter(workflow_id=workflow_id_param)
        
        if not user.is_superuser:
            try:
                profile = Profile.objects.get(user=user)
                if not profile.organization: 
                    return WorkflowStep.objects.none()
                
                # Assuming workflows are not directly organization-bound in the model for now.
                # If they were: queryset = queryset.filter(workflow__organization=profile.organization)
                # Further, filter based on whether user can see the workflows themselves.
                # This is a simplified permission for listing.
                if not (profile.is_org_admin or profile.can_view_all_tasks or profile.can_edit_workflows):
                    # Non-admins might only see steps of workflows assigned to tasks they can access
                    # or steps they are directly assigned to. This is complex for a general list.
                    # For now, a broad permission is assumed for listing.
                    # A more granular approach would filter based on tasks using these workflow steps.
                    # Example: accessible_workflow_ids = Task.objects.filter(...user access...).values_list('workflow_id', flat=True).distinct()
                    # queryset = queryset.filter(workflow_id__in=accessible_workflow_ids)
                    pass # No further specific filtering for listing steps if user has basic task/workflow view perms
            except Profile.DoesNotExist:
                return WorkflowStep.objects.none()
            
        return queryset.order_by('workflow__name', 'order')
    
    @action(detail=True, methods=['get'])
    def time_entries(self, request, pk=None):
        workflow_step = self.get_object() 
        user = request.user
        
        time_entries_qs = TimeEntry.objects.filter(
            workflow_step=workflow_step
        ).select_related(
            'user__profile', 'task__client', 'client'
        ).order_by('-date', '-created_at')
        
        # Permission check for viewing these time entries
        if not user.is_superuser:
            try:
                profile = Profile.objects.get(user=user)
                if not profile.organization: return TimeEntry.objects.none()
                # Filter time entries by the user's organization (via client)
                time_entries_qs = time_entries_qs.filter(client__organization=profile.organization)

                # If user is not admin/team_time_viewer, filter to their own or visible client entries
                if not (profile.is_org_admin or profile.can_view_team_time):
                    visible_client_ids = list(profile.visible_clients.values_list('id', flat=True))
                    time_entries_qs = time_entries_qs.filter(
                        Q(user=user) | Q(client_id__in=visible_client_ids)
                    )
            except Profile.DoesNotExist:
                return TimeEntry.objects.none()
        
        serializer = TimeEntrySerializer(time_entries_qs.distinct(), many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def current_tasks(self, request, pk=None):
        workflow_step = self.get_object()
        user = request.user
        
        current_tasks_qs = Task.objects.filter(
            current_workflow_step=workflow_step
        ).select_related(
            'client__organization', 'assigned_to__profile', 'created_by__profile'
        )
        
        if not user.is_superuser:
            try:
                profile = Profile.objects.get(user=user)
                if not profile.organization: return Task.objects.none()
                
                current_tasks_qs = current_tasks_qs.filter(client__organization=profile.organization)
                
                if not (profile.is_org_admin or profile.can_view_all_tasks):
                    visible_client_ids = list(profile.visible_clients.values_list('id', flat=True))
                    current_tasks_qs = current_tasks_qs.filter(
                        Q(client_id__in=visible_client_ids) | Q(assigned_to=user)
                    ).distinct()
            except Profile.DoesNotExist:
                 return Task.objects.none()
        
        serializer = TaskSerializer(current_tasks_qs.distinct(), many=True)
        return Response(serializer.data)

# --- (As definições de OrganizationViewSet, dashboard_summary, e os novos endpoints de verificação já devem estar no seu views.py) ---
# --- Certifique-se que o resto do seu views.py está correto e completo. ---
# --- O código abaixo são os novos endpoints de verificação, caso precise deles aqui novamente. ---

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def check_deadlines_and_notify_view(request):
    try:
        profile = Profile.objects.select_related('organization').get(user=request.user)
        if not (profile.is_org_admin): 
            return Response({'error': 'Apenas administradores podem executar esta verificação.'}, status=status.HTTP_403_FORBIDDEN)
        
        organization = profile.organization
        if not organization:
            return Response({'error': 'Administrador não associado a uma organização.'}, status=status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        thresholds_days = [3, 1, 0] 
        notifications_created_total = 0
        tasks_checked_total = 0

        for days_ahead in thresholds_days:
            target_deadline_date = (now + timedelta(days=days_ahead)).date()
            
            tasks_with_near_deadline = Task.objects.filter(
                deadline=target_deadline_date,
                status__in=['pending', 'in_progress'],
                client__organization=organization 
            ).select_related('assigned_to', 'current_workflow_step__assign_to', 'created_by', 'client') 
            
            tasks_checked_total += tasks_with_near_deadline.count()

            for task in tasks_with_near_deadline:
                notifications = NotificationService.notify_deadline_approaching(task, days_ahead)
                notifications_created_total += len(notifications)
        
        # Log organization action
        log_organization_action(
            request,
            action_type='CHECK_DEADLINES_AND_NOTIFY',
            action_description=f"Verificação de deadlines e notificações executada para organização {organization.name} (ID: {organization.id}) - {tasks_checked_total} tarefas verificadas, {notifications_created_total} notificações criadas.",
            related_object=organization
        )
        return Response({
            'success': True,
            'message': f'Verificação de deadlines concluída para a organização {organization.name}.',
            'tasks_checked': tasks_checked_total,
            'notifications_created': notifications_created_total
        })
        
    except Profile.DoesNotExist:
        return Response({'error': 'Perfil não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Erro ao verificar deadlines: {str(e)}")
        return Response({'error': f'Erro interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def check_overdue_steps_and_notify_view(request):
    try:
        profile = Profile.objects.select_related('organization').get(user=request.user)
        if not profile.is_org_admin:
            return Response({'error': 'Apenas administradores podem executar esta verificação.'}, status=status.HTTP_403_FORBIDDEN)
        
        organization = profile.organization
        if not organization:
            return Response({'error': 'Administrador não associado a uma organização.'}, status=status.HTTP_400_BAD_REQUEST)

        overdue_threshold_days = int(request.data.get('overdue_threshold_days', 5)) 
        if overdue_threshold_days <= 0:
            return Response({'error': 'overdue_threshold_days deve ser positivo.'}, status=status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        
        notifications_created_total = 0
        tasks_processed_total = 0

        active_workflow_tasks = Task.objects.filter(
            client__organization=organization,
            status__in=['pending', 'in_progress'],
            workflow__isnull=False,
            current_workflow_step__isnull=False
        ).select_related('current_workflow_step', 'current_workflow_step__assign_to', 
                         'assigned_to', 'created_by', 'client', 'client__account_manager') 

        for task in active_workflow_tasks:
            tasks_processed_total += 1
            last_significant_history = WorkflowHistory.objects.filter(
                task=task, 
                to_step=task.current_workflow_step, 
                action__in=['step_advanced', 'workflow_assigned'] 
            ).order_by('-created_at').first()

            step_became_current_at = task.updated_at 
            if last_significant_history:
                step_became_current_at = last_significant_history.created_at
            
            days_on_current_step = (now - step_became_current_at).days

            if days_on_current_step >= overdue_threshold_days:
                notifications = NotificationService.notify_step_overdue(task, task.current_workflow_step, days_on_current_step)
                notifications_created_total += len(notifications)
        
        # Log organization action
        log_organization_action(
            request,
            action_type='CHECK_OVERDUE_STEPS_AND_NOTIFY',
            action_description=f"Verificação de passos atrasados executada para organização {organization.name} (ID: {organization.id}) - {tasks_processed_total} tarefas processadas, {notifications_created_total} notificações criadas.",
            related_object=organization
        )
        return Response({
            'success': True,
            'message': f'Verificação de passos atrasados ({overdue_threshold_days} dias) concluída para {organization.name}.',
            'tasks_with_active_workflow_step_checked': tasks_processed_total,
            'overdue_step_notifications_created': notifications_created_total
        })

    except Profile.DoesNotExist:
        return Response({'error': 'Perfil não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Erro ao verificar passos atrasados: {str(e)}")
        return Response({'error': f'Erro interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def check_pending_approvals_and_notify_view(request):
    try:
        profile = Profile.objects.select_related('organization').get(user=request.user)
        if not profile.is_org_admin:
            return Response({'error': 'Apenas administradores podem executar esta verificação.'}, status=status.HTTP_403_FORBIDDEN)

        organization = profile.organization
        if not organization:
            return Response({'error': 'Administrador não associado a uma organização.'}, status=status.HTTP_400_BAD_REQUEST)

        reminder_threshold_days = int(request.data.get('reminder_threshold_days', 2))
        if reminder_threshold_days <= 0:
            return Response({'error': 'reminder_threshold_days deve ser positivo.'}, status=status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        
        tasks_needing_approval_reminder = Task.objects.filter(
            client__organization=organization,
            status__in=['pending', 'in_progress'],
            current_workflow_step__requires_approval=True
        ).exclude( 
            id__in=TaskApproval.objects.filter(
                task_id=models.F('task_id'), 
                workflow_step=models.F('task__current_workflow_step'), 
                approved=True
            ).values_list('task_id', flat=True)
        ).select_related('current_workflow_step', 'client', 'client__organization')
        
        notifications_sent_total = 0
        tasks_checked_count = 0

        for task in tasks_needing_approval_reminder:
            tasks_checked_count +=1
            
            step_became_current_history = WorkflowHistory.objects.filter(
                task=task, to_step=task.current_workflow_step,
                action__in=['step_advanced', 'workflow_assigned']
            ).order_by('-created_at').first()

            if step_became_current_history:
                days_pending_approval = (now - step_became_current_history.created_at).days
                if days_pending_approval >= reminder_threshold_days:
                    reminders = NotificationService.notify_approval_needed(
                        task, task.current_workflow_step, 
                        approvers=None, 
                        is_reminder=True
                    )
                    notifications_sent_total += len(reminders)
        
        # Log organization action
        log_organization_action(
            request,
            action_type='CHECK_PENDING_APPROVALS_AND_NOTIFY',
            action_description=f"Verificação de aprovações pendentes executada para organização {organization.name} (ID: {organization.id}) - {tasks_checked_count} tarefas verificadas, {notifications_sent_total} notificações enviadas.",
            related_object=organization
        )
        return Response({
            'success': True,
            'message': f'Verificação de aprovações pendentes ({reminder_threshold_days} dias) concluída para {organization.name}.',
            'tasks_checked_for_pending_approval': tasks_checked_count,
            'approval_reminder_notifications_sent': notifications_sent_total
        })
        
    except Profile.DoesNotExist:
        return Response({'error': 'Perfil não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Erro ao verificar aprovações pendentes: {str(e)}")
        return Response({'error': f'Erro interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_organization_profitability(request):
    """
    Triggers an asynchronous Celery task to update client profitability data 
    for the user's organization for the current and specified number of past months.
    """
    try:
        profile = Profile.objects.select_related('organization').get(user=request.user)
        
        if not profile.organization:
            return Response(
                {'error': 'Usuário não pertence a nenhuma organização'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not (profile.is_org_admin or profile.can_view_organization_profitability): # Or a more specific "can_trigger_recalculation"
            return Response(
                {'error': 'Sem permissão para iniciar o recálculo de rentabilidade da organização'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        months_back_str = request.data.get('months_back', '1') # Default to current month only if 0, or current + 1 prev if 1.
                                                              # Let's adjust to '1' meaning current month only for this endpoint.
        try:
            # months_back = 1 means current month
            # months_back = 2 means current month and previous month
            months_back = int(months_back_str)
            if not (1 <= months_back <= 12): # Max 12 months back for a manual trigger
                raise ValueError("O parâmetro 'months_back' deve estar entre 1 e 12.")
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        organization = profile.organization
        now = timezone.now()
        
        periods_to_update = []
        for i in range(months_back): # months_back=1 -> i=0 (current month)
                                   # months_back=2 -> i=0 (current), i=1 (prev)
            target_date_for_period = now - relativedelta(months=i)
            year_to_update = target_date_for_period.year
            month_to_update = target_date_for_period.month
            periods_to_update.append((year_to_update, month_to_update))
            
        # Dispatch the Celery task
        task_result = update_profitability_for_single_organization_task.delay(
            organization.id, 
            periods_to_update
        )
        
        logger.info(f"Dispatched profitability update task {task_result.id} for organization {organization.name} for periods: {periods_to_update}.")
        
        return Response({
            'success': True,
            'message': f'Recálculo de rentabilidade para a organização {organization.name} foi iniciado em segundo plano. Os dados serão atualizados em breve.',
            'task_id': task_result.id, # You can optionally return the Celery task ID
            'organization': organization.name,
            'periods_being_processed': periods_to_update,
            'processed_at': timezone.now().isoformat()
        }, status=status.HTTP_202_ACCEPTED) # HTTP 202 Accepted indicates the request is accepted for processing
        
    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil de usuário não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Organization.DoesNotExist: # Should not happen if profile.organization is enforced
        return Response(
            {'error': 'Organização não encontrada'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Erro ao despachar tarefa de atualização de rentabilidade: {str(e)}", exc_info=True)
        return Response(
            {'error': f'Erro interno ao processar a requisição: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
class NotificationSettingsViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSettingsSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return NotificationSettings.objects.filter(user=self.request.user)
    
    def get_object(self):
        """Retorna ou cria as configurações do usuário"""
        settings, created = NotificationSettings.objects.get_or_create(
            user=self.request.user,
            defaults={
                'deadline_days_notice': [3, 1, 0],
                'digest_time': '09:00',
            }
        )
        return settings
    
    @action(detail=False, methods=['get'])
    def my_settings(self, request):
        """Endpoint para obter configurações do usuário logado"""
        settings = self.get_object()
        serializer = self.get_serializer(settings)
        return Response(serializer.data)
    
    @action(detail=False, methods=['patch'])
    def update_settings(self, request):
        """Endpoint para atualizar configurações"""
        settings = self.get_object()
        serializer = self.get_serializer(settings, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def reset_to_defaults(self, request):
        """Reset para configurações padrão"""
        settings = self.get_object()
        
        # Reset para valores padrão
        settings.email_notifications_enabled = True
        settings.push_notifications_enabled = True
        settings.notify_step_ready = True
        settings.notify_step_completed = True
        settings.notify_approval_needed = True
        settings.notify_approval_completed = True
        settings.notify_workflow_completed = True
        settings.notify_deadline_approaching = True
        settings.notify_step_overdue = True
        settings.notify_workflow_assigned = True
        settings.notify_step_rejected = True
        settings.notify_manual_reminders = True
        settings.digest_frequency = 'immediate'
        settings.digest_time = timezone.datetime.strptime('09:00', '%H:%M').time()
        settings.deadline_days_notice = [3, 1, 0]
        settings.overdue_threshold_days = 5
        settings.approval_reminder_days = 2
        settings.quiet_start_time = None
        settings.quiet_end_time = None
        
        settings.save()
        
        serializer = self.get_serializer(settings)
        return Response({
            'message': 'Configurações resetadas para o padrão',
            'settings': serializer.data
        })
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notification_stats(request):
    """Estatísticas de notificações do usuário"""
    days = int(request.GET.get('days', 30))
    
    try:
        stats = NotificationMetricsService.get_user_notification_stats(
            request.user, days
        )
        return Response(stats)
    except Exception as e:
        logger.error(f"Erro ao calcular estatísticas de notificação: {e}")
        return Response(
            {'error': 'Erro ao calcular estatísticas'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def organization_notification_stats(request):
    """Estatísticas de notificações da organização (apenas admins)"""
    try:
        profile = Profile.objects.get(user=request.user)
        
        if not profile.is_org_admin:
            return Response(
                {'error': 'Apenas administradores podem ver estatísticas da organização'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        days = int(request.GET.get('days', 30))
        stats = NotificationMetricsService.get_organization_notification_stats(
            profile.organization, days
        )
        return Response(stats)
        
    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Erro ao calcular estatísticas da organização: {e}")
        return Response(
            {'error': 'Erro ao calcular estatísticas'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def workflow_notification_performance(request):
    """Performance de notificações de workflow"""
    try:
        profile = Profile.objects.get(user=request.user)
        
        if not (profile.is_org_admin or profile.can_view_analytics):
            return Response(
                {'error': 'Sem permissão para ver métricas de workflow'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        workflow_id = request.GET.get('workflow_id')
        days = int(request.GET.get('days', 30))
        
        stats = NotificationMetricsService.get_workflow_notification_performance(
            workflow_id, days
        )
        return Response(stats)
        
    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Erro ao calcular performance de workflow: {e}")
        return Response(
            {'error': 'Erro ao calcular métricas'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notification_reports(request):
    """Gera relatórios de notificações"""
    try:
        profile = Profile.objects.get(user=request.user)
        
        if not (profile.is_org_admin or profile.can_view_analytics):
            return Response(
                {'error': 'Sem permissão para gerar relatórios'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Parâmetros
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        format_type = request.GET.get('format', 'json')
        
        if not start_date_str or not end_date_str:
            return Response(
                {'error': 'start_date e end_date são obrigatórios'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            start_date = datetime.fromisoformat(start_date_str)
            end_date = datetime.fromisoformat(end_date_str)
        except ValueError:
            return Response(
                {'error': 'Formato de data inválido. Use YYYY-MM-DD'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        report = NotificationReportsService.generate_notification_report(
            profile.organization, start_date, end_date, format_type
        )
        
        if format_type == 'csv':
            response = Response(report, content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="notification_report.csv"'
            return response
        
        return Response(report)
        
    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def workflow_efficiency_report(request):
    """Relatório de eficiência de workflows"""
    try:
        profile = Profile.objects.get(user=request.user)
        
        if not (profile.is_org_admin or profile.can_view_analytics):
            return Response(
                {'error': 'Sem permissão para ver relatórios de workflow'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Parâmetros
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if not start_date_str or not end_date_str:
            # Usar últimos 30 dias como padrão
            end_date = timezone.now()
            start_date = end_date - timedelta(days=30)
        else:
            try:
                start_date = datetime.fromisoformat(start_date_str)
                end_date = datetime.fromisoformat(end_date_str)
            except ValueError:
                return Response(
                    {'error': 'Formato de data inválido'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        report = NotificationReportsService.generate_workflow_efficiency_report(
            profile.organization, start_date, end_date
        )
        
        return Response(report)
        
    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def trigger_escalation_check(request):
    """Endpoint para acionar verificação de escalação (apenas admins)"""
    try:
        profile = Profile.objects.get(user=request.user)
        
        if not profile.is_org_admin:
            return Response(
                {'error': 'Apenas administradores podem acionar escalação'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        escalated_count = NotificationEscalationService.check_and_escalate_overdue_notifications()
        
        return Response({
            'success': True,
            'escalated_notifications': escalated_count,
            'message': f'{escalated_count} notificações foram escaladas'
        })
        
    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    
class NotificationTemplateViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationTemplateSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        try:
            profile = user.profile
            if not profile.organization:
                return NotificationTemplate.objects.none()
            
            # Org Admins can manage templates for their organization.
            if not profile.is_org_admin:
                return NotificationTemplate.objects.none()
            
            # OPTIMIZATION
            return NotificationTemplate.objects.filter(
                organization=profile.organization
            ).select_related('organization', 'created_by')
            
        except Profile.DoesNotExist:
            return NotificationTemplate.objects.none()
    
    def perform_create(self, serializer):
        try:
            profile = Profile.objects.get(user=self.request.user)
            
            if not profile.is_org_admin:
                raise PermissionDenied("Apenas administradores podem criar templates")
            
            serializer.save(
                organization=profile.organization,
                created_by=self.request.user
            )
            
        except Profile.DoesNotExist:
            raise PermissionDenied("Perfil não encontrado")
    
    @action(detail=False, methods=['get'])
    def available_types(self, request):
        """Lista tipos de notificação disponíveis"""
        types = [
            {'value': choice[0], 'label': choice[1]}
            for choice in WorkflowNotification.NOTIFICATION_TYPES
        ]
        return Response(types)
    
    @action(detail=True, methods=['post'])
    def preview(self, request, pk=None):
        """Preview do template com dados de exemplo"""
        template = self.get_object()
        
        # Dados de exemplo
        example_context = {
            'user_name': 'João Silva',
            'user_first_name': 'João',
            'task_title': 'Declaração de IRS 2024',
            'client_name': 'Empresa XYZ Lda',
            'step_name': 'Revisão de Documentos',
            'workflow_name': 'Processo de Declaração',
            'organization_name': template.organization.name,
            'current_date': timezone.now().strftime('%d/%m/%Y'),
            'current_time': timezone.now().strftime('%H:%M'),
        }
        
        try:
            title, message = template.render(example_context)
            return Response({
                'title': title,
                'message': message,
                'context_used': example_context
            })
        except Exception as e:
            return Response({
                'error': f'Erro ao renderizar template: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def create_defaults(self, request):
        """Cria templates padrão para a organização"""
        try:
            profile = Profile.objects.get(user=request.user)
            
            if not profile.is_org_admin:
                raise PermissionDenied("Apenas administradores podem criar templates padrão")
            
            created_count = 0
            
            # Templates padrão para cada tipo
            default_templates = [
                {
                    'notification_type': 'step_ready',
                    'name': 'Passo Pronto - Padrão',
                    'title_template': '🔔 Passo pronto: {step_name}',
                    'message_template': 'Olá {user_first_name},\n\nA tarefa "{task_title}" (Cliente: {client_name}) chegou ao passo "{step_name}" e está pronta para ser trabalhada.\n\nWorkflow: {workflow_name}\nData: {current_date}',
                    'default_priority': 'normal',
                    'is_default': True
                },
                {
                    'notification_type': 'approval_needed',
                    'name': 'Aprovação Necessária - Padrão',
                    'title_template': '⚠️ Aprovação necessária: {step_name}',
                    'message_template': 'Caro {user_first_name},\n\nO passo "{step_name}" da tarefa "{task_title}" (Cliente: {client_name}) precisa de sua aprovação.\n\nPor favor, revise e aprove o mais breve possível.\n\nData: {current_date} às {current_time}',
                    'default_priority': 'high',
                    'is_default': True
                },
                {
                    'notification_type': 'step_overdue',
                    'name': 'Passo Atrasado - Padrão',
                    'title_template': '🚨 URGENTE: Passo atrasado - {step_name}',
                    'message_template': 'Atenção {user_first_name},\n\nO passo "{step_name}" da tarefa "{task_title}" (Cliente: {client_name}) está ATRASADO.\n\nPor favor, priorize esta atividade.\n\nWorkflow: {workflow_name}\nData: {current_date}',
                    'default_priority': 'urgent',
                    'is_default': True
                }
            ]
            
            for template_data in default_templates:
                # Verificar se já existe
                existing = NotificationTemplate.objects.filter(
                    organization=profile.organization,
                    notification_type=template_data['notification_type'],
                    is_default=True
                ).exists()
                
                if not existing:
                    NotificationTemplate.objects.create(
                        organization=profile.organization,
                        created_by=request.user,
                        **template_data
                    )
                    created_count += 1
            
            return Response({
                'success': True,
                'created_templates': created_count,
                'message': f'{created_count} templates padrão criados'
            })
            
        except Profile.DoesNotExist:
            return Response({
                'error': 'Perfil não encontrado'
            }, status=status.HTTP_404_NOT_FOUND)


class NotificationDigestViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = NotificationDigestSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        # OPTIMIZATION: `prefetch_related` is crucial for `notifications.count` in serializer
        # `annotate` is even better for the count.
        return NotificationDigest.objects.filter(
            user=user
        ).annotate(
            notifications_count=Count('notifications')
        ).order_by('-created_at')
        
    @action(detail=False, methods=['post'])
    def generate_digest(self, request):
        """Gera digest manual para o usuário"""
        digest_type = request.data.get('type', 'daily')
        
        if digest_type not in ['hourly', 'daily', 'weekly']:
            return Response({
                'error': 'Tipo de digest inválido'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            if digest_type == 'daily':
                digest = NotificationDigestService._create_daily_digest(request.user)
            else:
                return Response({
                    'error': 'Tipo de digest não implementado ainda'
                }, status=status.HTTP_501_NOT_IMPLEMENTED)
            
            if digest:
                serializer = self.get_serializer(digest)
                return Response({
                    'success': True,
                    'digest': serializer.data
                })
            else:
                return Response({
                    'message': 'Nenhuma notificação encontrada para o período'
                })
                
        except Exception as e:
            logger.error(f"Erro ao gerar digest: {e}")
            return Response({
                'error': 'Erro interno ao gerar digest'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_daily_digests_view(request):
    """Gera digests diários para todos os usuários (apenas admins)"""
    try:
        profile = Profile.objects.get(user=request.user)
        
        if not profile.is_org_admin:
            return Response(
                {'error': 'Apenas administradores podem gerar digests'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        generated_count = NotificationDigestService.generate_daily_digests()
        
        return Response({
            'success': True,
            'digests_generated': generated_count,
            'message': f'{generated_count} digests diários gerados'
        })
        
    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Erro ao gerar digests diários: {e}")
        return Response(
            {'error': 'Erro interno ao gerar digests'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_pending_digests_view(request):
    """Envia digests pendentes (apenas admins)"""
    try:
        profile = Profile.objects.get(user=request.user)
        
        if not profile.is_org_admin:
            return Response(
                {'error': 'Apenas administradores podem enviar digests'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        sent_count = NotificationDigestService.send_pending_digests()
        
        return Response({
            'success': True,
            'digests_sent': sent_count,
            'message': f'{sent_count} digests enviados'
        })
        
    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Erro ao enviar digests: {e}")
        return Response(
            {'error': 'Erro interno ao enviar digests'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_fiscal_obligations_manual(request):
    """
    Endpoint para geração manual de obrigações fiscais.
    Apenas administradores podem executar.
    """
    try:
        profile = Profile.objects.get(user=request.user)
        if not profile.is_org_admin:
            return Response(
                {'error': 'Apenas administradores podem gerar obrigações fiscais manualmente'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        organization = profile.organization
        if not organization:
            return Response(
                {'error': 'Administrador não associado a uma organização'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # Parâmetros opcionais
        months_ahead = int(request.data.get('months_ahead', 3))
        clean_old = request.data.get('clean_old', False)
        days_old = int(request.data.get('days_old', 30))

        # Limpar obrigações obsoletas se solicitado
        cleaned_count = 0
        if clean_old:
            cleaned_count = FiscalObligationGenerator.clean_old_pending_obligations(
                days_old=days_old,
                organization=organization
            )

        # Gerar obrigações
        results = FiscalObligationGenerator.generate_for_next_months(
            months_ahead=months_ahead,
            organization=organization
        )

        # Calcular totais
        total_created = sum(result['tasks_created'] for result in results)
        total_skipped = sum(result['tasks_skipped'] for result in results)
        total_errors = sum(len(result['errors']) for result in results)

        return Response({
            'success': True,
            'message': f'Geração concluída para {organization.name}',
            'summary': {
                'months_processed': len(results),
                'tasks_created': total_created,
                'tasks_skipped': total_skipped,
                'errors': total_errors,
                'old_tasks_cleaned': cleaned_count
            },
            'detailed_results': results
        })

    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Erro na geração manual de obrigações: {e}")
        return Response(
            {'error': f'Erro interno: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def fiscal_obligations_stats(request):
    """
    Endpoint para obter estatísticas do sistema de obrigações fiscais.
    """
    try:
        profile = Profile.objects.get(user=request.user)
        
        # Verificar permissões - admins ou usuários com permissão de analytics
        if not (profile.is_org_admin or profile.can_view_analytics):
            return Response(
                {'error': 'Sem permissão para ver estatísticas fiscais'}, 
                status=status.HTTP_403_FORBIDDEN
            )

        organization = profile.organization
        stats = FiscalObligationGenerator.get_generation_stats(organization)
        
        # Adicionar informações extras
        if organization:
            stats['organization_info'] = {
                'name': organization.name,
                'active_clients': organization.clients.filter(is_active=True).count(),
                'clients_with_tags': organization.clients.exclude(fiscal_tags=[]).count(),
                'active_definitions': FiscalObligationDefinition.objects.filter(
                    Q(organization__isnull=True) | Q(organization=organization),
                    is_active=True
                ).count()
            }

        return Response(stats)

    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Erro ao obter estatísticas fiscais: {e}")
        return Response(
            {'error': f'Erro interno: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def test_client_fiscal_obligations(request):
    """
    Endpoint para testar quais obrigações seriam geradas para um cliente específico.
    """
    try:
        profile = Profile.objects.get(user=request.user)
        
        client_id = request.data.get('client_id')
        year = request.data.get('year', timezone.now().year)
        month = request.data.get('month', timezone.now().month)
        
        if not client_id:
            return Response(
                {'error': 'client_id é obrigatório'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            client = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            return Response(
                {'error': 'Cliente não encontrado'}, 
                status=status.HTTP_404_NOT_FOUND
            )

        # Verificar acesso ao cliente
        if not profile.can_access_client(client):
            return Response(
                {'error': 'Sem acesso a este cliente'}, 
                status=status.HTTP_403_FORBIDDEN
            )

        # Buscar definições aplicáveis
        definitions_query = FiscalObligationDefinition.objects.filter(is_active=True)
        
        if client.organization:
            definitions_query = definitions_query.filter(
                Q(organization__isnull=True) | Q(organization=client.organization)
            )

        applicable_obligations = []
        for definition in definitions_query:
            # Verificar se se aplica ao período
            if not FiscalObligationGenerator._should_generate_for_period(definition, year, month):
                continue
                
            # Verificar se se aplica ao cliente
            if definition.applies_to_client_tags:
                if definition.applies_to_client_tags != ['ALL'] and not any(
                    tag in (client.fiscal_tags or []) for tag in definition.applies_to_client_tags
                ):
                    continue
            
            # Calcular deadline
            deadline_info = FiscalObligationGenerator._calculate_deadline(definition, year, month)
            if not deadline_info:
                continue
            
            # Verificar se deve gerar agora
            should_generate = FiscalObligationGenerator._should_generate_now(
                deadline_info['deadline'], 
                definition.generation_trigger_offset_days
            )
            
            # Verificar se já existe
            period_key = FiscalObligationGenerator._generate_period_key(definition, year, month)
            existing_task = client.tasks.filter(
                source_fiscal_obligation=definition,
                obligation_period_key=period_key
            ).first()
            
            # Gerar título simulado
            simulated_title = FiscalObligationGenerator._generate_task_title(
                definition, client, deadline_info, year, month
            )
            
            applicable_obligations.append({
                'definition_id': definition.id,
                'definition_name': definition.name,
                'periodicity': definition.get_periodicity_display(),
                'required_tags': definition.applies_to_client_tags or [],
                'deadline': deadline_info['deadline'],
                'period_description': deadline_info['period_description'],
                'should_generate_now': should_generate,
                'existing_task': {
                    'id': existing_task.id,
                    'title': existing_task.title,
                    'status': existing_task.status
                } if existing_task else None,
                'simulated_title': simulated_title,
                'trigger_offset_days': definition.generation_trigger_offset_days,
                'priority': definition.get_default_priority_display()
            })

        return Response({
            'client': {
                'id': client.id,
                'name': client.name,
                'fiscal_tags': client.fiscal_tags or []
            },
            'test_period': {
                'year': year,
                'month': month,
                'month_name': datetime(year, month, 1).strftime('%B %Y')
            },
            'applicable_obligations': applicable_obligations,
            'summary': {
                'total_definitions_checked': definitions_query.count(),
                'applicable_count': len(applicable_obligations),
                'would_generate_count': sum(1 for o in applicable_obligations if o['should_generate_now'] and not o['existing_task']),
                'already_exists_count': sum(1 for o in applicable_obligations if o['existing_task'])
            }
        })

    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Erro no teste de obrigações fiscais: {e}")
        return Response(
            {'error': f'Erro interno: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

class FiscalSystemSettingsViewSet(viewsets.ModelViewSet):
    serializer_class = FiscalSystemSettingsSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        try:
            profile = Profile.objects.get(user=user)
            if profile.organization and profile.is_org_admin:
                return FiscalSystemSettings.objects.filter(organization=profile.organization)
            return FiscalSystemSettings.objects.none()
        except Profile.DoesNotExist:
            return FiscalSystemSettings.objects.none()
    
    def get_object(self):
        """Retorna ou cria configurações da organização do usuário."""
        try:
            profile = Profile.objects.get(user=self.request.user)
            if not profile.is_org_admin:
                raise PermissionDenied("Apenas administradores podem gerenciar configurações fiscais")
            
            if not profile.organization:
                raise ValidationError("Usuário não pertence a uma organização")
            
            settings = FiscalSystemSettings.get_for_organization(profile.organization)
            return settings
        except Profile.DoesNotExist:
            raise PermissionDenied("Perfil não encontrado")
    
    @action(detail=False, methods=['get'])
    def my_settings(self, request):
        """Endpoint para obter configurações da organização do usuário."""
        settings = self.get_object()
        serializer = self.get_serializer(settings)
        return Response(serializer.data)
    
    @action(detail=False, methods=['patch'])
    def update_settings(self, request):
        """Endpoint para atualizar configurações."""
        settings = self.get_object()
        serializer = self.get_serializer(settings, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def test_webhook(self, request):
        """Testa o webhook configurado."""
        settings = self.get_object()
        
        if not settings.webhook_url:
            return Response(
                {'error': 'Nenhum webhook configurado'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            test_data = {
                'type': 'webhook_test',
                'organization': settings.organization.name,
                'message': 'Teste de webhook do sistema fiscal',
                'timestamp': timezone.now().isoformat()
            }
            
            FiscalNotificationService._send_webhook(settings, test_data)
            
            return Response({'success': True, 'message': 'Webhook testado com sucesso'})
            
        except Exception as e:
            return Response(
                {'error': f'Erro no teste do webhook: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'])
    def send_test_email(self, request):
        """Envia email de teste."""
        settings = self.get_object()
        
        try:
            test_data = {
                'type': 'email_test',
                'organization': settings.organization.name,
                'stats': {
                    'tasks_created': 5,
                    'tasks_skipped': 2,
                    'errors': []
                },
                'timestamp': timezone.now().isoformat(),
                'success': True
            }
            
            FiscalNotificationService._send_generation_email(settings, test_data)
            
            return Response({'success': True, 'message': 'Email de teste enviado'})
            
        except Exception as e:
            return Response(
                {'error': f'Erro no envio do email: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_fiscal_obligations_manual(request):
    """
    Endpoint para geração manual de obrigações fiscais.
    """
    try:
        serializer = FiscalGenerationRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        profile = Profile.objects.get(user=request.user)
        if not profile.is_org_admin:
            return Response(
                {'error': 'Apenas administradores podem gerar obrigações fiscais manualmente'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        organization = profile.organization
        if not organization:
            return Response(
                {'error': 'Administrador não associado a uma organização'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        validated_data = serializer.validated_data
        months_ahead = validated_data['months_ahead']
        clean_old = validated_data['clean_old']
        days_old = validated_data['days_old']

        # Limpar obrigações obsoletas se solicitado
        cleaned_count = 0
        if clean_old:
            cleaned_count = FiscalObligationGenerator.clean_old_pending_obligations(
                days_old=days_old,
                organization=organization
            )

        # Gerar obrigações
        results = FiscalObligationGenerator.generate_for_next_months(
            months_ahead=months_ahead,
            organization=organization
        )

        # Calcular totais
        total_created = sum(result['tasks_created'] for result in results)
        total_skipped = sum(result['tasks_skipped'] for result in results)
        total_errors = sum(len(result['errors']) for result in results)

        return Response({
            'success': True,
            'message': f'Geração concluída para {organization.name}',
            'summary': {
                'months_processed': len(results),
                'tasks_created': total_created,
                'tasks_skipped': total_skipped,
                'errors': total_errors,
                'old_tasks_cleaned': cleaned_count
            },
            'detailed_results': results
        })

    except Profile.DoesNotExist:
        return Response(
            {'error': 'Perfil não encontrado'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Erro na geração manual de obrigações: {e}")
        return Response(
            {'error': f'Erro interno: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def fiscal_upcoming_deadlines(request):
    """
    Returns fiscal obligation tasks with upcoming deadlines for the user's organization.
    Default: next 30 days.
    Query Params:
        - days: Number of days ahead to check for deadlines (e.g., ?days=7 for next week). Default is 30.
        - limit: Max number of tasks to return. Default is 20.
    """
    try:
        profile = Profile.objects.get(user=request.user)
        if not profile.organization:
            return Response({'error': 'Usuário não associado a uma organização.'}, status=status.HTTP_400_BAD_REQUEST)

        # Permissions: Org admins or users with analytics/task view permissions
        if not (profile.is_org_admin or profile.can_view_analytics or profile.can_view_all_tasks):
            # If user can only see assigned tasks, filter for those. This might be too restrictive for a dashboard view.
            # For a general dashboard, typically broader view is preferred if allowed.
            # Consider if non-admins should see all upcoming fiscal deadlines for their org, or only their assigned ones.
            # For now, restricting to admins/analytics viewers for org-wide view.
             return Response({'error': 'Sem permissão para visualizar prazos fiscais da organização.'}, status=status.HTTP_403_FORBIDDEN)


        days_ahead_str = request.query_params.get('days', '30')
        limit_str = request.query_params.get('limit', '20')

        try:
            days_ahead = int(days_ahead_str)
            limit = int(limit_str)
            if not (1 <= days_ahead <= 90): # Limit days_ahead to a reasonable range
                raise ValueError("Parâmetro 'days' deve estar entre 1 e 90.")
            if not (1 <= limit <= 100): # Limit the number of results
                 raise ValueError("Parâmetro 'limit' deve estar entre 1 e 100.")
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        today = timezone.now().date()
        future_date = today + timedelta(days=days_ahead)

        upcoming_tasks = Task.objects.filter(
            client__organization=profile.organization,
            source_fiscal_obligation__isnull=False, # Ensure it's a fiscal obligation task
            deadline__gte=today,
            deadline__lte=future_date,
            status__in=['pending', 'in_progress'] # Only active tasks
        ).select_related(
            'client', 
            'assigned_to', 
            'source_fiscal_obligation' # To get obligation name if needed
        ).order_by('deadline', 'priority')[:limit]

        # Using TaskSerializer but might need a more specific one if you want less/different fields
        serializer = TaskSerializer(upcoming_tasks, many=True) 
        return Response(serializer.data)

    except Profile.DoesNotExist:
        return Response({'error': 'Perfil não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Erro ao buscar prazos fiscais futuros: {e}")
        return Response({'error': f'Erro interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_ai_advisor_initial_context(request):
    try:
        profile = Profile.objects.select_related('organization').get(user=request.user) # Fetch organization along with profile
        if not profile.organization:
            return Response({"error": "Usuário não está associado a uma organização."}, status=status.HTTP_400_BAD_REQUEST)

        if not profile.is_org_admin: # Or a more specific 'can_use_ai_advisor' permission
            return Response({"error": "Sem permissão para aceder ao consultor AI."}, status=status.HTTP_403_FORBIDDEN)

        organization = profile.organization
        context_data = {
            "organization_name": organization.name,
            "current_date": timezone.now().strftime("%Y-%m-%d"),
            "data_summary_period": "Últimos 90 dias e dados atuais relevantes" # Clarify the scope
        }
        
        # --- Timeframe for aggregations ---
        ninety_days_ago = timezone.now().date() - timedelta(days=90)
        today = timezone.now().date() # Use today consistently

        # 1. Clients Overview
        # ... (Client overview logic remains the same) ...
        active_clients = Client.objects.filter(organization=organization, is_active=True)
        client_details_sample = []
        top_clients = active_clients.order_by('-monthly_fee')[:5]
        low_fee_clients = active_clients.filter(monthly_fee__lte=0).order_by('name')[:3]
        
        clients_for_sample = list(top_clients) + list(low_fee_clients)
        clients_for_sample_ids = {c.id for c in clients_for_sample}
        
        if len(clients_for_sample_ids) < 8:
            additional_clients_needed = 8 - len(clients_for_sample_ids)
            other_clients = active_clients.exclude(id__in=clients_for_sample_ids).order_by('?')[:additional_clients_needed]
            clients_for_sample.extend(other_clients)
            clients_for_sample_ids.update(c.id for c in other_clients)
            
        unique_sample_clients = {c.id: c for c in clients_for_sample}.values()

        for client in unique_sample_clients:
            profit_record = ClientProfitability.objects.filter(
                client=client
            ).order_by('-year', '-month').first()
            client_details_sample.append({
                "name": client.name,
                "monthly_fee": float(client.monthly_fee or 0),
                "fiscal_tags": client.fiscal_tags or [],
                "recent_profit_margin": float(profit_record.profit_margin) if profit_record and profit_record.profit_margin is not None else None,
                "active_tasks_count": Task.objects.filter(client=client, status__in=['pending', 'in_progress']).count(),
            })
        context_data['clients_overview'] = {
            "total_active_clients": active_clients.count(),
            "clients_sample_details": client_details_sample,
            "clients_with_no_fee": active_clients.filter(monthly_fee__lte=0).count()
        }

        # 2. Profitability Snapshot (Organization Level for last 90 days)
        # ... (Profitability snapshot logic remains the same) ...
        profitability_snapshot = ClientProfitability.objects.filter(
            client__organization=organization,
            last_updated__gte=ninety_days_ago 
        ).aggregate(
            avg_profit_margin_org=Avg('profit_margin', filter=Q(profit_margin__isnull=False)),
            total_profit_org=Sum('profit', filter=Q(profit__isnull=False)),
            count_profitable_periods=Count('id', filter=Q(is_profitable=True)),
            count_unprofitable_periods=Count('id', filter=Q(is_profitable=False))
        )
        context_data['profitability_snapshot_organization'] = {
            "average_profit_margin": profitability_snapshot.get('avg_profit_margin_org'),
            "total_profit_last_90_days_approx": profitability_snapshot.get('total_profit_org'),
            "profitable_client_months_count": profitability_snapshot.get('count_profitable_periods'),
            "unprofitable_client_months_count": profitability_snapshot.get('count_unprofitable_periods'),
        }

        # 3. Tasks Overview
        org_tasks = Task.objects.filter(client__organization=organization)
        active_org_tasks = org_tasks.filter(status__in=['pending', 'in_progress'])
        
        duration_expression = ExpressionWrapper(F('completed_at') - F('created_at'), output_field=fields.DurationField())
        completed_tasks_recent = org_tasks.filter(
            status='completed',
            completed_at__gte=ninety_days_ago,
            created_at__isnull=False,
            completed_at__isnull=False
        )
        avg_completion_duration_data = completed_tasks_recent.annotate(duration=duration_expression).aggregate(avg_duration=Avg('duration'))
        
        avg_completion_days = None
        if avg_completion_duration_data and avg_completion_duration_data.get('avg_duration') is not None:
            avg_duration_value = avg_completion_duration_data['avg_duration']
            # Add an explicit check for timedelta, though Avg(DurationField) should return this or None
            if isinstance(avg_duration_value, timedelta):
                avg_completion_days = avg_duration_value.total_seconds() / (60*60*24)
            else:
                # This case is unexpected with standard Django ORM usage but good for logging
                logger.warning(
                    f"Unexpected type for avg_duration: {type(avg_duration_value)}. Value: {avg_duration_value}. Org: {organization.id}"
                )
                avg_completion_days = None # Ensure it's None if not a timedelta

        context_data['tasks_overview'] = {
            "total_tasks": org_tasks.count(),
            "active_tasks": active_org_tasks.count(),
            "overdue_tasks": active_org_tasks.filter(deadline__lt=today).count(),
            "completed_last_90_days": completed_tasks_recent.count(),
            # This line was already robust:
            "avg_task_completion_days_last_90_days": round(avg_completion_days, 1) if avg_completion_days is not None else None,
        }

        tasks_per_category = TaskCategory.objects.annotate(
            active_task_count=Count('tasks', filter=Q(tasks__client__organization=organization, tasks__status__in=['pending', 'in_progress']))
        ).filter(active_task_count__gt=0).order_by('-active_task_count')[:5]
        
        context_data['tasks_overview']['active_tasks_per_category_top_5'] = [
            {"category_name": cat.name, "count": cat.active_task_count} for cat in tasks_per_category
        ]

        # --- START: NEW SECTION FOR DETAILED TASK SAMPLE ---
        detailed_tasks_sample = []
        sample_task_ids = set()
        MAX_SAMPLE_TASKS = 15 # Max tasks to include in the detailed sample

        # a. Overdue tasks (up to 5)
        overdue_sample = active_org_tasks.filter(
            deadline__lt=today
        ).select_related('assigned_to', 'client').order_by('deadline', '-priority')[:5]
        for task in overdue_sample:
            if len(detailed_tasks_sample) < MAX_SAMPLE_TASKS and task.id not in sample_task_ids:
                detailed_tasks_sample.append({
                    "title": task.title,
                    "client_name": task.client.name if task.client else "N/A",
                    "status": task.get_status_display(),
                    "priority": task.get_priority_display(),
                    "deadline": task.deadline.strftime("%Y-%m-%d") if task.deadline else "N/A",
                    "assigned_to": task.assigned_to.username if task.assigned_to else "Não atribuído"
                })
                sample_task_ids.add(task.id)

        # b. Tasks due today or tomorrow (up to 5, excluding already added overdue)
        if len(detailed_tasks_sample) < MAX_SAMPLE_TASKS:
            due_soon_sample = active_org_tasks.filter(
                deadline__gte=today,
                deadline__lte=today + timedelta(days=1)
            ).exclude(id__in=sample_task_ids).select_related('assigned_to', 'client').order_by('deadline', 'priority')[:5]
            for task in due_soon_sample:
                if len(detailed_tasks_sample) < MAX_SAMPLE_TASKS and task.id not in sample_task_ids:
                    detailed_tasks_sample.append({
                        "title": task.title,
                        "client_name": task.client.name if task.client else "N/A",
                        "status": task.get_status_display(),
                        "priority": task.get_priority_display(),
                        "deadline": task.deadline.strftime("%Y-%m-%d") if task.deadline else "N/A",
                        "assigned_to": task.assigned_to.username if task.assigned_to else "Não atribuído"
                    })
                    sample_task_ids.add(task.id)
        
        # c. Other active high priority tasks or recently updated active tasks (to fill up to MAX_SAMPLE_TASKS)
        if len(detailed_tasks_sample) < MAX_SAMPLE_TASKS:
            remaining_needed = MAX_SAMPLE_TASKS - len(detailed_tasks_sample)
            other_active_sample = active_org_tasks.exclude(
                id__in=sample_task_ids
            ).select_related('assigned_to', 'client').order_by('-priority', '-updated_at')[:remaining_needed]
            for task in other_active_sample:
                # No need to check MAX_SAMPLE_TASKS again due to slice, but keep id check
                if task.id not in sample_task_ids:
                    detailed_tasks_sample.append({
                        "title": task.title,
                        "client_name": task.client.name if task.client else "N/A",
                        "status": task.get_status_display(),
                        "priority": task.get_priority_display(),
                        "deadline": task.deadline.strftime("%Y-%m-%d") if task.deadline else "N/A",
                        "assigned_to": task.assigned_to.username if task.assigned_to else "Não atribuído"
                    })
                    sample_task_ids.add(task.id)

        context_data['tasks_overview']['detailed_tasks_sample'] = detailed_tasks_sample
        # --- END: NEW SECTION FOR DETAILED TASK SAMPLE ---


        # 4. Fiscal Obligations Snapshot
        # ... (Fiscal obligations logic remains the same) ...
        upcoming_fiscal_days = 30
        future_date_fiscal = today + timedelta(days=upcoming_fiscal_days) # Corrected: today
        
        upcoming_fiscal_tasks = Task.objects.filter(
            client__organization=organization,
            source_fiscal_obligation__isnull=False,
            deadline__gte=today, # Corrected: today
            deadline__lte=future_date_fiscal,
            status__in=['pending', 'in_progress']
        ).select_related('source_fiscal_obligation', 'client').order_by('deadline')

        context_data['fiscal_obligations_snapshot'] = {
            "total_active_fiscal_definitions": FiscalObligationDefinition.objects.filter(
                Q(organization=organization) | Q(organization__isnull=True), is_active=True
            ).count(),
            "upcoming_deadlines_next_30_days_count": upcoming_fiscal_tasks.count(),
            "upcoming_deadlines_sample": [
                {
                    "obligation_name": task.source_fiscal_obligation.name if task.source_fiscal_obligation else task.title,
                    "client_name": task.client.name,
                    "deadline": task.deadline.strftime("%Y-%m-%d")
                } for task in upcoming_fiscal_tasks[:3]
            ]
        }

        # 5. Team Performance Snippet (Corrected field names based on previous implementation)
        # ... (Team performance logic remains the same) ...
        active_users_count = Profile.objects.filter(organization=organization, user__is_active=True).count()
        if active_users_count > 0:
            tasks_completed_last_month_org = org_tasks.filter(status='completed', completed_at__gte=today - timedelta(days=30)).count() # Corrected: today
            avg_tasks_per_user = tasks_completed_last_month_org / active_users_count if active_users_count > 0 else 0
            context_data['team_performance_snippet'] = {
                "active_team_members": active_users_count,
                "avg_tasks_completed_per_member_last_30_days": round(avg_tasks_per_user,1)
            }
        else:
            context_data['team_performance_snippet'] = {
                 "active_team_members": 0,
                 "avg_tasks_completed_per_member_last_30_days": 0
            }


        logger.info(f"Generated AI advisor context for org {organization.id}, user {request.user.username}. Size approx {len(json.dumps(context_data, cls=CustomJSONEncoder))} bytes.")

        return Response(context_data)

    except Profile.DoesNotExist:
        return Response({"error": "Perfil de usuário não encontrado."}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error generating AI advisor initial context: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({"error": "Erro ao preparar dados para o Consultor AI."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def start_ai_advisor_session(request):
    try:
        profile = Profile.objects.get(user=request.user)
        if not profile.is_org_admin: # Or your specific permission check
            return Response({"error": "Sem permissão para iniciar sessão com o Consultor AI."}, status=status.HTTP_403_FORBIDDEN)

        context_data = request.data.get('context', {})
        if not context_data: # This context comes from the /ai-advisor/get-initial-context/ call
            return Response({"error": "Dados de contexto são necessários para iniciar a sessão."}, status=status.HTTP_400_BAD_REQUEST)

        advisor_service = AIAdvisorService()
        session_id, initial_message = advisor_service.start_session(context_data, request.user)

        if session_id is None: # Indicates an error occurred within the service
            return Response({"error": initial_message or "Não foi possível iniciar a sessão com o Consultor AI."}, status=status.HTTP_502_BAD_GATEWAY)

        # Log organization action
        log_organization_action(
            request,
            action_type='START_AI_ADVISOR_SESSION',
            action_description=f"Sessão do Consultor AI iniciada pelo usuário {request.user.username} (ID: {request.user.id}) com contexto: {str(context_data)[:200]}",
            related_object=None
        )
        return Response({
            "session_id": session_id,
            "initial_message": initial_message
        }, status=status.HTTP_201_CREATED)

    except Profile.DoesNotExist:
        return Response({"error": "Perfil de usuário não encontrado."}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error in start_ai_advisor_session view: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({"error": "Erro interno ao iniciar a sessão com o Consultor AI."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def query_ai_advisor(request):
    try:
        profile = Profile.objects.get(user=request.user)
        if not profile.is_org_admin: # Or your specific permission check
            return Response({"error": "Sem permissão para consultar o Consultor AI."}, status=status.HTTP_403_FORBIDDEN)

        session_id = request.data.get('session_id')
        user_query_text = request.data.get('query')

        if not session_id or not user_query_text:
            return Response({"error": "session_id e query são obrigatórios."}, status=status.HTTP_400_BAD_REQUEST)

        advisor_service = AIAdvisorService()
        ai_response_text, error_message = advisor_service.process_query(session_id, user_query_text, request.user)

        # Log organization action
        log_organization_action(
            request,
            action_type='QUERY_AI_ADVISOR',
            action_description=f"Usuário {request.user.username} (ID: {request.user.id}) consultou o Consultor AI (sessão: {session_id}) com query: {user_query_text[:200]}",
            related_object=None
        )
        if error_message:
            # Check if it's a session not found error specifically
            if "Sessão inválida ou expirada" in error_message:
                 return Response({"error": error_message}, status=status.HTTP_404_NOT_FOUND)
            return Response({"error": error_message or "Não foi possível obter resposta do Consultor AI."}, status=status.HTTP_502_BAD_GATEWAY)
        return Response({"response": ai_response_text})

    except Profile.DoesNotExist:
        return Response({"error": "Perfil de usuário não encontrado."}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error in query_ai_advisor view: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({"error": "Erro interno ao consultar o Consultor AI."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from .services.report_generation_service import ReportGenerationService
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import uuid
import os
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_report(request):
    """
    Endpoint to INITIATE a report generation.
    This creates a report record and dispatches a Celery task to do the work.
    """
    try:
        profile = Profile.objects.select_related('organization').get(user=request.user)
        
        if not (profile.is_org_admin or profile.can_export_reports or profile.can_create_custom_reports):
            return Response({'error': 'Sem permissão para gerar relatórios'}, status=status.HTTP_403_FORBIDDEN)
        
        organization = profile.organization
        if not organization:
            return Response({'error': 'Utilizador não está associado a uma organização'}, status=status.HTTP_400_BAD_REQUEST)

        report_type = request.data.get('report_type')
        report_format = request.data.get('format', 'pdf')
        report_name = request.data.get('name', f'Relatório {report_type}')
        params = request.data.get('parameters', {})
        
        if not report_type:
            return Response({'error': 'Tipo de relatório é obrigatório'}, status=status.HTTP_400_BAD_REQUEST)
        
        # --- Create the initial report record with PENDING status ---
        generated_report = GeneratedReport.objects.create(
            name=report_name,
            report_type=report_type,
            report_format=report_format,
            organization=organization,
            generated_by=request.user,
            parameters=params,
            description=params.get('description', f'Relatório {report_type} gerado em {timezone.now().strftime("%d/%m/%Y")}'),
            status='PENDING' # Initial status
        )
        
        # --- Dispatch the Celery task ---
        generate_report_task.delay(generated_report.id)
        
        logger.info(f"Dispatched report generation task for report {generated_report.id} by user {request.user.username}")

        # --- Return an immediate response to the user ---
        serializer = GeneratedReportSerializer(generated_report)
        return Response({
            'success': True,
            'message': 'A geração do seu relatório foi iniciada. Será notificado quando estiver concluído.',
            'report': serializer.data
        }, status=status.HTTP_202_ACCEPTED) # 202 Accepted is the correct code for async operations

    except Profile.DoesNotExist:
        return Response({'error': 'Perfil não encontrado'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Erro ao iniciar a geração do relatório: {e}", exc_info=True)
        return Response({'error': f'Erro interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_report_generation_context(request):
    """
    Endpoint para obter contexto necessário para geração de relatórios
    (clientes, utilizadores, etc.)
    """
    try:
        profile = Profile.objects.get(user=request.user)
        
        if not (profile.is_org_admin or profile.can_export_reports or profile.can_create_custom_reports):
            return Response(
                {'error': 'Sem permissão para aceder ao contexto de relatórios'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        organization = profile.organization
        if not organization:
            return Response(
                {'error': 'Utilizador não está associado a uma organização'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Obter dados para os formulários de relatório
        context = {
            'organization': {
                'id': organization.id,
                'name': organization.name
            },
            'clients': [],
            'users': [],
            'categories': [],
            'report_types': [
                {'value': 'client_summary', 'label': 'Resumo de Cliente(s)'},
                {'value': 'profitability_analysis', 'label': 'Análise de Rentabilidade'},
                {'value': 'task_performance', 'label': 'Performance de Tarefas'},
                {'value': 'time_tracking_summary', 'label': 'Resumo de Registo de Tempos'},
                {'value': 'custom_report', 'label': 'Relatório Personalizado'},
            ],
            'formats': [
                {'value': 'pdf', 'label': 'PDF'},
                {'value': 'csv', 'label': 'CSV'},
                {'value': 'xlsx', 'label': 'Excel (XLSX)'},
            ]
        }
        
        # Clientes ativos
        clients = Client.objects.filter(
            organization=organization, 
            is_active=True
        ).select_related('account_manager').order_by('name')
        
        context['clients'] = [
            {
                'id': client.id,
                'name': client.name,
                'account_manager': client.account_manager.username if client.account_manager else None,
                'monthly_fee': float(client.monthly_fee) if client.monthly_fee else 0
            }
            for client in clients
        ]
        
        # Utilizadores da organização
        org_users = Profile.objects.filter(
            organization=organization,
            user__is_active=True
        ).select_related('user').order_by('user__username')
        
        context['users'] = [
            {
                'id': profile_item.user.id,
                'username': profile_item.user.username,
                'first_name': profile_item.user.first_name,
                'last_name': profile_item.user.last_name,
                'role': profile_item.role
            }
            for profile_item in org_users
        ]
        
        # Categorias de tarefas
        categories = TaskCategory.objects.all().order_by('name')
        context['categories'] = [
            {
                'id': category.id,
                'name': category.name,
                'color': category.color
            }
            for category in categories
        ]
        
        org_users_profiles = Profile.objects.filter(
            organization=organization,
            user__is_active=True
        ).select_related('user').order_by('user__username')
        
        context['users'] = [ # This should be a list of Profile objects, not User objects directly
            {
                'id': profile_item.user.id, # Use user ID
                'username': profile_item.user.username,
                'first_name': profile_item.user.first_name,
                'last_name': profile_item.user.last_name,
                'role': profile_item.role
            }
            for profile_item in org_users_profiles
        ]
        
        return Response(context)
        
    except Profile.DoesNotExist:
        return Response({'error': 'Perfil não encontrado'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Erro ao obter contexto de relatórios: {e}", exc_info=True)
        return Response({
            'error': f'Erro interno: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Adicionar também este endpoint para download direto (opcional)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_report(request, report_id):
    """
    Endpoint para download direto de relatório por ID.
    """
    try:
        profile = Profile.objects.get(user=request.user)
        
        if not (profile.is_org_admin or profile.can_export_reports):
            return Response(
                {'error': 'Sem permissão para fazer download de relatórios'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Buscar relatório
        try:
            report = GeneratedReport.objects.get(
                id=report_id,
                organization=profile.organization
            )
        except GeneratedReport.DoesNotExist:
            return Response({'error': 'Relatório não encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        # Redirect para URL de storage ou servir arquivo diretamente
        return Response({'download_url': report.storage_url})
        
    except Profile.DoesNotExist:
        return Response({'error': 'Perfil não encontrado'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Erro no download do relatório {report_id}: {e}", exc_info=True)
        return Response({
            'error': f'Erro interno: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def time_entry_context(request):
    """
    Fornece o contexto necessário para o formulário de Registo de Tempo,
    filtrado pelas permissões do utilizador.
    """
    user = request.user
    try:
        profile = Profile.objects.get(user=user)
        if not profile.organization:
            return Response({"error": "Utilizador não está numa organização."}, status=status.HTTP_400_BAD_REQUEST)

        # 1. Obter todas as tarefas ativas da organização que o user pode ver
        # Use o TaskManager para obter a base de tarefas seguras para o usuário
        user_tasks_qs = Task.objects.for_user(user).filter(
            status__in=['pending', 'in_progress']
        ).select_related('client', 'category', 'current_workflow_step')

        # 2. Obter os clientes únicos a partir das tarefas relevantes E dos clientes visíveis
        # Isto garante que o utilizador pode registar tempo para um cliente mesmo que não tenha tarefas ativas para ele
        visible_clients_qs = Client.objects.for_user(user).filter(is_active=True)
        
        # 3. Obter todas as categorias (geralmente não são restritas por permissão)
        categories_qs = TaskCategory.objects.all().order_by('name')
        
        # 4. Serializar os dados para a resposta
        client_serializer = ClientSerializer(visible_clients_qs, many=True)
        task_serializer = TaskSerializer(user_tasks_qs, many=True, context={'request': request})
        category_serializer = TaskCategorySerializer(categories_qs, many=True)
        
        return Response({
            "clients": client_serializer.data,
            "tasks": task_serializer.data,
            "categories": category_serializer.data
        })
        
    except Profile.DoesNotExist:
        return Response({"error": "Perfil não encontrado."}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Erro ao construir contexto de registo de tempo: {e}", exc_info=True)
        return Response({"error": "Erro interno ao buscar contexto."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class OrganizationActionLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = OrganizationActionLog.objects.all().order_by('-timestamp')
    serializer_class = OrganizationActionLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.is_superuser:
            return qs
        if hasattr(user, 'profile') and user.profile.organization:
            return qs.filter(organization=user.profile.organization)
        return qs.none()