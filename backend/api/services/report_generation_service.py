# api/services/report_generation_service.py
import io
import csv
import logging
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Any, Optional, Tuple

from django.db.models import Sum, Count, Avg, Q, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncMonth, TruncDay
from django.utils import timezone
from django.conf import settings

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.axes import XCategoryAxis, YValueAxis
from reportlab.graphics.charts.legends import Legend

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart as OpenpyxlBarChart, PieChart as OpenpyxlPieChart, LineChart as OpenpyxlLineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.utils import get_column_letter


from ..models import (
    Client, Task, TimeEntry, ClientProfitability,
    Organization, Profile, GeneratedReport, TaskCategory, User
)

logger = logging.getLogger(__name__)

class ReportGenerationService:
    """
    Serviço centralizado para geração de relatórios em diferentes formatos.
    """

    @staticmethod
    def _get_pdf_styles():
        styles = getSampleStyleSheet()
        
        C = {
            'primary': colors.HexColor('#3B82F6'), 'primary_dark': colors.HexColor('#1E40AF'),
            'secondary': colors.HexColor('#64748B'), 'text_light': colors.HexColor('#F8FAFC'),
            'text_dark': colors.HexColor('#1E293B'), 'text_muted': colors.HexColor('#475569'),
            'bg_light': colors.HexColor('#F1F5F9'), 'bg_header_table': colors.HexColor('#334155'),
            'border_color': colors.HexColor('#CBD5E1'), 'green_light_bg': colors.HexColor('#ECFDF5'),
            'green_dark_text': colors.HexColor('#047857'), 'red_light_bg': colors.HexColor('#FEF2F2'),
            'red_dark_text': colors.HexColor('#B91C1C'), 'yellow_light_bg': colors.HexColor('#FEFCE8'),
            'yellow_dark_text': colors.HexColor('#854D0E'),
            'gray_bg': colors.HexColor('#E5E7EB') 
        }

        base_font = 'Helvetica'
        base_font_bold = 'Helvetica-Bold'

        custom_styles = [
            ('ReportTitle', {'fontName': base_font_bold, 'fontSize': 20, 'alignment': 1, 'spaceAfter': 6, 'textColor': C['primary_dark']}),
            ('ReportSubtitle', {'fontName': base_font, 'fontSize': 12, 'alignment': 1, 'spaceAfter': 12, 'textColor': C['secondary']}),
            ('SmallTextMuted', {'fontName': base_font, 'fontSize': 8, 'alignment': 0, 'spaceAfter': 6, 'textColor': C['text_muted']}),
            ('SectionTitle', {'fontName': base_font_bold, 'fontSize': 14, 'spaceBefore': 16, 'spaceAfter': 10, 'textColor': C['primary']}),
            ('SubSectionTitle', {'fontName': base_font_bold, 'fontSize': 11, 'spaceBefore': 10, 'spaceAfter': 6, 'textColor': C['primary_dark']}),
            ('TableHeader', {'fontName': base_font_bold, 'fontSize': 9, 'alignment': 1, 'textColor': C['text_light']}),
            ('TableCellText', {'fontName': base_font, 'fontSize': 9, 'alignment': 0, 'textColor': C['text_dark'], 'leading': 11}),
            ('TableCellTextSmall', {'fontName': base_font, 'fontSize': 8, 'alignment': 0, 'textColor': C['text_dark'], 'leading': 10}),
            ('TableCellNumber', {'fontName': base_font, 'fontSize': 9, 'alignment': 2, 'textColor': C['text_dark'], 'leading': 11}),
            ('TableCellCenter', {'fontName': base_font, 'fontSize': 9, 'alignment': 1, 'textColor': C['text_dark'], 'leading': 11}),
            ('AnalysisText', {'fontName': base_font, 'fontSize': 9, 'leading': 12, 'spaceBefore': 6, 'spaceAfter': 6, 'textColor': C['text_dark']}),
            ('InsightTitle', {'fontName': base_font_bold, 'fontSize': 10, 'textColor': C['primary']}),
        ]

        for name, attrs in custom_styles:
            if name not in styles:
                styles.add(ParagraphStyle(name=name, **attrs))
            else:
                existing_style = styles[name]
                for key, value in attrs.items():
                    setattr(existing_style, key, value)
        
        if 'BodyText' in styles:
            styles['BodyText'].fontName = base_font
            styles['BodyText'].fontSize = 10
            styles['BodyText'].leading = 14
            styles['BodyText'].textColor = C['text_dark']
        
        return styles, C

    @staticmethod
    def _add_page_elements(canvas, doc, title="Relatório", org_name=""):
        styles, C = ReportGenerationService._get_pdf_styles()
        canvas.saveState()
        
        canvas.setFont(styles['SmallTextMuted'].fontName, styles['SmallTextMuted'].fontSize - 1)
        canvas.setFillColor(styles['SmallTextMuted'].textColor)
        header_text = f"{org_name} - {title}"
        canvas.drawString(doc.leftMargin, doc.height + doc.topMargin + 0.1*inch, header_text)

        page_num_text = f"Página {doc.page}"
        gen_date_text = f"Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        
        canvas.drawString(doc.leftMargin, 0.5*inch, gen_date_text)
        canvas.drawRightString(doc.width + doc.leftMargin, 0.5*inch, page_num_text)

        canvas.setStrokeColor(C['border_color'])
        canvas.setLineWidth(0.5)
        canvas.line(doc.leftMargin, 0.75*inch, doc.width + doc.leftMargin, 0.75*inch)
        
        canvas.restoreState()

    @staticmethod
    def _build_pdf_doc(buffer, story_elements, title="Relatório", org_name="", pagesize=A4):
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            title=title,
            leftMargin=0.75*inch,
            rightMargin=0.75*inch,
            topMargin=1.0*inch,
            bottomMargin=1.0*inch 
        )
        
        doc.build(
            story_elements, 
            onFirstPage=lambda canvas, doc_obj: ReportGenerationService._add_page_elements(canvas, doc_obj, title, org_name),
            onLaterPages=lambda canvas, doc_obj: ReportGenerationService._add_page_elements(canvas, doc_obj, title, org_name)
        )
        return doc
        
    @staticmethod
    def _get_default_table_style(C_colors, num_data_rows=0, header_bg=None, alt_row_bg=None):
        header_background = header_bg or C_colors['bg_header_table']
        alt_row_background = alt_row_bg or C_colors['gray_bg']
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), header_background),
            ('TEXTCOLOR', (0,0), (-1,0), C_colors['text_light']),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('GRID', (0,0), (-1,-1), 0.5, C_colors['border_color']),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('TEXTCOLOR', (0,1), (-1,-1), C_colors['text_dark']),
            ('TOPPADDING', (0,1), (-1,-1), 5),
            ('BOTTOMPADDING', (0,1), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ])
        for i in range(1, num_data_rows + 1):
            if i % 2 == 1: 
                 style.add('BACKGROUND', (0,i), (-1,i), alt_row_background)
        return style

    @staticmethod
    def _apply_xlsx_header_style(cell, font_bold=True, fill_color="002060", font_color="FFFFFF"):
        cell.font = Font(bold=font_bold, color=font_color, size=10)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

    @staticmethod
    def _apply_xlsx_cell_style(cell, alignment='left', number_format=None, font_bold=False, wrap_text=False):
        cell.font = Font(bold=font_bold, size=9)
        cell.alignment = Alignment(horizontal=alignment, vertical="center", wrap_text=wrap_text)
        if number_format:
            cell.number_format = number_format
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
    
    @staticmethod
    def _format_currency_excel(value):
        try: return float(value) if value is not None else 0.0
        except (ValueError, TypeError): return 0.0

    @staticmethod
    def _format_percentage_excel(value):
        try: return float(value)/100 if value is not None else 0.0
        except (ValueError, TypeError): return 0.0
    
    # --- Report: Client Summary ---
    @staticmethod
    def generate_client_summary_report(
        organization: Organization, client_ids: List[str] = None,
        include_profitability: bool = True, include_tasks: bool = True,
        include_time_entries: bool = True, date_from: datetime = None,
        date_to: datetime = None, format_type: str = 'pdf'
    ) -> Tuple[io.BytesIO, str]:
        clients_query = Client.objects.filter(organization=organization, is_active=True)
        if client_ids:
            clients_query = clients_query.filter(id__in=client_ids)
        clients = clients_query.select_related('account_manager').order_by('name')
        
        logger.info(f"Client Summary Report: Found {clients.count()} clients for org {organization.name} with client_ids {client_ids}")

        report_data = {'organization': organization, 'clients_data': [], 'generation_date': timezone.now(), 'date_from': date_from, 'date_to': date_to}
        report_data['total_clients'] = clients.count()
        report_data['total_monthly_fees'] = clients.aggregate(total=Sum('monthly_fee'))['total'] or Decimal('0.00')

        for client_obj in clients:
            client_detail = {'client_obj': client_obj, 'active_tasks_count': 0, 'completed_tasks_count': 0, 'total_time_minutes': 0, 'recent_profitability': None, 'top_tasks': [], 'recent_time_entries': []}
            
            task_filter = Q(client=client_obj)
            time_filter = Q(client=client_obj)
            if date_from:
                task_filter &= (Q(created_at__gte=date_from) | Q(completed_at__gte=date_from))
                time_filter &= Q(date__gte=date_from.date())
            if date_to:
                task_filter &= (Q(created_at__lte=date_to) | Q(completed_at__lte=date_to))
                time_filter &= Q(date__lte=date_to.date())

            if include_tasks:
                client_detail['active_tasks_count'] = Task.objects.filter(task_filter & Q(status__in=['pending', 'in_progress'])).count()
                client_detail['completed_tasks_count'] = Task.objects.filter(task_filter & Q(status='completed')).count()
                client_detail['top_tasks'] = Task.objects.filter(task_filter).select_related('category').order_by('-priority', '-created_at')[:5]
            
            if include_time_entries:
                time_stats = TimeEntry.objects.filter(time_filter).aggregate(total_minutes=Sum('minutes_spent'))
                client_detail['total_time_minutes'] = time_stats['total_minutes'] or 0
                client_detail['recent_time_entries'] = TimeEntry.objects.filter(time_filter).select_related('user', 'task', 'category').order_by('-date')[:5]
            
            if include_profitability:
                profit_filter = Q(client=client_obj)
                if date_from and date_to:
                    current_date_for_period = date_from
                    year_month_pairs = set()
                    while current_date_for_period <= date_to:
                        year_month_pairs.add((current_date_for_period.year, current_date_for_period.month))
                        if current_date_for_period.month == 12:
                            current_date_for_period = current_date_for_period.replace(year=current_date_for_period.year + 1, month=1)
                        else:
                            current_date_for_period = current_date_for_period.replace(month=current_date_for_period.month + 1)
                    
                    q_objects_profit = Q()
                    for ym_pair in year_month_pairs:
                        q_objects_profit |= Q(year=ym_pair[0], month=ym_pair[1])
                    profit_filter &= q_objects_profit
                
                recent_profit_agg = ClientProfitability.objects.filter(profit_filter).aggregate(
                    avg_profit_margin=Avg('profit_margin'),
                    total_profit=Sum('profit')
                )
                client_detail['recent_profitability'] = {
                    'profit_margin': recent_profit_agg['avg_profit_margin'],
                    'profit': recent_profit_agg['total_profit']
                } if recent_profit_agg['avg_profit_margin'] is not None or recent_profit_agg['total_profit'] is not None else None
            report_data['clients_data'].append(client_detail)
        
        logger.info(f"Client Summary Report Data: {len(report_data['clients_data'])} clients processed for PDF/CSV/XLSX.")

        if format_type == 'pdf':
            return ReportGenerationService._generate_client_summary_pdf(report_data)
        elif format_type == 'csv':
            return ReportGenerationService._generate_client_summary_csv(report_data)
        elif format_type == 'xlsx':
            return ReportGenerationService._generate_client_summary_xlsx(report_data)
        raise ValueError(f"Formato não suportado: {format_type}")

    @staticmethod
    def _generate_client_summary_pdf(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.BytesIO()
        styles, C = ReportGenerationService._get_pdf_styles()
        story = []
        
        story.append(Paragraph("Resumo de Clientes", styles['ReportTitle']))
        story.append(Paragraph(data['organization'].name, styles['ReportSubtitle']))
        story.append(Paragraph(f"Gerado em: {data['generation_date'].strftime('%d/%m/%Y %H:%M')}", styles['SmallTextMuted']))
        if data['date_from'] and data['date_to']:
            story.append(Paragraph(f"Período: {data['date_from'].strftime('%d/%m/%Y')} a {data['date_to'].strftime('%d/%m/%Y')}", styles['SmallTextMuted']))
        story.append(Spacer(1, 0.2*inch))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C['border_color'], spaceBefore=5, spaceAfter=15))

        story.append(Paragraph("Sumário Geral da Organização", styles['SectionTitle']))
        summary_table_data = [
            [Paragraph("Total de Clientes Analisados:", styles['TableCellText']), Paragraph(str(data['total_clients']), styles['TableCellNumber'])],
            [Paragraph("Avenças Mensais Totais:", styles['TableCellText']), Paragraph(f"€{data['total_monthly_fees']:.2f}", styles['TableCellNumber'])],
        ]
        summary_table = Table(summary_table_data, colWidths=[3*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, C['border_color']), ('BACKGROUND', (0,0), (0,-1), C['bg_light'])]))
        story.append(summary_table)
        story.append(Spacer(1, 0.2*inch))

        if not data['clients_data']:
            story.append(Paragraph("Nenhum cliente encontrado para os critérios selecionados.", styles['BodyText']))
        else:
            for client_data in data['clients_data']:
                story.append(PageBreak())
                client_obj = client_data['client_obj']
                story.append(Paragraph(f"Detalhes do Cliente: {client_obj.name}", styles['SectionTitle']))
                
                client_info_data = [
                    [Paragraph("NIF:", styles['TableCellText']), Paragraph(client_obj.nif or 'N/A', styles['TableCellText'])],
                    [Paragraph("Email:", styles['TableCellText']), Paragraph(client_obj.email or 'N/A', styles['TableCellText'])],
                    [Paragraph("Telefone:", styles['TableCellText']), Paragraph(client_obj.phone or 'N/A', styles['TableCellText'])],
                    [Paragraph("Morada:", styles['TableCellText']), Paragraph(client_obj.address or 'N/A', styles['TableCellTextSmall'])],
                    [Paragraph("Gestor de Conta:", styles['TableCellText']), Paragraph(client_obj.account_manager.username if client_obj.account_manager else 'N/A', styles['TableCellText'])],
                    [Paragraph("Avença Mensal:", styles['TableCellText']), Paragraph(f"€{client_obj.monthly_fee or Decimal('0.00'):.2f}", styles['TableCellNumber'])],
                ]
                client_info_table = Table(client_info_data, colWidths=[1.5*inch, 4*inch])
                client_info_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, C['border_color']), ('BACKGROUND', (0,0), (0,-1), C['bg_light'])]))
                story.append(client_info_table)
                story.append(Spacer(1, 0.1*inch))

                if client_data.get('recent_profitability'):
                    profit_detail = client_data['recent_profitability']
                    profit_value = profit_detail.get('profit')
                    profit_margin_value = profit_detail.get('profit_margin')
                    
                    profit_str = f"{profit_value:.2f}€" if profit_value is not None else "N/A"
                    margin_str = f"{profit_margin_value:.1f}%" if profit_margin_value is not None else "N/A"
                    
                    profit_text = f"Lucro (período): {profit_str}, Margem Média (período): {margin_str}"
                    story.append(Paragraph(profit_text, styles['TableCellText']))
                
                if client_data.get('top_tasks') and client_data['top_tasks']:
                    story.append(Paragraph("Tarefas Recentes (Top 5 no período)", styles['SubSectionTitle']))
                    task_headers = [Paragraph(h, styles['TableHeader']) for h in ['Título', 'Prioridade', 'Status', 'Prazo']]
                    task_table_data = [task_headers]
                    for task in client_data['top_tasks']:
                        task_table_data.append([
                            Paragraph(task.title[:50], styles['TableCellTextSmall']), # Truncate long titles
                            Paragraph(task.get_priority_display(), styles['TableCellCenter']),
                            Paragraph(task.get_status_display(), styles['TableCellCenter']),
                            Paragraph(task.deadline.strftime('%d/%m/%Y') if task.deadline else 'N/A', styles['TableCellCenter']),
                        ])
                    task_table = Table(task_table_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch], repeatRows=1)
                    task_table.setStyle(ReportGenerationService._get_default_table_style(C, num_data_rows=len(client_data['top_tasks'])))
                    story.append(task_table)
                    story.append(Spacer(1, 0.1*inch))
                
                if client_data.get('recent_time_entries') and client_data['recent_time_entries']:
                    story.append(Paragraph("Registos de Tempo Recentes (Top 5 no período)", styles['SubSectionTitle']))
                    time_headers = [Paragraph(h, styles['TableHeader']) for h in ['Data', 'Utilizador', 'Minutos', 'Descrição']]
                    time_table_data = [time_headers]
                    for entry in client_data['recent_time_entries']:
                        time_table_data.append([
                            Paragraph(entry.date.strftime('%d/%m/%Y'), styles['TableCellCenter']),
                            Paragraph(entry.user.username, styles['TableCellCenter']),
                            Paragraph(str(entry.minutes_spent), styles['TableCellNumber']),
                            Paragraph(entry.description[:50] + ('...' if len(entry.description) > 50 else ''), styles['TableCellTextSmall']),
                        ])
                    time_table = Table(time_table_data, colWidths=[0.8*inch, 1.2*inch, 0.7*inch, 2.8*inch], repeatRows=1)
                    time_table.setStyle(ReportGenerationService._get_default_table_style(C, num_data_rows=len(client_data['recent_time_entries'])))
                    story.append(time_table)

        ReportGenerationService._build_pdf_doc(buffer, story, title="Resumo de Clientes", org_name=data['organization'].name)
        buffer.seek(0)
        return buffer, 'application/pdf'

    @staticmethod
    def _generate_client_summary_csv(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        writer.writerow(['Relatório de Resumo de Clientes'])
        writer.writerow([f"Organização: {data['organization'].name}"])
        writer.writerow([f"Gerado em: {data['generation_date'].strftime('%d/%m/%Y %H:%M')}"])
        if data.get('date_from') and data.get('date_to'): writer.writerow([f"Período: {data['date_from'].strftime('%d/%m/%Y')} a {data['date_to'].strftime('%d/%m/%Y')}"])
        writer.writerow([])
        
        if not data['clients_data']:
            writer.writerow(["Nenhum cliente encontrado para os critérios selecionados."])
        else:
            writer.writerow(['Nome Cliente', 'NIF', 'Email', 'Telefone', 'Morada', 'Gestor de Conta', 'Avença Mensal (€)', 'Tarefas Ativas (Período)', 'Tarefas Concluídas (Período)', 'Tempo Total (min)', 'Lucro Agregado (Período)', 'Margem Média (%)'])
            for client_detail in data['clients_data']:
                client_obj = client_detail['client_obj']
                profit = client_detail.get('recent_profitability')
                writer.writerow([
                    client_obj.name, client_obj.nif or '', client_obj.email or '', client_obj.phone or '', client_obj.address or '',
                    client_obj.account_manager.username if client_obj.account_manager else '',
                    float(client_obj.monthly_fee or 0), client_detail['active_tasks_count'], client_detail['completed_tasks_count'],
                    client_detail['total_time_minutes'] or 0,
                    float(profit['profit']) if profit and profit.get('profit') is not None else 'N/A',
                    float(profit['profit_margin']) if profit and profit.get('profit_margin') is not None else 'N/A',
                ])
        
        csv_content = buffer.getvalue()
        buffer.close()
        return io.BytesIO(csv_content.encode('utf-8-sig')), 'text/csv'

    @staticmethod
    def _generate_client_summary_xlsx(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        
        ws_summary = wb.active
        ws_summary.title = "Sumário Geral"
        ws_summary['A1'] = f"Relatório de Resumo de Clientes - {data['organization'].name}"
        ws_summary.merge_cells('A1:F1'); ReportGenerationService._apply_xlsx_header_style(ws_summary['A1'], font_bold=True, fill_color="1E40AF")
        ws_summary['A2'] = f"Gerado em: {data['generation_date'].strftime('%d/%m/%Y %H:%M')}"
        if data.get('date_from') and data.get('date_to'): ws_summary['A3'] = f"Período: {data['date_from'].strftime('%d/%m/%Y')} a {data['date_to'].strftime('%d/%m/%Y')}"
        
        ws_summary['A5'] = "Total de Clientes Analisados:"; ReportGenerationService._apply_xlsx_cell_style(ws_summary['A5'], font_bold=True)
        ws_summary['B5'] = data['total_clients']; ReportGenerationService._apply_xlsx_cell_style(ws_summary['B5'])
        ws_summary['A6'] = "Avenças Totais (€):"; ReportGenerationService._apply_xlsx_cell_style(ws_summary['A6'], font_bold=True)
        ws_summary['B6'] = ReportGenerationService._format_currency_excel(data['total_monthly_fees']); ReportGenerationService._apply_xlsx_cell_style(ws_summary['B6'], number_format='#,##0.00€')

        if not data['clients_data']:
            ws_summary['A8'] = "Nenhum cliente encontrado para os critérios selecionados."
        else:
            ws_details = wb.create_sheet("Detalhes Clientes")
            headers = ['Cliente', 'NIF', 'Email', 'Telefone', 'Morada', 'Gestor', 'Avença (€)', 'Tarefas Ativas', 'Tarefas Concluídas', 'Tempo (min)', 'Lucro (€)', 'Margem (%)']
            for col_idx, header_text in enumerate(headers, 1):
                ReportGenerationService._apply_xlsx_header_style(ws_details.cell(row=1, column=col_idx, value=header_text))

            for row_idx, client_detail in enumerate(data['clients_data'], 2):
                client_obj = client_detail['client_obj']
                profit = client_detail.get('recent_profitability')
                
                row_values = [
                    client_obj.name, client_obj.nif, client_obj.email, client_obj.phone, client_obj.address,
                    client_obj.account_manager.username if client_obj.account_manager else '',
                    ReportGenerationService._format_currency_excel(client_obj.monthly_fee),
                    client_detail['active_tasks_count'], client_detail['completed_tasks_count'],
                    client_detail['total_time_minutes'] or 0,
                    ReportGenerationService._format_currency_excel(profit['profit'] if profit and profit.get('profit') is not None else None),
                    ReportGenerationService._format_percentage_excel(profit['profit_margin'] if profit and profit.get('profit_margin') is not None else None)
                ]
                for col_idx, cell_value in enumerate(row_values, 1):
                    cell = ws_details.cell(row=row_idx, column=col_idx, value=cell_value)
                    if col_idx == 7 or col_idx == 11: cell.number_format = '#,##0.00€'
                    elif col_idx == 12: cell.number_format = '0.00%'
                    ReportGenerationService._apply_xlsx_cell_style(cell)
            
            for sheet_name_key in [ws_summary.title, ws_details.title]:
                current_sheet = wb[sheet_name_key]
                for i, column_cells in enumerate(current_sheet.columns):
                    length = 0
                    if column_cells:
                        str_values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
                        if str_values: length = max(len(val) for val in str_values)
                    current_sheet.column_dimensions[get_column_letter(i + 1)].width = length + 3

        wb.save(buffer)
        buffer.seek(0)
        return buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
    # --- Report: Profitability Analysis ---
    @staticmethod
    def generate_profitability_analysis_report(
        organization: Organization, client_ids: List[str] = None,
        year: int = None, month: int = None, format_type: str = 'pdf'
    ) -> Tuple[io.BytesIO, str]:
        profit_query = ClientProfitability.objects.filter(client__organization=organization)
        if client_ids: profit_query = profit_query.filter(client_id__in=client_ids)
        
        current_period_label = "Geral (Todos os Períodos)"
        if year and month:
            profit_query = profit_query.filter(year=year, month=month)
            try:
                current_period_label = f"{datetime(year,month,1).strftime('%B %Y')}"
            except ValueError: 
                current_period_label = f"{month}/{year}"
        elif year:
            profit_query = profit_query.filter(year=year)
            current_period_label = f"Ano {year}"

        records = profit_query.select_related('client').order_by('-profit_margin')
        
        stats = records.aggregate(
            total_profit=Sum('profit'), avg_margin=Avg('profit_margin'),
            profitable_count=Count('id', filter=Q(is_profitable=True)),
            unprofitable_count=Count('id', filter=Q(is_profitable=False))
        )
        
        report_data = {
            'organization': organization, 'profitability_records': list(records), 
            'stats': stats, 'generation_date': timezone.now(), 
            'total_records': records.count(), 'month_name': current_period_label
        }

        if format_type == 'pdf':
            return ReportGenerationService._generate_profitability_pdf(report_data)
        elif format_type == 'csv':
            return ReportGenerationService._generate_profitability_csv(report_data)
        elif format_type == 'xlsx':
            return ReportGenerationService._generate_profitability_xlsx(report_data)
        raise ValueError(f"Formato não suportado: {format_type}")

    @staticmethod
    def _generate_profitability_pdf(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.BytesIO()
        styles, C = ReportGenerationService._get_pdf_styles()
        story = []
        story.append(Paragraph(f"Análise de Rentabilidade - {data['month_name']}", styles['ReportTitle']))
        story.append(Paragraph(data['organization'].name, styles['ReportSubtitle']))
        story.append(Paragraph(f"Gerado em: {data['generation_date'].strftime('%d/%m/%Y %H:%M')}", styles['SmallTextMuted']))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C['border_color'], spaceBefore=5, spaceAfter=15))
        
        story.append(Paragraph("Sumário de Rentabilidade", styles['SectionTitle']))
        stats = data['stats']
        summary_data = [
            [Paragraph('Total de Registos:', styles['TableCellText']), Paragraph(str(data['total_records']), styles['TableCellNumber'])],
            [Paragraph('Clientes Rentáveis:', styles['TableCellText']), Paragraph(str(stats.get('profitable_count',0) or 0), styles['TableCellNumber'])],
            [Paragraph('Clientes Não Rentáveis:', styles['TableCellText']), Paragraph(str(stats.get('unprofitable_count',0) or 0), styles['TableCellNumber'])],
            [Paragraph('Lucro Total (Período):', styles['TableCellText']), Paragraph(f"€{stats.get('total_profit', 0):.2f}" if stats.get('total_profit') is not None else 'N/A', styles['TableCellNumber'])],
            [Paragraph('Margem Média (%):', styles['TableCellText']), Paragraph(f"{stats.get('avg_margin', 0):.1f}%" if stats.get('avg_margin') is not None else 'N/A', styles['TableCellNumber'])],
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, C['border_color']), ('BACKGROUND', (0,0), (0,-1), C['bg_light'])]))
        story.append(summary_table)
        story.append(Spacer(1, 0.2*inch))

        if not data['profitability_records']:
            story.append(Paragraph("Nenhum dado de rentabilidade encontrado para este período.", styles['BodyText']))
        else:
            story.append(Paragraph("Detalhes por Cliente", styles['SectionTitle']))
            headers = [Paragraph(h, styles['TableHeader']) for h in ['Cliente', 'Avença', 'Custo Tempo', 'Despesas', 'Lucro', 'Margem %', 'Status']]
            table_data = [headers]
            for record in data['profitability_records']:
                status_text = 'Rentável' if record.is_profitable else ('Não Rentável' if record.is_profitable == False else 'N/D')
                status_color = C['green_dark_text'] if record.is_profitable else (C['red_dark_text'] if record.is_profitable == False else C['text_muted'])
                status_style = ParagraphStyle(name=f'StatusCell_Profit_{record.id}', parent=styles['TableCellCenter'], textColor=status_color)
                
                table_data.append([
                    Paragraph(record.client.name, styles['TableCellTextSmall']),
                    Paragraph(f"€{record.monthly_fee:.2f}", styles['TableCellNumber']),
                    Paragraph(f"€{record.time_cost:.2f}", styles['TableCellNumber']),
                    Paragraph(f"€{record.total_expenses:.2f}", styles['TableCellNumber']),
                    Paragraph(f"€{record.profit:.2f}" if record.profit is not None else "N/A", styles['TableCellNumber']),
                    Paragraph(f"{record.profit_margin:.1f}%" if record.profit_margin is not None else "N/A", styles['TableCellCenter']),
                    Paragraph(status_text, status_style)
                ])
            detail_table = Table(table_data, colWidths=[1.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.8*inch], repeatRows=1)
            detail_table.setStyle(ReportGenerationService._get_default_table_style(C, num_data_rows=len(data['profitability_records'])))
            story.append(detail_table)

        ReportGenerationService._build_pdf_doc(buffer, story, title=f"Análise Rentabilidade {data['month_name']}", org_name=data['organization'].name)
        buffer.seek(0)
        return buffer, 'application/pdf'

    @staticmethod
    def _generate_profitability_csv(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(['Análise de Rentabilidade'])
        writer.writerow([f"Organização: {data['organization'].name}"])
        writer.writerow([f"Período: {data['month_name']}"])
        writer.writerow([f"Gerado em: {data['generation_date'].strftime('%d/%m/%Y %H:%M')}"])
        writer.writerow([])
        
        stats = data['stats']
        writer.writerow(['Estatísticas Resumo'])
        writer.writerow(['Total de Registos:', data['total_records']])
        writer.writerow(['Clientes Rentáveis:', stats.get('profitable_count',0) or 0])
        writer.writerow(['Clientes Não Rentáveis:', stats.get('unprofitable_count',0) or 0])
        writer.writerow(['Lucro Total (€):', float(stats.get('total_profit',0)) if stats.get('total_profit') is not None else 0])
        writer.writerow(['Margem Média (%):', float(stats.get('avg_margin',0)) if stats.get('avg_margin') is not None else 0])
        writer.writerow([])
        
        if not data['profitability_records']:
            writer.writerow(["Nenhum dado de rentabilidade encontrado."])
        else:
            writer.writerow(['Cliente', 'Ano', 'Mês', 'Avença (€)', 'Custo Tempo (€)', 'Despesas (€)', 'Lucro (€)', 'Margem (%)', 'Rentável'])
            for record in data['profitability_records']:
                writer.writerow([
                    record.client.name, record.year, record.month,
                    float(record.monthly_fee), float(record.time_cost), float(record.total_expenses),
                    float(record.profit) if record.profit is not None else 'N/A',
                    float(record.profit_margin) if record.profit_margin is not None else 'N/A',
                    'Sim' if record.is_profitable else ('Não' if record.is_profitable == False else 'N/D')
                ])
        csv_content = buffer.getvalue()
        buffer.close()
        return io.BytesIO(csv_content.encode('utf-8-sig')), 'text/csv'

    @staticmethod
    def _generate_profitability_xlsx(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Sumário Rentabilidade"
        ws_summary['A1'] = f"Análise de Rentabilidade - {data['month_name']} ({data['organization'].name})"
        ws_summary.merge_cells('A1:E1'); ReportGenerationService._apply_xlsx_header_style(ws_summary['A1'], font_bold=True, fill_color="1E40AF")
        
        row_idx = 3
        stats = data['stats']
        summary_items = [
            ("Total de Registos:", data['total_records']),
            ("Clientes Rentáveis:", stats.get('profitable_count',0) or 0),
            ("Clientes Não Rentáveis:", stats.get('unprofitable_count',0) or 0),
            ("Lucro Total (€):", ReportGenerationService._format_currency_excel(stats.get('total_profit'))),
            ("Margem Média (%):", ReportGenerationService._format_percentage_excel(stats.get('avg_margin'))),
        ]
        for label, value in summary_items:
            ReportGenerationService._apply_xlsx_cell_style(ws_summary.cell(row=row_idx, column=1, value=label), font_bold=True)
            cell = ws_summary.cell(row=row_idx, column=2, value=value)
            if "€" in label: cell.number_format = '#,##0.00€'
            elif "%" in label: cell.number_format = '0.00%'
            ReportGenerationService._apply_xlsx_cell_style(cell)
            row_idx +=1

        # Detailed Data Sheet
        if data['profitability_records']:
            ws_details = wb.create_sheet("Detalhes Rentabilidade")
            headers = ['Cliente', 'Ano', 'Mês', 'Avença (€)', 'Custo Tempo (€)', 'Despesas (€)', 'Lucro (€)', 'Margem (%)', 'Rentável']
            for col_idx, header_text in enumerate(headers, 1):
                ReportGenerationService._apply_xlsx_header_style(ws_details.cell(row=1, column=col_idx, value=header_text))

            for row_idx_det, record in enumerate(data['profitability_records'], 2):
                row_values = [
                    record.client.name, record.year, record.month,
                    ReportGenerationService._format_currency_excel(record.monthly_fee),
                    ReportGenerationService._format_currency_excel(record.time_cost),
                    ReportGenerationService._format_currency_excel(record.total_expenses),
                    ReportGenerationService._format_currency_excel(record.profit),
                    ReportGenerationService._format_percentage_excel(record.profit_margin),
                    'Sim' if record.is_profitable else ('Não' if record.is_profitable == False else 'N/D')
                ]
                for col_idx, cell_value in enumerate(row_values, 1):
                    cell = ws_details.cell(row=row_idx_det, column=col_idx, value=cell_value)
                    if col_idx in [4,5,6,7]: cell.number_format = '#,##0.00€'
                    elif col_idx == 8: cell.number_format = '0.00%'
                    ReportGenerationService._apply_xlsx_cell_style(cell)
            
            # Bar Chart for Top 5 Profitable Clients
            if len(data['profitability_records']) > 0:
                chart_sheet = wb.create_sheet("Top Rentáveis")
                top_profitable = sorted([r for r in data['profitability_records'] if r.profit is not None], key=lambda x: x.profit, reverse=True)[:10]
                chart_sheet.cell(row=1, column=1, value="Cliente")
                chart_sheet.cell(row=1, column=2, value="Lucro (€)")
                for i, rec in enumerate(top_profitable, 2):
                    chart_sheet.cell(row=i, column=1, value=rec.client.name)
                    chart_sheet.cell(row=i, column=2, value=ReportGenerationService._format_currency_excel(rec.profit))

                bar_chart = OpenpyxlBarChart()
                bar_chart.title = "Top 10 Clientes Mais Rentáveis"
                bar_chart.style = 10
                labels_ref = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(top_profitable)+1)
                data_ref = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(top_profitable)+1)
                bar_chart.add_data(data_ref, titles_from_data=True)
                bar_chart.set_categories(labels_ref)
                bar_chart.y_axis.title = "Lucro (€)"
                bar_chart.x_axis.title = "Cliente"
                chart_sheet.add_chart(bar_chart, "D2")
        
        for sheet_name_key in wb.sheetnames:
            current_sheet = wb[sheet_name_key]
            for i, column_cells in enumerate(current_sheet.columns):
                length = 0
                if column_cells:
                    str_values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
                    if str_values: length = max(len(val) for val in str_values)
                current_sheet.column_dimensions[get_column_letter(i + 1)].width = max(length + 2, 12) # Min width 12

        wb.save(buffer)
        buffer.seek(0)
        return buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
    # --- Report: Time Tracking Summary ---
    @staticmethod
    def generate_time_tracking_summary_report(
        organization: Organization, user_ids: List[str] = None, client_ids: List[str] = None,
        date_from: datetime = None, date_to: datetime = None, format_type: str = 'pdf'
    ) -> Tuple[io.BytesIO, str]:
        time_query = TimeEntry.objects.filter(client__organization=organization)
        if user_ids: time_query = time_query.filter(user_id__in=user_ids)
        if client_ids: time_query = time_query.filter(client_id__in=client_ids)
        
        actual_date_from = date_from or (timezone.now() - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        actual_date_to = date_to or timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        time_query = time_query.filter(date__gte=actual_date_from.date(), date__lte=actual_date_to.date())

        time_entries = list(time_query.select_related('user', 'client', 'task', 'category').order_by('-date', 'user__username')[:500]) 

        total_stats = time_query.aggregate(total_minutes=Sum('minutes_spent'), total_entries=Count('id'), unique_users=Count('user', distinct=True), unique_clients=Count('client', distinct=True))
        user_stats = list(time_query.values('user__username').annotate(total_minutes=Sum('minutes_spent'), entry_count=Count('id'), client_count=Count('client', distinct=True)).order_by('-total_minutes'))
        client_stats = list(time_query.values('client__name').annotate(total_minutes=Sum('minutes_spent'), entry_count=Count('id'), user_count=Count('user', distinct=True)).order_by('-total_minutes'))
        category_stats = list(time_query.filter(category__isnull=False).values('category__name').annotate(total_minutes=Sum('minutes_spent'), entry_count=Count('id')).order_by('-total_minutes'))

        report_data = {
            'organization': organization, 'time_entries': time_entries, 'total_stats': total_stats,
            'user_stats': user_stats, 'client_stats': client_stats, 'category_stats': category_stats,
            'date_from': actual_date_from, 'date_to': actual_date_to,
            'generation_date': timezone.now()
        }
        if format_type == 'pdf':
            return ReportGenerationService._generate_time_tracking_pdf(report_data)
        elif format_type == 'csv':
            return ReportGenerationService._generate_time_tracking_csv(report_data)
        elif format_type == 'xlsx':
            return ReportGenerationService._generate_time_tracking_xlsx(report_data)
        raise ValueError(f"Formato não suportado: {format_type}")

    @staticmethod
    def _generate_time_tracking_pdf(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.BytesIO()
        styles, C = ReportGenerationService._get_pdf_styles()
        story = []
        story.append(Paragraph("Resumo de Registo de Tempos", styles['ReportTitle']))
        story.append(Paragraph(data['organization'].name, styles['ReportSubtitle']))
        story.append(Paragraph(f"Período: {data['date_from'].strftime('%d/%m/%Y')} a {data['date_to'].strftime('%d/%m/%Y')}", styles['SmallTextMuted']))
        story.append(Paragraph(f"Gerado em: {data['generation_date'].strftime('%d/%m/%Y %H:%M')}", styles['SmallTextMuted']))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C['border_color'], spaceBefore=5, spaceAfter=15))

        story.append(Paragraph("Sumário Geral", styles['SectionTitle']))
        ts = data['total_stats']
        summary_data = [
            [Paragraph('Total de Registos:', styles['TableCellText']), Paragraph(str(ts.get('total_entries',0) or 0), styles['TableCellNumber'])],
            [Paragraph('Total de Minutos:', styles['TableCellText']), Paragraph(str(ts.get('total_minutes',0) or 0), styles['TableCellNumber'])],
            [Paragraph('Total de Horas:', styles['TableCellText']), Paragraph(f"{(ts.get('total_minutes',0) or 0) / 60:.1f}h", styles['TableCellNumber'])],
            [Paragraph('Nº Utilizadores Ativos:', styles['TableCellText']), Paragraph(str(ts.get('unique_users',0) or 0), styles['TableCellNumber'])],
            [Paragraph('Nº Clientes Envolvidos:', styles['TableCellText']), Paragraph(str(ts.get('unique_clients',0) or 0), styles['TableCellNumber'])],
        ]
        story.append(Table(summary_data, colWidths=[3*inch, 2.5*inch], style=TableStyle([('GRID', (0,0), (-1,-1), 0.5, C['border_color']), ('BACKGROUND', (0,0), (0,-1), C['bg_light'])])))
        story.append(Spacer(1, 0.2*inch))

        for title, stats_list, val_key, count_key, name_key, item_label in [
            ("Tempo por Utilizador", data['user_stats'], 'total_minutes', 'entry_count', 'user__username', "Utilizador"),
            ("Tempo por Cliente", data['client_stats'], 'total_minutes', 'entry_count', 'client__name', "Cliente"),
            ("Tempo por Categoria", data['category_stats'], 'total_minutes', 'entry_count', 'category__name', "Categoria")
        ]:
            if stats_list:
                story.append(Paragraph(title, styles['SubSectionTitle']))
                headers = [Paragraph(h, styles['TableHeader']) for h in [item_label, "Total Minutos", "Nº Registos"]]
                table_data = [headers]
                for item in stats_list[:10]: # Top 10
                    table_data.append([
                        Paragraph(item[name_key], styles['TableCellTextSmall']),
                        Paragraph(str(item.get(val_key,0) or 0), styles['TableCellNumber']),
                        Paragraph(str(item.get(count_key,0) or 0), styles['TableCellNumber']),
                    ])
                story.append(Table(table_data, colWidths=[3*inch, 1.25*inch, 1.25*inch], repeatRows=1, style=ReportGenerationService._get_default_table_style(C, len(stats_list[:10]))))
                story.append(Spacer(1, 0.1*inch))
        
        if data['time_entries']:
            story.append(PageBreak())
            story.append(Paragraph(f"Lista Detalhada de Registos (Primeiros {len(data['time_entries'])})", styles['SectionTitle']))
            detail_headers = [Paragraph(h, styles['TableHeader']) for h in ['Data', 'Utilizador', 'Cliente', 'Tarefa', 'Min', 'Descrição']]
            detail_table_data = [detail_headers]
            for entry in data['time_entries']:
                detail_table_data.append([
                    Paragraph(entry.date.strftime('%d/%m/%y'), styles['TableCellCenter']),
                    Paragraph(entry.user.username, styles['TableCellTextSmall']),
                    Paragraph(entry.client.name, styles['TableCellTextSmall']),
                    Paragraph(entry.task.title[:20] if entry.task else '-', styles['TableCellTextSmall']),
                    Paragraph(str(entry.minutes_spent), styles['TableCellNumber']),
                    Paragraph(entry.description[:40] + ('...' if len(entry.description) > 40 else ''), styles['TableCellTextSmall']),
                ])
            detail_table = Table(detail_table_data, colWidths=[0.6*inch, 1*inch, 1.2*inch, 1.2*inch, 0.5*inch, 2*inch], repeatRows=1)
            detail_table.setStyle(ReportGenerationService._get_default_table_style(C, len(data['time_entries'])))
            story.append(detail_table)

        ReportGenerationService._build_pdf_doc(buffer, story, title="Resumo de Tempos", org_name=data['organization'].name, pagesize=landscape(A4))
        buffer.seek(0)
        return buffer, 'application/pdf'
        
    @staticmethod
    def _generate_time_tracking_csv(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Resumo de Registo de Tempos", f"Organização: {data['organization'].name}"])
        writer.writerow([f"Período: {data['date_from'].strftime('%d/%m/%Y')} a {data['date_to'].strftime('%d/%m/%Y')}"])
        writer.writerow([f"Gerado em: {data['generation_date'].strftime('%d/%m/%Y %H:%M')}"])
        writer.writerow([])
        
        ts = data['total_stats']
        writer.writerow(['Sumário Geral'])
        writer.writerow(['Total de Registos:', ts.get('total_entries',0) or 0])
        writer.writerow(['Total de Minutos:', ts.get('total_minutes',0) or 0])
        writer.writerow(['Total de Horas:', f"{(ts.get('total_minutes',0) or 0) / 60:.1f}"])
        writer.writerow(['Utilizadores Únicos:', ts.get('unique_users',0) or 0])
        writer.writerow(['Clientes Únicos:', ts.get('unique_clients',0) or 0])
        writer.writerow([])
        
        writer.writerow(['Registos Detalhados'])
        writer.writerow(['Data', 'Utilizador', 'Cliente', 'Tarefa', 'Categoria', 'Descrição', 'Minutos', 'Horas'])
        for entry in data['time_entries']: 
            writer.writerow([
                entry.date.strftime('%Y-%m-%d'), entry.user.username, entry.client.name,
                entry.task.title if entry.task else '',
                entry.category.name if entry.category else '',
                entry.description, entry.minutes_spent, round(entry.minutes_spent / 60, 2)
            ])
        csv_content = buffer.getvalue()
        buffer.close()
        return io.BytesIO(csv_content.encode('utf-8-sig')), 'text/csv'

    @staticmethod
    def _generate_time_tracking_xlsx(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Sumário Tempos"
        ws_summary['A1'] = f"Resumo de Registo de Tempos - {data['organization'].name}"
        ws_summary.merge_cells('A1:E1'); ReportGenerationService._apply_xlsx_header_style(ws_summary['A1'], font_bold=True, fill_color="1E40AF")
        # ... Add overall summary stats (total_stats) to this sheet ...
        ts = data['total_stats']
        row_idx = 3
        ws_summary.cell(row=row_idx, column=1, value="Período:"); ReportGenerationService._apply_xlsx_cell_style(ws_summary.cell(row=row_idx,column=1), font_bold=True)
        ws_summary.cell(row=row_idx, column=2, value=f"{data['date_from'].strftime('%d/%m/%Y')} - {data['date_to'].strftime('%d/%m/%Y')}"); row_idx+=1
        ws_summary.cell(row=row_idx, column=1, value="Total Registos:"); ReportGenerationService._apply_xlsx_cell_style(ws_summary.cell(row=row_idx,column=1), font_bold=True)
        ws_summary.cell(row=row_idx, column=2, value=ts.get('total_entries',0) or 0); row_idx+=1
        ws_summary.cell(row=row_idx, column=1, value="Total Horas:"); ReportGenerationService._apply_xlsx_cell_style(ws_summary.cell(row=row_idx,column=1), font_bold=True)
        ws_summary.cell(row=row_idx, column=2, value=(ts.get('total_minutes',0) or 0)/60).number_format = '0.00 "h"'; row_idx+=2

        # Sheet for User Stats
        if data['user_stats']:
            ws_user = wb.create_sheet("Tempo por Utilizador")
            headers_user = ["Utilizador", "Total Minutos", "Nº Registos", "Nº Clientes"]
            for col_idx, header in enumerate(headers_user,1): ReportGenerationService._apply_xlsx_header_style(ws_user.cell(row=1, column=col_idx, value=header))
            for r_idx, stat in enumerate(data['user_stats'], 2):
                ws_user.cell(row=r_idx, column=1, value=stat['user__username'])
                ws_user.cell(row=r_idx, column=2, value=stat.get('total_minutes',0) or 0)
                ws_user.cell(row=r_idx, column=3, value=stat.get('entry_count',0) or 0)
                ws_user.cell(row=r_idx, column=4, value=stat.get('client_count',0) or 0)
                for c_idx in range(1, len(headers_user)+1): ReportGenerationService._apply_xlsx_cell_style(ws_user.cell(row=r_idx, column=c_idx))

        # Sheet for Client Stats & Category Stats (similar to User Stats)
        # ...

        # Detailed Time Entries Sheet
        ws_details = wb.create_sheet("Registos Detalhados")
        headers_details = ['Data', 'Utilizador', 'Cliente', 'Tarefa', 'Categoria', 'Descrição', 'Minutos']
        for col_idx, header in enumerate(headers_details,1): ReportGenerationService._apply_xlsx_header_style(ws_details.cell(row=1, column=col_idx, value=header))
        for r_idx, entry in enumerate(data['time_entries'], 2):
            row_values = [
                entry.date, entry.user.username, entry.client.name,
                entry.task.title if entry.task else '', entry.category.name if entry.category else '',
                entry.description, entry.minutes_spent
            ]
            for c_idx, val in enumerate(row_values,1):
                cell = ws_details.cell(row=r_idx, column=c_idx, value=val)
                if c_idx == 1: cell.number_format = 'YYYY-MM-DD'
                ReportGenerationService._apply_xlsx_cell_style(cell, wrap_text=(c_idx==6))

        for sheet_name_key in wb.sheetnames:
            current_sheet = wb[sheet_name_key]
            for i, column_cells in enumerate(current_sheet.columns):
                length = 0
                if column_cells:
                    str_values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
                    if str_values: length = max(len(val) for val in str_values)
                current_sheet.column_dimensions[get_column_letter(i + 1)].width = max(length + 2, 12)

        wb.save(buffer)
        buffer.seek(0)
        return buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # --- Report: Task Performance (New) ---
    @staticmethod
    def generate_task_performance_report(
        organization: Organization, date_from: datetime = None, date_to: datetime = None,
        client_ids: List[str] = None, user_ids: List[str] = None, 
        category_ids: List[str] = None, statuses: List[str] = None,
        format_type: str = 'pdf'
    ) -> Tuple[io.BytesIO, str]:
        task_query = Task.objects.filter(client__organization=organization)

        actual_date_from = date_from or (timezone.now() - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        actual_date_to = date_to or timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        
        task_query = task_query.filter(created_at__gte=actual_date_from, created_at__lte=actual_date_to)

        if client_ids: task_query = task_query.filter(client_id__in=client_ids)
        if user_ids: task_query = task_query.filter(Q(assigned_to_id__in=user_ids) | Q(collaborators__id__in=user_ids)).distinct()
        if category_ids: task_query = task_query.filter(category_id__in=category_ids)
        if statuses: task_query = task_query.filter(status__in=statuses)
        
        tasks = list(task_query.select_related('client', 'category', 'assigned_to', 'created_by', 'workflow').order_by('-created_at')[:500])

        total_tasks = task_query.count()
        status_distribution = list(task_query.values('status').annotate(count=Count('id')).order_by('-count'))
        category_distribution = list(task_query.filter(category__isnull=False).values('category__name').annotate(count=Count('id')).order_by('-count'))
        assignee_distribution = list(task_query.filter(assigned_to__isnull=False).values('assigned_to__username').annotate(count=Count('id')).order_by('-count'))
        
        completed_tasks_in_filtered_set = task_query.filter(status='completed', completed_at__isnull=False, created_at__isnull=False)
        
        avg_completion_time_data = completed_tasks_in_filtered_set.annotate(
            duration=ExpressionWrapper(F('completed_at') - F('created_at'), output_field=DurationField())
        ).aggregate(avg_duration=Avg('duration'))
        
        avg_completion_seconds = avg_completion_time_data['avg_duration'].total_seconds() if avg_completion_time_data['avg_duration'] else None
        
        overdue_tasks_count = task_query.filter(deadline__lt=timezone.now().date(), status__in=['pending', 'in_progress']).count()

        report_data = {
            'organization': organization, 'tasks': tasks, 'generation_date': timezone.now(),
            'date_from': actual_date_from, 'date_to': actual_date_to, 
            'filters_applied': {
                'clients': Client.objects.filter(id__in=client_ids).values_list('name', flat=True) if client_ids else "Todos",
                'users': User.objects.filter(id__in=user_ids).values_list('username', flat=True) if user_ids else "Todos",
                'categories': TaskCategory.objects.filter(id__in=category_ids).values_list('name', flat=True) if category_ids else "Todas",
                'statuses': statuses or "Todos"
            },
            'summary_stats': {
                'total_tasks': total_tasks, 'status_distribution': status_distribution,
                'overdue_tasks_count': overdue_tasks_count, 
                'avg_completion_days': (avg_completion_seconds / (24*3600)) if avg_completion_seconds is not None else None,
            },
            'category_distribution': category_distribution,
            'assignee_distribution': assignee_distribution,
        }

        if format_type == 'pdf':
            return ReportGenerationService._generate_task_performance_pdf(report_data)
        elif format_type == 'csv':
            return ReportGenerationService._generate_task_performance_csv(report_data)
        elif format_type == 'xlsx':
            return ReportGenerationService._generate_task_performance_xlsx(report_data)
        raise ValueError(f"Formato não suportado: {format_type}")

    @staticmethod
    def _generate_task_performance_pdf(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.BytesIO()
        styles, C = ReportGenerationService._get_pdf_styles()
        story = []
        
        story.append(Paragraph("Relatório de Performance de Tarefas", styles['ReportTitle']))
        story.append(Paragraph(data['organization'].name, styles['ReportSubtitle']))
        story.append(Paragraph(f"Gerado em: {data['generation_date'].strftime('%d/%m/%Y %H:%M')}", styles['SmallTextMuted']))
        if data['date_from'] and data['date_to']: story.append(Paragraph(f"Período Tarefas Criadas: {data['date_from'].strftime('%d/%m/%Y')} a {data['date_to'].strftime('%d/%m/%Y')}", styles['SmallTextMuted']))
        story.append(Spacer(1, 0.2*inch))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C['border_color'], spaceBefore=5, spaceAfter=15))

        story.append(Paragraph("Sumário Geral", styles['SectionTitle']))
        summary = data['summary_stats']
        summary_data = [
            [Paragraph("Total de Tarefas (filtros):", styles['TableCellText']), Paragraph(str(summary['total_tasks']), styles['TableCellNumber'])],
            [Paragraph("Tarefas Atrasadas:", styles['TableCellText']), Paragraph(str(summary['overdue_tasks_count']), styles['TableCellNumber'])],
            [Paragraph("Tempo Médio Conclusão (dias):", styles['TableCellText']), Paragraph(f"{summary['avg_completion_days']:.1f}" if summary['avg_completion_days'] is not None else "N/A", styles['TableCellNumber'])],
        ]
        story.append(Table(summary_data, colWidths=[3*inch, 2.5*inch], style=TableStyle([('GRID', (0,0), (-1,-1), 0.5, C['border_color']), ('BACKGROUND', (0,0), (0,-1), C['bg_light'])])))
        story.append(Spacer(1, 0.1*inch))

        story.append(Paragraph("Distribuição por Status", styles['SubSectionTitle']))
        status_headers = [Paragraph(h, styles['TableHeader']) for h in ["Status", "Quantidade"]]
        status_table_data = [status_headers]
        for s_data in summary['status_distribution']:
            status_display = dict(Task.STATUS_CHOICES).get(s_data['status'], s_data['status'])
            status_table_data.append([Paragraph(status_display, styles['TableCellText']), Paragraph(str(s_data['count']), styles['TableCellNumber'])])
        story.append(Table(status_table_data, colWidths=[2*inch, 1*inch], repeatRows=1, style=ReportGenerationService._get_default_table_style(C, num_data_rows=len(summary['status_distribution']))))
        story.append(Spacer(1, 0.2*inch))
        
        # (Similar tables for category_distribution and assignee_distribution)

        story.append(Paragraph("Análise e Recomendações", styles['SectionTitle']))
        # ... (Analysis text based on data) ...
        story.append(PageBreak())
        
        story.append(Paragraph(f"Lista Detalhada de Tarefas (até {len(data['tasks'])} de {summary['total_tasks']})", styles['SectionTitle']))
        task_detail_headers = [Paragraph(h, styles['TableHeader']) for h in ['Título', 'Cliente', 'Status', 'Prioridade', 'Prazo', 'Responsável']]
        task_detail_data = [task_detail_headers]
        for task in data['tasks']:
            task_detail_data.append([
                Paragraph(task.title[:40], styles['TableCellTextSmall']),
                Paragraph(task.client.name[:25], styles['TableCellTextSmall']),
                Paragraph(task.get_status_display(), styles['TableCellCenter']),
                Paragraph(task.get_priority_display(), styles['TableCellCenter']),
                Paragraph(task.deadline.strftime('%d/%m/%y') if task.deadline else '-', styles['TableCellCenter']),
                Paragraph(task.assigned_to.username if task.assigned_to else '-', styles['TableCellCenter']),
            ])
        task_detail_table = Table(task_detail_data, colWidths=[1.8*inch, 1.2*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.8*inch], repeatRows=1)
        task_detail_table.setStyle(ReportGenerationService._get_default_table_style(C, num_data_rows=len(data['tasks'])))
        story.append(task_detail_table)

        ReportGenerationService._build_pdf_doc(buffer, story, title="Performance de Tarefas", org_name=data['organization'].name, pagesize=landscape(A4))
        buffer.seek(0)
        return buffer, 'application/pdf'

    @staticmethod
    def _generate_task_performance_csv(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Relatório de Performance de Tarefas", f"Organização: {data['organization'].name}"])
        if data['date_from'] and data['date_to']: writer.writerow([f"Período Tarefas Criadas: {data['date_from'].strftime('%d/%m/%Y')} a {data['date_to'].strftime('%d/%m/%Y')}"])
        writer.writerow(["Filtros Aplicados:", f"Clientes: {data['filters_applied']['clients']}", f"Utilizadores: {data['filters_applied']['users']}", f"Categorias: {data['filters_applied']['categories']}", f"Status: {data['filters_applied']['statuses']}"])
        writer.writerow([])
        
        summary = data['summary_stats']
        writer.writerow(['Sumário Geral'])
        writer.writerow(['Total de Tarefas (filtros):', summary['total_tasks']])
        writer.writerow(['Tarefas Atrasadas:', summary['overdue_tasks_count']])
        writer.writerow(['Tempo Médio Conclusão (dias):', f"{summary['avg_completion_days']:.1f}" if summary['avg_completion_days'] is not None else "N/A"])
        writer.writerow([])

        writer.writerow(['ID Tarefa', 'Título', 'Cliente', 'Status', 'Prioridade', 'Prazo', 'Responsável', 'Categoria', 'Workflow', 'Criado em', 'Concluído em', 'Estimado (min)', 'Descrição'])
        for task in data['tasks']:
            writer.writerow([
                task.id, task.title, task.client.name, task.get_status_display(), task.get_priority_display(),
                task.deadline.strftime('%Y-%m-%d') if task.deadline else '',
                task.assigned_to.username if task.assigned_to else '',
                task.category.name if task.category else '',
                task.workflow.name if task.workflow else '',
                task.created_at.strftime('%Y-%m-%d %H:%M') if task.created_at else '',
                task.completed_at.strftime('%Y-%m-%d %H:%M') if task.completed_at else '',
                task.estimated_time_minutes or '',
                task.description or ''
            ])
        csv_content = buffer.getvalue()
        buffer.close()
        return io.BytesIO(csv_content.encode('utf-8-sig')), 'text/csv'

    @staticmethod
    def _generate_task_performance_xlsx(data: Dict) -> Tuple[io.BytesIO, str]:
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Sumário Performance"
        ws_summary['A1'] = f"Relatório Performance Tarefas - {data['organization'].name}"
        ws_summary.merge_cells('A1:G1'); ReportGenerationService._apply_xlsx_header_style(ws_summary['A1'], fill_color="1E40AF")
        if data['date_from'] and data['date_to']: ws_summary['A2'] = f"Período Criação: {data['date_from'].strftime('%d/%m/%Y')} a {data['date_to'].strftime('%d/%m/%Y')}"
        
        row_idx = 4
        summary = data['summary_stats']
        ws_summary.cell(row=row_idx, column=1, value="Total Tarefas (filtros):"); ReportGenerationService._apply_xlsx_cell_style(ws_summary.cell(row=row_idx,column=1), font_bold=True)
        ws_summary.cell(row=row_idx, column=2, value=summary['total_tasks']); row_idx+=1
        ws_summary.cell(row=row_idx, column=1, value="Tarefas Atrasadas:"); ReportGenerationService._apply_xlsx_cell_style(ws_summary.cell(row=row_idx,column=1), font_bold=True)
        ws_summary.cell(row=row_idx, column=2, value=summary['overdue_tasks_count']); row_idx+=1
        ws_summary.cell(row=row_idx, column=1, value="Tempo Médio Conclusão (dias):"); ReportGenerationService._apply_xlsx_cell_style(ws_summary.cell(row=row_idx,column=1), font_bold=True)
        ws_summary.cell(row=row_idx, column=2, value=summary['avg_completion_days'] if summary['avg_completion_days'] is not None else 'N/A').number_format = '0.0'
        row_idx+=2

        ReportGenerationService._apply_xlsx_header_style(ws_summary.cell(row=row_idx, column=1, value="Status"))
        ReportGenerationService._apply_xlsx_header_style(ws_summary.cell(row=row_idx, column=2, value="Quantidade"))
        row_idx+=1
        for s_dist in summary['status_distribution']:
            status_display = dict(Task.STATUS_CHOICES).get(s_dist['status'], s_dist['status'])
            ws_summary.cell(row=row_idx, column=1, value=status_display)
            ws_summary.cell(row=row_idx, column=2, value=s_dist['count'])
            ReportGenerationService._apply_xlsx_cell_style(ws_summary.cell(row=row_idx, column=1))
            ReportGenerationService._apply_xlsx_cell_style(ws_summary.cell(row=row_idx, column=2), alignment='right')
            row_idx+=1
        
        # Detailed Task List Sheet
        ws_details = wb.create_sheet("Lista Tarefas Detalhada")
        headers = ['ID', 'Título', 'Cliente', 'Status', 'Prioridade', 'Prazo', 'Responsável', 'Categoria', 'Workflow', 'Criado em', 'Concluído em', 'Estimado (min)', 'Descrição']
        for col_idx, header_text in enumerate(headers, 1):
            ReportGenerationService._apply_xlsx_header_style(ws_details.cell(row=1, column=col_idx, value=header_text))

        for r_idx, task in enumerate(data['tasks'], 2):
            row_values = [
                str(task.id), task.title, task.client.name, task.get_status_display(), task.get_priority_display(),
                task.deadline, task.assigned_to.username if task.assigned_to else None,
                task.category.name if task.category else None, task.workflow.name if task.workflow else None,
                task.created_at, task.completed_at, task.estimated_time_minutes, task.description
            ]
            for c_idx, val in enumerate(row_values,1):
                cell = ws_details.cell(row=r_idx, column=c_idx, value=val)
                if c_idx == 6: cell.number_format = 'YYYY-MM-DD' # Deadline
                elif c_idx == 10 or c_idx == 11: cell.number_format = 'YYYY-MM-DD HH:MM' # Created/Completed At
                ReportGenerationService._apply_xlsx_cell_style(cell, wrap_text=(c_idx in [2,13]))

        # Chart for Tasks by Status
        chart_sheet_status = wb.create_sheet("Gráfico Status Tarefas")
        chart_sheet_status.cell(row=1, column=1, value="Status")
        chart_sheet_status.cell(row=1, column=2, value="Quantidade")
        for i, item in enumerate(summary['status_distribution'], 2):
            status_display = dict(Task.STATUS_CHOICES).get(item['status'], item['status'])
            chart_sheet_status.cell(row=i, column=1, value=status_display)
            chart_sheet_status.cell(row=i, column=2, value=item['count'])
        
        pie_chart = OpenpyxlPieChart()
        labels_ref = Reference(chart_sheet_status, min_col=1, min_row=2, max_row=len(summary['status_distribution'])+1)
        data_ref = Reference(chart_sheet_status, min_col=2, min_row=1, max_row=len(summary['status_distribution'])+1)
        pie_chart.add_data(data_ref, titles_from_data=True)
        pie_chart.set_categories(labels_ref)
        pie_chart.title = "Distribuição de Tarefas por Status"
        chart_sheet_status.add_chart(pie_chart, "E2")

        for sheet_name_key in wb.sheetnames:
            current_sheet = wb[sheet_name_key]
            for i, column_cells in enumerate(current_sheet.columns):
                length = 0
                if column_cells:
                    str_values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
                    if str_values: length = max(len(val) for val in str_values)
                current_sheet.column_dimensions[get_column_letter(i + 1)].width = max(length + 3, 12)


        wb.save(buffer)
        buffer.seek(0)
        return buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # --- Report: Custom Report (Placeholder) ---
    @staticmethod
    def generate_custom_report(*args, **kwargs) -> Tuple[io.BytesIO, str]:
        buffer = io.BytesIO() 
        format_type = kwargs.get('format_type', 'pdf')

        if format_type == 'csv':
            s_buffer = io.StringIO()
            writer = csv.writer(s_buffer)
            writer.writerow(["Relatório Personalizado - Em Desenvolvimento"])
            s_buffer.seek(0)
            return io.BytesIO(s_buffer.read().encode('utf-8-sig')), 'text/csv'
        elif format_type == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = "Relatório Personalizado - Em Desenvolvimento"
            wb.save(buffer)
            buffer.seek(0)
            return buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else: # Default to PDF
            styles, C = ReportGenerationService._get_pdf_styles()
            story = [
                Paragraph("Relatório Personalizado", styles['ReportTitle']),
                Paragraph("Esta funcionalidade de relatório personalizado ainda está em desenvolvimento.", styles['BodyText']),
            ]
            ReportGenerationService._build_pdf_doc(buffer, story, title="Relatório Personalizado")
            buffer.seek(0)
            return buffer, 'application/pdf'